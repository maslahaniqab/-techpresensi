import os
import io
import csv
import math
import calendar
import smtplib
import time
import uuid
import json
from email.message import EmailMessage
from functools import wraps
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from urllib.parse import quote

import openpyxl
from dateutil import parser as date_parser
from python_calamine import CalamineWorkbook

WIB = ZoneInfo("Asia/Jakarta")


def now_wib():
    return datetime.now(WIB).replace(tzinfo=None)


def today_wib():
    return now_wib().date()

from io import BytesIO

from flask import Flask, render_template, redirect, url_for, request, flash, jsonify, Response, send_file
from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    current_user,
)
from werkzeug.utils import secure_filename
from xhtml2pdf import pisa

from models import (
    db, User, Employee, Attendance, Settings, Payroll, PengajuanIzin,
    LaporanPekerjaan, PengajuanLembur, IklanMarketplace, ProdukIklan,
    PengeluaranOperasional, ItemLabaRugi, PenjualanMarketplace, Produk,
    IklanMeta, HariLibur, PesananMarketplace, PendapatanPesanan,
    BahanBaku, BahanBakuKebutuhan, BahanBakuTransaksi, ProdukSpekUkuran,
    Vendor, Gudang, AkunPembayaran, PurchaseOrder, PurchaseOrderItemProduk,
    PurchaseOrderBahanPakai, PurchaseOrderPembayaran,
)

HARI_NAMA = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]

# Kolom ukuran di ProdukSpekUkuran (lingkar_dada s/d pergelangan) dipakai sbg "slot"
# generik yg label & maknanya beda-beda tergantung kategori produknya. "" = kategori
# kosong/belum diisi (baris lama sblm fitur kategori ini ada) -> fallback ke label
# generik yg sama seperti sebelumnya, jadi data lama tetap kebaca normal.
KATEGORI_SPEK_FIELDS = {
    "": [
        ("lingkar_dada", "Lingkar Dada (LD)"), ("panjang_atas", "Panjang Atas (PA)"),
        ("lingkar_pinggang", "Lingkar Pinggang"), ("ld_lengan", "LD Lengan"), ("pergelangan", "Pergelangan"),
    ],
    "Cadar": [
        ("lingkar_dada", "Lingkar Kepala"), ("panjang_atas", "Lingkar Wajah"), ("lingkar_pinggang", "Panjang Cadar"),
    ],
    "Khimar": [
        ("lingkar_dada", "Lingkar Kepala"), ("panjang_atas", "Panjang Khimar"), ("lingkar_pinggang", "Lebar Bawah"),
        ("ld_lengan", "Panjang Depan"), ("pergelangan", "Panjang Belakang"),
    ],
    "Pashmina": [
        ("lingkar_dada", "Panjang"), ("panjang_atas", "Lebar"),
    ],
    "Handshock": [
        ("lingkar_dada", "Panjang"), ("panjang_atas", "Lingkar Pergelangan"),
    ],
    "Celamis": [
        ("lingkar_dada", "Lingkar Pinggang"), ("panjang_atas", "Lingkar Pinggul"), ("lingkar_pinggang", "Panjang Celana"),
    ],
    "Abaya": [
        ("lingkar_dada", "Lingkar Dada"), ("lingkar_pinggang", "Lingkar Pinggang"), ("panjang_atas", "Panjang Baju"),
        ("ld_lengan", "Panjang Lengan"), ("pergelangan", "Lingkar Lengan"),
    ],
    "Oneset": [
        ("panjang_atas", "Panjang Atas"), ("pergelangan", "Panjang Bawah"),
        ("lingkar_dada", "Lingkar Dada"), ("lingkar_pinggang", "Lingkar Pinggang"),
    ],
    "Gamis": [
        ("lingkar_dada", "Lingkar Dada"), ("lingkar_pinggang", "Lingkar Pinggang"), ("panjang_atas", "Panjang Baju"),
        ("ld_lengan", "Panjang Lengan"), ("pergelangan", "Lingkar Lengan"),
    ],
}

HARI_LIBUR_2026 = [
    ("2026-01-01", "Tahun Baru Masehi"),
    ("2026-01-16", "Isra Mikraj Nabi Muhammad SAW"),
    ("2026-02-16", "Cuti Bersama Tahun Baru Imlek"),
    ("2026-02-17", "Tahun Baru Imlek 2577 Kongzili"),
    ("2026-03-18", "Cuti Bersama Hari Suci Nyepi"),
    ("2026-03-19", "Hari Suci Nyepi (Tahun Baru Saka 1948)"),
    ("2026-03-20", "Cuti Bersama Idulfitri"),
    ("2026-03-21", "Idulfitri 1447 H"),
    ("2026-03-22", "Idulfitri 1447 H"),
    ("2026-03-23", "Cuti Bersama Idulfitri"),
    ("2026-03-24", "Cuti Bersama Idulfitri"),
    ("2026-04-03", "Wafat Yesus Kristus"),
    ("2026-04-05", "Kebangkitan Yesus Kristus (Paskah)"),
    ("2026-05-01", "Hari Buruh Internasional"),
    ("2026-05-14", "Kenaikan Yesus Kristus"),
    ("2026-05-15", "Cuti Bersama Kenaikan Yesus Kristus"),
    ("2026-05-27", "Iduladha 1447 H"),
    ("2026-05-28", "Cuti Bersama Iduladha"),
    ("2026-05-31", "Hari Raya Waisak 2570 BE"),
    ("2026-06-01", "Hari Lahir Pancasila"),
    ("2026-06-16", "Tahun Baru Islam 1448 H (1 Muharam)"),
    ("2026-08-17", "Proklamasi Kemerdekaan RI"),
    ("2026-08-25", "Maulid Nabi Muhammad SAW"),
    ("2026-12-24", "Cuti Bersama Hari Raya Natal"),
    ("2026-12-25", "Kelahiran Yesus Kristus (Natal)"),
]

BULAN_NAMA = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]

STATUS_PEGAWAI_BULANAN = ("Karyawan Tetap", "Probation")

JABATAN_LIST = [
    "Admin Ops",
    "Admin Sales & Marketplace",
    "Admin Affiliator",
    "Admin Gudang",
    "Host Live",
    "Manager",
    "Human Resource",
    "Supervisor",
]

# Skema BPJS Kesehatan sesuai Perpres 82/2018: total iuran 5% dari gaji (gaji pokok +
# tunjangan tetap), dibagi 4% ditanggung perusahaan dan 1% dipotong dari gaji karyawan,
# dengan batas atas gaji yang dihitung Rp 12.000.000/bulan.
BATAS_GAJI_BPJS_KESEHATAN = 12_000_000
PERSEN_BPJS_KESEHATAN_KARYAWAN = 0.01
PERSEN_BPJS_KESEHATAN_PERUSAHAAN = 0.04

KATEGORI_PENGELUARAN_RUTIN = [
    "Internet Indihome",
    "Internet Biznet",
    "Listrik",
    "Air",
    "PAM",
    "Telekomunikasi",
    "Bensin, Parkir, Tol Kendaraan",
    "Perlengkapan Kantor",
    "Sewa Gedung",
    "Penyusutan Peralatan",
    "Biaya Afiliasi Tokopedia Softsell",
]

KELOMPOK_LABA_RUGI = [
    "Pendapatan",
    "Beban Pokok Penjualan",
    "Beban Operasional",
    "Pendapatan Non Operasional",
    "Beban Non Operasional",
]

PRESET_LABA_RUGI = {
    "Pendapatan": [
        "Penjualan",
        "Diskon Penjualan",
        "Beban Diskon Shopee",
        "Beban Diskon Tokopedia",
        "Beban Diskon TikTok Shop",
        "Beban Diskon Lazada",
    ],
    "Beban Pokok Penjualan": [
        "Beban Pokok Penjualan (HPP)",
        "Shopee - Beban Kelebihan/Kekurangan Ongkir",
        "Shopee - Biaya Layanan dan Tambahan Platform",
        "Shopee - Biaya Afiliasi",
        "TikTok Shop - Beban Kelebihan/Kekurangan Ongkir",
        "TikTok Shop - Biaya Layanan dan Tambahan Platform",
        "TikTok Shop - Biaya Afiliasi",
        "Tokopedia - Biaya Layanan dan Tambahan Platform",
        "Lazada - Biaya Layanan dan Tambahan Platform",
        "Beban Ongkir Manual",
    ],
    "Beban Operasional": [
        "Iklan Shopee (Tambahan Manual)",
        "Iklan Tokopedia (Tambahan Manual)",
        "Iklan TikTok Shop (Tambahan Manual)",
        "Iklan Lazada (Tambahan Manual)",
        "Iklan Blibli (Tambahan Manual)",
        "Iklan Meta (Tambahan Manual)",
        "Pajak Iklan (Tambahan Manual)",
        "Beban Gaji, Upah dan Honorer (Tambahan Manual)",
    ],
    "Pendapatan Non Operasional": [
        "Pendapatan Bunga Bank",
        "Pendapatan Lain-lain",
    ],
    "Beban Non Operasional": [
        "Beban Administrasi Bank",
        "Beban Lain-lain",
    ],
}

MARKETPLACE_LIST = ["Shopee", "Tokopedia", "TikTok Shop", "Lazada", "Blibli"]

KOLOM_TARGET_IKLAN = [
    ("tanggal", "Tanggal", True, ["tanggal", "date", "tgl", "periode", "reporting starts", "reporting ends", "per hari"]),
    ("biaya", "Biaya Iklan", True, [
        "biaya", "cost", "spend", "spent", "pengeluaran", "biaya iklan", "amount spent", "belanja iklan",
        "dibelanjakan",
    ]),
    ("impresi", "Impresi/Dilihat", False, ["impresi", "impression", "dilihat", "views", "tayangan", "reach", "jangkauan"]),
    ("klik", "Klik", False, ["klik", "click", "jumlah klik"]),
    ("pesanan", "Pesanan/Konversi", False, [
        "pesanan", "konversi", "conversion", "order", "dibeli", "produk terjual", "checkout", "terjual",
        "purchase", "result", "hasil",
    ]),
    ("omzet", "Omzet Penjualan", False, [
        "omzet", "omset", "penjualan", "revenue", "gmv", "nilai penjualan", "sales", "conversion value",
        "purchase value", "purchases conversion value", "nilai konversi", "nilai pembelian", "penghasilan bruto",
    ]),
]

KOLOM_TARGET_META = KOLOM_TARGET_IKLAN + [
    ("pajak", "Pajak Iklan", False, ["pajak", "tax", "ppn"]),
]

TARIF_PAJAK_META = 0.11  # PPN 11% -- "Biaya" Iklan Meta sudah termasuk pajak ini


def hitung_pajak_meta(biaya):
    return round((biaya or 0) * TARIF_PAJAK_META)


KOLOM_TARGET_PRODUK = [
    ("nama_produk", "Nama Produk", True, ["nama produk", "produk", "product", "nama barang", "item name", "item"]),
    ("tanggal", "Tanggal", True, ["tanggal", "date", "tgl", "periode"]),
    ("biaya", "Biaya Iklan", True, ["biaya", "cost", "spend", "pengeluaran", "biaya iklan"]),
    ("impresi", "Impresi/Dilihat", False, ["impresi", "impression", "dilihat", "views", "tayangan", "reach"]),
    ("klik", "Klik", False, ["klik", "click", "jumlah klik"]),
    ("pesanan", "Pesanan/Konversi", False, ["pesanan", "konversi", "conversion", "order", "dibeli", "produk terjual", "checkout", "terjual"]),
    ("omzet", "Omzet Penjualan", False, ["omzet", "omset", "penjualan", "revenue", "gmv", "nilai penjualan", "sales"]),
]

KOLOM_TARGET_PENJUALAN = [
    ("tanggal", "Tanggal (Waktu Pesanan Selesai)", True, ["waktu pesanan selesai", "tanggal", "date", "tgl", "periode"]),
    ("no_pesanan", "No. Pesanan", False, ["no. pesanan", "no pesanan", "nomor pesanan", "order id", "id pesanan"]),
    ("total_penjualan", "Total Penjualan (Total Pembayaran)", True, ["total pembayaran", "total penjualan", "penjualan", "omzet", "sales"]),
    ("total_diskon", "Total Diskon", False, ["total diskon"]),
]

# Header persis yang sudah pernah dikonfirmasi cocok untuk laporan Pesanan Selesai dari
# marketplace tertentu -- kalau SEMUA kolom wajib pada salah satu profil ini ketemu persis
# di file yang diupload, langkah "Cocokkan Kolom" manual dilewati dan file langsung diproses.
PROFIL_HEADER_PENJUALAN = [
    {
        "nama": "Shopee (header Bahasa Indonesia)",
        "kolom": {
            "tanggal": "Waktu Pesanan Selesai",
            "no_pesanan": "No. Pesanan",
            "total_penjualan": "Total Pembayaran",
            "total_diskon": "Total Diskon",
        },
    },
    {
        "nama": "Format deskriptif Bahasa Inggris (mis. TikTok Shop/Tokopedia)",
        "kolom": {
            "tanggal": "Order paid time.",
            "no_pesanan": "Platform unique order ID.",
            "total_penjualan": "Order total amount paid by the buyer.",
            "total_diskon": None,
        },
    },
    {
        "nama": "Lazada (nama kolom API)",
        "kolom": {
            "tanggal": "deliveredDate",
            "no_pesanan": "orderNumber",
            "total_penjualan": "paidPrice",
            "total_diskon": None,  # paidPrice sudah bersih setelah diskon, jangan dipotong lagi
        },
    },
]


def cocokkan_profil_header_penjualan(headers):
    """Cari profil header yang cocok persis dengan file yang diupload. Mengembalikan
    (nama_profil, mapping) jika semua kolom WAJIB pada satu profil ditemukan persis
    di headers, atau (None, None) jika tidak ada yang cocok sepenuhnya."""
    headers_lower = {}
    for i, h in enumerate(headers):
        kunci = str(h).strip().lower()
        if kunci not in headers_lower:
            headers_lower[kunci] = i

    for profil in PROFIL_HEADER_PENJUALAN:
        mapping = {}
        lengkap = True
        for key, _label, wajib, _kk in KOLOM_TARGET_PENJUALAN:
            nama_kolom = profil["kolom"].get(key)
            idx = headers_lower.get(nama_kolom.strip().lower()) if nama_kolom else None
            mapping[key] = idx
            if wajib and idx is None:
                lengkap = False
                break
        if lengkap:
            return profil["nama"], mapping
    return None, None


BATAS_ANGKA_WAJAR_PENJUALAN = 1_000_000_000_000  # Rp 1 triliun per baris -- di atas ini hampir pasti salah kolom (mis. ID pesanan ikut terbaca sebagai angka)

_KATA_TANGGAL_PRIORITAS = ["selesai", "complete", "completed", "deliver", "paid", "payment", "finish"]
_KATA_TOTAL_PRIORITAS = ["paid", "payment", "amount", "pembayaran", "penjualan", "price", "total"]
_KATA_TOTAL_HINDARI = ["id", "code", "number", "sku", "qty", "quantity", "weight", "berat", "phone", "postcode", "zip"]


def deteksi_otomatis_kolom_penjualan(headers, rows):
    """Deteksi kolom Tanggal/No. Pesanan/Total Penjualan/Total Diskon dari ISI datanya
    (bukan cuma nama header), dipakai saat file tidak cocok dengan profil header manapun
    yang sudah dikenal (mis. marketplace/format baru). Mengembalikan dict mapping
    (index kolom atau None kalau tidak ketemu)."""
    sampel = rows[:15]
    headers_lower = [str(h).strip().lower() for h in headers]
    n = len(headers)

    def tingkat_valid_tanggal(idx):
        ok = sum(1 for r in sampel if idx < len(r) and parse_tanggal_iklan(r[idx]))
        return ok / max(1, len(sampel))

    def nilai_numerik(idx):
        hasil = []
        for r in sampel:
            if idx < len(r) and r[idx] not in (None, ""):
                v = parse_angka_iklan(r[idx])
                if v:
                    hasil.append(v)
        return hasil

    idx_tanggal = None
    skor_tanggal_terbaik = 0
    for i in range(n):
        rate = tingkat_valid_tanggal(i)
        if rate < 0.7:
            continue
        bonus = 1 if any(k in headers_lower[i] for k in _KATA_TANGGAL_PRIORITAS) else 0
        skor = rate + bonus
        if skor > skor_tanggal_terbaik:
            skor_tanggal_terbaik, idx_tanggal = skor, i

    idx_total = None
    skor_total_terbaik = -1
    for i in range(n):
        if i == idx_tanggal:
            continue
        h = headers_lower[i]
        if any(k in h for k in _KATA_TOTAL_HINDARI):
            continue
        nilai = nilai_numerik(i)
        if len(nilai) < len(sampel) * 0.5:
            continue
        median = sorted(abs(v) for v in nilai)[len(nilai) // 2]
        if median >= BATAS_ANGKA_WAJAR_PENJUALAN or median < 500:
            continue
        skor = 2 if any(k in h for k in _KATA_TOTAL_PRIORITAS) else 1
        if skor > skor_total_terbaik:
            skor_total_terbaik, idx_total = skor, i

    idx_no_pesanan = None
    for i, h in enumerate(headers_lower):
        if "order" in h and ("number" in h or "id" in h):
            idx_no_pesanan = i
            break
        if "no. pesanan" in h or "no pesanan" in h or "nomor pesanan" in h:
            idx_no_pesanan = i
            break

    idx_diskon = None
    for i, h in enumerate(headers_lower):
        if i in (idx_tanggal, idx_total):
            continue
        if "discount" in h or "diskon" in h:
            idx_diskon = i
            break

    return {
        "tanggal": idx_tanggal,
        "no_pesanan": idx_no_pesanan,
        "total_penjualan": idx_total,
        "total_diskon": idx_diskon,
    }


_KATA_KUNCI_HEADER = set()
for _key, _label, _wajib, _kk in KOLOM_TARGET_IKLAN + KOLOM_TARGET_PRODUK + KOLOM_TARGET_PENJUALAN:
    _KATA_KUNCI_HEADER.update(_kk)


def _parse_csv_delimiter_terbaik(text):
    """Coba beberapa delimiter umum, pakai yang menghasilkan baris paling lebar
    (laporan marketplace kadang pakai ';' bukan ',', dan Sniffer gampang salah tebak
    kalau baris-baris awal file cuma metadata satu kolom)."""
    kandidat = []
    for delim in (",", ";", "\t"):
        try:
            reader = csv.reader(io.StringIO(text), delimiter=delim)
            baris = [row for row in reader if any(str(c).strip() for c in row)]
        except csv.Error:
            continue
        lebar_maks = max((len(r) for r in baris), default=0)
        kandidat.append((lebar_maks, baris))
    if not kandidat:
        return []
    kandidat.sort(key=lambda x: x[0], reverse=True)
    return kandidat[0][1]


def _cari_baris_header(semua):
    """Sebagian laporan marketplace (mis. Shopee) diawali beberapa baris metadata
    (judul laporan, nama toko, periode, dll) sebelum baris header kolom sebenarnya.
    Cari baris paling mirip header: kolom terisi terbanyak + paling banyak cocok
    kata kunci nama kolom yang kita kenal."""
    skor_terbaik, idx_terbaik = -1, 0
    for idx, row in enumerate(semua[:50]):
        sel = [str(c).strip() for c in row]
        non_kosong = sum(1 for c in sel if c)
        if non_kosong < 2:
            continue
        cocok = sum(1 for c in sel if any(kk in c.lower() for kk in _KATA_KUNCI_HEADER))
        skor = non_kosong + cocok * 5
        if skor > skor_terbaik:
            skor_terbaik, idx_terbaik = skor, idx
    return idx_terbaik


def deteksi_otomatis_kolom_iklan(headers, rows, kolom_target):
    """Deteksi kolom Tanggal/Biaya/dst dari nama header + validasi isi datanya, tanpa
    perlu form cocokkan kolom manual. Kolom WAJIB (tanggal, biaya) divalidasi isinya
    (harus mayoritas berupa tanggal/angka valid) sebelum dianggap cocok; kalau kolom
    yang cocok nama tapi isinya tidak valid, dicoba kolom lain yang juga cocok nama,
    lalu (khusus tanggal) fallback ke pemindaian isi tanpa syarat nama kolom."""
    sampel = rows[:20] or rows
    headers_lower = [str(h).strip().lower() for h in headers]
    n = len(headers)

    def rate_tanggal(idx):
        ok = sum(1 for r in sampel if idx < len(r) and parse_tanggal_iklan(r[idx]))
        return ok / max(1, len(sampel))

    def terlihat_numerik(value):
        """Cek apakah nilai TERLIHAT seperti angka (termasuk nol yang sah), tanpa lewat
        parse_angka_iklan yang juga mengembalikan 0 untuk teks yang gagal diparse --
        supaya kolom yang isinya kebetulan banyak angka 0 (mis. belum ada biaya iklan
        hari itu) tidak salah dianggap "bukan kolom angka"."""
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return True
        s = str(value).strip()
        if not s:
            return False
        s_bersih = s.replace("Rp", "").replace("rp", "").replace("%", "").strip()
        return any(ch.isdigit() for ch in s_bersih)

    def rate_numerik(idx):
        total = ok = 0
        for r in sampel:
            if idx < len(r) and str(r[idx]).strip() != "":
                total += 1
                if terlihat_numerik(r[idx]):
                    ok += 1
        return (ok / total) if total else 0

    def rate_non_kosong(idx):
        total = ok = 0
        for r in sampel:
            if idx < len(r):
                total += 1
                if str(r[idx]).strip():
                    ok += 1
        return (ok / total) if total else 0

    hasil = {}
    idx_terpakai = set()

    for key, _label, wajib, kata_kunci in kolom_target:
        if key == "tanggal":
            cek, ambang = rate_tanggal, 0.6
        elif key == "nama_produk":
            cek, ambang = rate_non_kosong, 0.5
        else:
            cek, ambang = rate_numerik, 0.5
        idx_pilihan = None
        for idx, h in enumerate(headers_lower):
            if idx in idx_terpakai:
                continue
            if any(kk in h for kk in kata_kunci) and cek(idx) >= ambang:
                idx_pilihan = idx
                break
        if idx_pilihan is None and key == "tanggal":
            skor_terbaik = 0
            for idx in range(n):
                if idx in idx_terpakai:
                    continue
                rate = rate_tanggal(idx)
                if rate >= 0.7 and rate > skor_terbaik:
                    skor_terbaik, idx_pilihan = rate, idx
        hasil[key] = idx_pilihan
        if idx_pilihan is not None:
            idx_terpakai.add(idx_pilihan)

    return hasil


def deteksi_kolom_produk(headers):
    """Deteksi kolom Nama Produk/Variasi/HPP/Harga Jual dari nama header (mis. laporan
    'Online Products' Shopee), tanpa form cocokkan kolom manual."""
    headers_lower = [str(h).strip().lower() for h in headers]
    idx_terpakai = set()

    def cari(kata_kunci):
        for idx, h in enumerate(headers_lower):
            if idx in idx_terpakai:
                continue
            if any(kk in h for kk in kata_kunci):
                idx_terpakai.add(idx)
                return idx
        return None

    return {
        "nama_produk": cari(["nama produk", "product name", "nama barang"]),
        "variasi": cari(["variasi", "variant", "varian"]),
        "hpp": cari(["hpp", "harga pokok", "cost"]),
        "harga_jual": cari(["harga jual", "selling price", "price"]),
    }


KOLOM_LAIN_INCOME_SHOPEE = [
    "Premi", "Biaya Transaksi", "Biaya Kampanye", "Biaya Komisi AMS",
    "Biaya Proses Pesanan", "FBS Fee", "Biaya Isi Saldo Otomatis (dari Penghasilan)", "PPh 22",
]


_MAX_KOLOM_LAPORAN_SHOPEE = 80  # cukup lebar utk laporan asli (~50-53 kolom); mencegah baca kolom "hantu"


def _cek_header_marketplace(sel):
    """sel = list header (sudah lower+strip). Kembalikan (marketplace, tipe) atau (None, None).
    tipe bernilai 'order'/'income'."""
    if "no. pesanan" in sel and "status pesanan" in sel:
        return "Shopee", "order"
    if "no. pesanan" in sel and ("total penghasilan" in sel or "lihat berdasarkan" in sel):
        # "total penghasilan" = varian file Income terbaru (1 baris/pesanan, sheet "Income").
        # "lihat berdasarkan" = varian lama yg baris Order & Sku dicampur dalam 1 sheet --
        # dicek belakangan karena kolom itu JUGA muncul di sheet lain yg bukan Income asli
        # (mis. "Order Processing Fee"), jadi jangan sampai sheet itu kepilih duluan.
        return "Shopee", "income"
    if "order id" in sel and "product name" in sel and "seller sku" in sel:
        return "TikTok Shop", "order"
    if "id pesanan/penyesuaian" in sel and "jenis transaksi" in sel:
        return "TikTok Shop", "income"
    return None, None


def baca_laporan_marketplace(file_storage):
    """Baca file xlsx/csv lalu deteksi ini Laporan Pesanan (Order) atau Laporan
    Pendapatan (Income) marketplace mana (Shopee/TikTok Shop/dst) dari nama kolomnya --
    mengembalikan (marketplace, tipe, headers, baris_data, error).

    Dibatasi _MAX_KOLOM_LAPORAN_SHOPEE kolom saat baca xlsx supaya tidak ikut membaca
    kolom "hantu" (sisa format/formula lama) yang bisa melebar sampai ratusan/ribuan
    kolom kosong pada file laporan asli -- kalau tidak dibatasi, baca 1 file bisa
    makan waktu sangat lama dan bikin request timeout di server.
    """
    filename = file_storage.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    raw = file_storage.read()
    pesan_gagal = (
        "File ini bukan Laporan Pesanan (Order) atau Laporan Pendapatan (Income) yang dikenali "
        "(baru mendukung Shopee & TikTok Shop). Pastikan file yang diupload adalah hasil export asli "
        "dari Seller Center/Partner Center marketplace terkait."
    )

    if ext == "csv":
        text = raw.decode("utf-8-sig", errors="ignore")
        baris = _parse_csv_delimiter_terbaik(text)
        for idx_baris in range(min(6, len(baris))):
            sel = [str(c).strip().lower() for c in baris[idx_baris][:_MAX_KOLOM_LAPORAN_SHOPEE]]
            marketplace, tipe = _cek_header_marketplace(sel)
            if tipe:
                return marketplace, tipe, baris[idx_baris], baris[idx_baris + 1:], None
        return None, None, None, None, pesan_gagal

    if ext not in ("xlsx", "xlsm"):
        return None, None, None, None, "Format file tidak didukung. Gunakan file CSV atau XLSX."

    # Catatan: sebelumnya pakai openpyxl, tapi laporan marketplace asli sering punya "kolom
    # hantu" (dimensi sheet ke ratusan/ribuan kolom kosong, sisa format/formula lama) yang
    # bikin openpyxl SANGAT lambat (puluhan detik, bisa timeout) -- read_only=True openpyxl
    # pun malah salah baca (cuma dapat 1 baris) untuk file ini. python-calamine tidak kena
    # masalah itu (baca file yang sama dalam <2 detik, kolom yang dilaporkan pun akurat).
    try:
        wb = CalamineWorkbook.from_filelike(io.BytesIO(raw))
    except Exception:
        return None, None, None, None, "File Excel tidak bisa dibaca. Pastikan file tidak rusak."

    marketplace_ditemukan = tipe_ditemukan = headers = baris_data = None
    for nama_sheet in wb.sheet_names:
        baris_sheet_penuh = wb.get_sheet_by_name(nama_sheet).to_python()
        idx_header = None
        for i, row in enumerate(baris_sheet_penuh[:6]):
            sel = [str(c).strip().lower() if c is not None else "" for c in row[:_MAX_KOLOM_LAPORAN_SHOPEE]]
            marketplace, tipe = _cek_header_marketplace(sel)
            if tipe:
                idx_header, marketplace_ditemukan, tipe_ditemukan = i, marketplace, tipe
                headers = ["" if c is None else c for c in row]
                break
        if idx_header is not None:
            baris_data = [
                ["" if c is None else c for c in row]
                for row in baris_sheet_penuh[idx_header + 1:]
                if any(cell not in (None, "") for cell in row)
            ]
            break

    if tipe_ditemukan is None:
        return None, None, None, None, pesan_gagal
    return marketplace_ditemukan, tipe_ditemukan, headers, baris_data, None


def parse_order_shopee(headers, rows_data):
    idx = {str(h).strip().lower(): i for i, h in enumerate(headers)}

    def ambil(row, nama):
        i = idx.get(nama.lower())
        return row[i] if i is not None and i < len(row) else None

    def ambil_pertama(row, *nama_list):
        """Beberapa varian export Order Shopee pakai nama kolom berbeda untuk konsep yang
        sama (mis. 'Subtotal Pesanan' di satu versi, 'Dibayar Pembeli' di versi lain) --
        coba tiap nama sampai ketemu kolom yang ada."""
        for nama in nama_list:
            if nama.lower() in idx:
                return ambil(row, nama)
        return None

    hasil = []
    for row in rows_data:
        no_pesanan = str(ambil(row, "No. Pesanan") or "").strip()
        if not no_pesanan:
            continue
        tanggal = parse_tanggal_iklan(ambil(row, "Waktu Pesanan Dibuat"))
        if not tanggal:
            continue
        nama_produk = str(ambil(row, "Nama Produk") or "").strip()
        variasi = str(ambil(row, "Nama Variasi") or "").strip()
        nama_final = f"{nama_produk} - {variasi}" if variasi and variasi != "-" else nama_produk
        hasil.append({
            "no_pesanan": no_pesanan,
            "tanggal_pesanan": tanggal,
            "status_pesanan": str(ambil(row, "Status Pesanan") or "").strip()[:32],
            "nama_produk": nama_final.strip()[:256],
            "sku": str(ambil(row, "Nomor Referensi SKU") or "").strip()[:128],
            "jumlah": round(parse_angka_iklan(ambil(row, "Jumlah"))),
            "subtotal": round(parse_angka_iklan(ambil_pertama(row, "Subtotal Pesanan", "Dibayar Pembeli"))),
            "subtotal_kotor": round(
                parse_angka_iklan(ambil(row, "Harga Awal")) * parse_angka_iklan(ambil(row, "Jumlah"))
            ),
        })
    return hasil


def parse_income_shopee(headers, rows_data):
    idx = {str(h).strip().lower(): i for i, h in enumerate(headers)}

    def ambil(row, nama):
        i = idx.get(nama.lower())
        return row[i] if i is not None and i < len(row) else None

    # Varian file Income yang punya kolom "Lihat berdasarkan" mencampur baris ringkasan
    # per Order dengan baris rincian per Sku dalam 1 sheet, jadi perlu difilter cuma
    # ambil yang "Order". Varian lain (sheet "Income" versi lebih baru) sudah 1 baris
    # = 1 pesanan tanpa kolom ini sama sekali -- jangan difilter, nanti malah baris
    # kebuang semua & Total Penghasilan keitung 0 padahal ada datanya.
    ada_kolom_lihat_berdasarkan = "lihat berdasarkan" in idx

    hasil = []
    for row in rows_data:
        if ada_kolom_lihat_berdasarkan and str(ambil(row, "Lihat berdasarkan") or "").strip().lower() != "order":
            continue  # lewati baris rincian 'Sku', pakai baris ringkasan 'Order' saja
        no_pesanan = str(ambil(row, "No. Pesanan") or "").strip()
        if not no_pesanan:
            continue
        biaya_lainnya = sum(abs(parse_angka_iklan(ambil(row, k))) for k in KOLOM_LAIN_INCOME_SHOPEE)
        hasil.append({
            "no_pesanan": no_pesanan,
            "tanggal_dana_dilepas": parse_tanggal_iklan(ambil(row, "Tanggal Dana Dilepaskan")),
            "total_penghasilan": round(parse_angka_iklan(ambil(row, "Total Penghasilan"))),
            "biaya_admin": round(abs(parse_angka_iklan(ambil(row, "Biaya Administrasi (termasuk PPN 11%)")))),
            "biaya_layanan": round(abs(parse_angka_iklan(ambil(row, "Biaya Layanan")))),
            "biaya_lainnya": round(biaya_lainnya),
        })
    return hasil


def parse_order_tiktok(headers, rows_data):
    idx = {str(h).strip().lower(): i for i, h in enumerate(headers)}

    def ambil(row, nama):
        i = idx.get(nama.lower())
        return row[i] if i is not None and i < len(row) else None

    hasil = []
    for row in rows_data:
        order_id = str(ambil(row, "Order ID") or "").strip()
        if not order_id or not order_id.isdigit():
            continue  # lewati baris deskripsi kolom yang kadang ikut terbawa (bukan data asli)
        tanggal = parse_tanggal_iklan(ambil(row, "Created Time"))
        if not tanggal:
            continue
        nama_produk = str(ambil(row, "Product Name") or "").strip()
        variasi = str(ambil(row, "Variation") or "").strip()
        nama_final = f"{nama_produk} - {variasi}" if variasi and variasi != "-" else nama_produk
        hasil.append({
            "no_pesanan": order_id,
            "tanggal_pesanan": tanggal,
            "status_pesanan": str(ambil(row, "Order Status") or "").strip()[:32],
            "nama_produk": nama_final.strip()[:256],
            "sku": str(ambil(row, "Seller SKU") or "").strip()[:128],
            "jumlah": round(parse_angka_iklan(ambil(row, "Quantity"))),
            "subtotal": round(parse_angka_iklan(ambil(row, "SKU Subtotal After Discount"))),
            "subtotal_kotor": round(parse_angka_iklan(ambil(row, "SKU Subtotal Before Discount"))),
        })
    return hasil


def parse_income_tiktok(headers, rows_data):
    idx = {str(h).strip().lower(): i for i, h in enumerate(headers)}

    def ambil(row, nama):
        i = idx.get(nama.lower())
        return row[i] if i is not None and i < len(row) else None

    hasil = []
    for row in rows_data:
        jenis = str(ambil(row, "Jenis transaksi") or "").strip().lower()
        if jenis and jenis != "pesanan":
            continue  # lewati baris penyesuaian/adjustment, pakai baris pesanan asli saja
        order_id = str(ambil(row, "ID Pesanan/Penyesuaian") or "").strip()
        if not order_id:
            continue
        biaya_admin = abs(parse_angka_iklan(ambil(row, "Biaya komisi platform")))
        total_biaya = abs(parse_angka_iklan(ambil(row, "Total Biaya")))
        hasil.append({
            "no_pesanan": order_id,
            "tanggal_dana_dilepas": parse_tanggal_iklan(ambil(row, "Waktu pembayaran pesanan")),
            "total_penghasilan": round(parse_angka_iklan(ambil(row, "Jumlah penyelesaian pembayaran"))),
            "biaya_admin": round(biaya_admin),
            "biaya_layanan": 0,
            "biaya_lainnya": round(max(total_biaya - biaya_admin, 0)),
        })
    return hasil


PARSER_ORDER_MARKETPLACE = {
    "Shopee": parse_order_shopee,
    "TikTok Shop": parse_order_tiktok,
    # Order Manual pakai format kolom yang sama persis dengan export Order Shopee
    # (No. Pesanan, Waktu Pesanan Dibuat, Status Pesanan, dst) -- untuk pesanan
    # langsung/di luar marketplace (WA, COD, dsb) yang diinput manual oleh admin
    # lewat template, jadi parsernya bisa dipakai bareng.
    "Manual": parse_order_shopee,
}
PARSER_INCOME_MARKETPLACE = {
    "Shopee": parse_income_shopee,
    "TikTok Shop": parse_income_tiktok,
}


STATUS_BATAL_MARKETPLACE = ("Batal", "Dibatalkan", "Cancelled", "Cancel")


def hitung_profit_agregat(bulan=None):
    """Hitung profit per produk untuk periode tertentu ('YYYY-MM' atau None = semua),
    digabung dari SEMUA marketplace yang datanya sudah diupload. Profit = Total
    Penghasilan (Income, dialokasikan proporsional ke tiap baris produk dalam pesanan
    yang sama berdasarkan share Subtotal) dikurangi HPP x Qty (dari menu Data Produk &
    Harga Jual). Pesanan berstatus batal tidak dihitung."""
    tahun_f = bulan_f = None
    q = PesananMarketplace.query.filter(PesananMarketplace.status_pesanan.notin_(STATUS_BATAL_MARKETPLACE))
    if bulan:
        try:
            tahun_f, bulan_f = (int(x) for x in bulan.split("-"))
            q = q.filter(
                db.extract("year", PesananMarketplace.tanggal_pesanan) == tahun_f,
                db.extract("month", PesananMarketplace.tanggal_pesanan) == bulan_f,
            )
        except ValueError:
            tahun_f = bulan_f = None
    item_list = q.all()

    kosong = {
        "produk": [], "tren": [],
        "ringkasan": {
            "total_omzet": 0, "total_income": 0, "total_hpp": 0, "total_profit": 0, "margin": 0,
            "jumlah_produk": 0, "jumlah_produk_ada_hpp": 0, "jumlah_pesanan": 0,
            "total_qty": 0, "total_potongan": 0, "total_biaya_admin": 0, "total_biaya_layanan": 0,
            "total_biaya_lainnya": 0, "total_biaya_iklan": 0, "total_profit_real": 0, "margin_real": 0,
        },
    }
    if not item_list:
        return kosong

    kunci_pesanan_set = {(it.marketplace, it.no_pesanan) for it in item_list}
    marketplace_set = {it.marketplace for it in item_list}
    income_map = {
        (p.marketplace, p.no_pesanan): p
        for p in PendapatanPesanan.query.filter(PendapatanPesanan.marketplace.in_(marketplace_set)).all()
        if (p.marketplace, p.no_pesanan) in kunci_pesanan_set
    }
    nama_produk_set = {it.nama_produk for it in item_list}
    hpp_map = {
        p.nama_produk: (p.hpp or 0)
        for p in Produk.query.filter(Produk.nama_produk.in_(nama_produk_set)).all()
    }

    by_order = {}
    for it in item_list:
        by_order.setdefault((it.marketplace, it.no_pesanan), []).append(it)

    agregat_produk = {}
    tren_harian = {}

    for kunci_order, items in by_order.items():
        inc = income_map.get(kunci_order)
        total_income_order = inc.total_penghasilan if inc else None
        total_subtotal_order = sum(it.subtotal for it in items)
        for it in items:
            if total_income_order is None:
                income_alokasi = None
            elif total_subtotal_order:
                income_alokasi = round(total_income_order * (it.subtotal / total_subtotal_order))
            else:
                income_alokasi = round(total_income_order / len(items))

            hpp_satuan = hpp_map.get(it.nama_produk, 0)
            hpp_total_item = hpp_satuan * it.jumlah

            a = agregat_produk.setdefault(it.nama_produk, {
                "nama_produk": it.nama_produk, "qty": 0, "omzet": 0, "income": 0, "hpp_total": 0,
                "profit": 0, "hpp_satuan": hpp_satuan, "ada_hpp": hpp_satuan > 0, "ada_income": False,
            })
            a["qty"] += it.jumlah
            a["omzet"] += it.subtotal
            a["hpp_total"] += hpp_total_item
            if income_alokasi is not None:
                a["income"] += income_alokasi
                a["ada_income"] = True
                a["profit"] += income_alokasi - hpp_total_item

            tgl = it.tanggal_pesanan.isoformat()
            t = tren_harian.setdefault(tgl, {"tanggal": tgl, "omzet": 0, "income": 0, "hpp": 0, "profit": 0})
            t["omzet"] += it.subtotal
            t["hpp"] += hpp_total_item
            if income_alokasi is not None:
                t["income"] += income_alokasi
                t["profit"] += income_alokasi - hpp_total_item

    daftar_produk = list(agregat_produk.values())
    for p in daftar_produk:
        p["margin"] = (p["profit"] / p["income"] * 100) if p["income"] else 0

    # Rincian biaya per ORDER (bukan per baris produk, biar tidak dobel-hitung kalau
    # 1 pesanan berisi beberapa produk) -- dipakai buat breakdown "Profit Kamu Bocor
    # di Sini" di Dashboard Profit.
    total_biaya_admin = total_biaya_layanan = total_biaya_lainnya = 0
    for kunci_order in by_order:
        inc = income_map.get(kunci_order)
        if inc:
            total_biaya_admin += inc.biaya_admin or 0
            total_biaya_layanan += inc.biaya_layanan or 0
            total_biaya_lainnya += inc.biaya_lainnya or 0

    # Biaya iklan dari periode & marketplace yang sama (data dari Upload Data > File
    # Data Iklan, tabel IklanMarketplace) -- supaya Dashboard Profit ikut menghitung
    # potongan biaya iklan, bukan cuma HPP & fee marketplace.
    q_iklan = IklanMarketplace.query.filter(IklanMarketplace.marketplace.in_(marketplace_set))
    if tahun_f and bulan_f:
        q_iklan = q_iklan.filter(
            db.extract("year", IklanMarketplace.tanggal) == tahun_f,
            db.extract("month", IklanMarketplace.tanggal) == bulan_f,
        )
    total_biaya_iklan = sum(r.biaya or 0 for r in q_iklan.all())

    ringkasan = {
        "total_omzet": sum(p["omzet"] for p in daftar_produk),
        "total_income": sum(p["income"] for p in daftar_produk),
        "total_hpp": sum(p["hpp_total"] for p in daftar_produk),
        "total_profit": sum(p["profit"] for p in daftar_produk),
        "jumlah_produk": len(daftar_produk),
        "jumlah_produk_ada_hpp": sum(1 for p in daftar_produk if p["hpp_satuan"] > 0),
        "jumlah_pesanan": len(by_order),
        "total_qty": sum(p["qty"] for p in daftar_produk),
        "total_biaya_admin": total_biaya_admin,
        "total_biaya_layanan": total_biaya_layanan,
        "total_biaya_lainnya": total_biaya_lainnya,
        "total_biaya_iklan": total_biaya_iklan,
    }
    ringkasan["margin"] = (ringkasan["total_profit"] / ringkasan["total_income"] * 100) if ringkasan["total_income"] else 0
    ringkasan["total_potongan"] = ringkasan["total_omzet"] - ringkasan["total_income"]
    ringkasan["total_profit_real"] = ringkasan["total_income"] - ringkasan["total_hpp"] - total_biaya_iklan
    ringkasan["margin_real"] = (
        ringkasan["total_profit_real"] / ringkasan["total_omzet"] * 100
    ) if ringkasan["total_omzet"] else 0

    tren = [tren_harian[t] for t in sorted(tren_harian.keys())]
    return {"produk": daftar_produk, "tren": tren, "ringkasan": ringkasan}


def hitung_ringkasan_gabungan(order_item_list, income_item_list):
    """Ringkasan reconciliation Order+Income tepat setelah upload: dipakai buat halaman
    Upload Data supaya user langsung lihat apakah datanya nyambung dengan baik, sebelum
    lanjut ke tab Order & Income / HPP. Basis tanggal yang dipakai di sini SELALU tanggal
    pesanan selesai dari file Order -- beda dengan sheet Summary resmi Shopee yang basisnya
    tanggal dana dicairkan, jadi wajar kalau angkanya tidak identik."""
    pesanan_unik = {item["no_pesanan"] for item in order_item_list}
    produk_unik = {item["nama_produk"] for item in order_item_list}
    total_sebelum_diskon = sum(item.get("subtotal_kotor", item["subtotal"]) for item in order_item_list)
    total_setelah_diskon = sum(item["subtotal"] for item in order_item_list)

    fee_per_pesanan = {}
    for item in income_item_list:
        fee_per_pesanan[item["no_pesanan"]] = fee_per_pesanan.get(item["no_pesanan"], 0) + (
            item["biaya_admin"] + item["biaya_layanan"] + item["biaya_lainnya"]
        )
    total_fee = sum(fee_per_pesanan.values())

    income_no_set = set(fee_per_pesanan.keys())
    baris_order_tanpa_income = sum(1 for item in order_item_list if item["no_pesanan"] not in income_no_set)
    fee_teralokasi = sum(v for no, v in fee_per_pesanan.items() if no in pesanan_unik)
    fee_belum_teralokasi = total_fee - fee_teralokasi

    return {
        "jumlah_pesanan": len(pesanan_unik),
        "jumlah_produk_unik": len(produk_unik),
        "total_sebelum_diskon": total_sebelum_diskon,
        "total_setelah_diskon": total_setelah_diskon,
        "total_fee": total_fee,
        "baris_order_tanpa_income": baris_order_tanpa_income,
        "fee_teralokasi": fee_teralokasi,
        "fee_belum_teralokasi": fee_belum_teralokasi,
    }


def hitung_pendapatan_gross_marketplace(bulan, tahun):
    """Omzet kotor (Subtotal Pesanan, sebelum potongan biaya platform) per marketplace
    untuk 1 bulan -- SATU sumber data yang dipakai bareng oleh menu Penjualan dan
    Laporan Laba/Rugi. Prioritas data dari Profitabilitas (Order & Income); kalau
    suatu marketplace belum ada datanya di Profitabilitas untuk bulan itu, fallback ke
    data Penjualan lama (PenjualanMarketplace) supaya laporan bulan-bulan lama sebelum
    ada Profitabilitas tidak mendadak jadi kosong."""
    awal = date(tahun, bulan, 1)
    akhir = date(tahun, bulan, calendar.monthrange(tahun, bulan)[1])

    omzet_profit = {}
    pesanan_unik_profit = {}
    for r in PesananMarketplace.query.filter(
        PesananMarketplace.tanggal_pesanan >= awal, PesananMarketplace.tanggal_pesanan <= akhir,
        PesananMarketplace.status_pesanan.notin_(STATUS_BATAL_MARKETPLACE),
    ).all():
        omzet_profit[r.marketplace] = omzet_profit.get(r.marketplace, 0) + (r.subtotal or 0)
        pesanan_unik_profit.setdefault(r.marketplace, set()).add(r.no_pesanan)

    legacy_by_mp = {}
    for r in PenjualanMarketplace.query.filter(
        PenjualanMarketplace.tanggal >= awal, PenjualanMarketplace.tanggal <= akhir
    ).all():
        agg = legacy_by_mp.setdefault(r.marketplace, {"total_penjualan": 0, "total_diskon": 0, "jumlah_pesanan": 0})
        agg["total_penjualan"] += r.total_penjualan or 0
        agg["total_diskon"] += r.total_diskon or 0
        agg["jumlah_pesanan"] += r.jumlah_pesanan or 0

    hasil = {}
    for mp in set(omzet_profit) | set(legacy_by_mp):
        if mp in omzet_profit:
            hasil[mp] = {
                "pendapatan": omzet_profit[mp],
                "jumlah_pesanan": len(pesanan_unik_profit.get(mp, ())),
                "sumber": "profitabilitas",
            }
        else:
            leg = legacy_by_mp[mp]
            hasil[mp] = {
                "pendapatan": leg["total_penjualan"] - leg["total_diskon"],
                "jumlah_pesanan": leg["jumlah_pesanan"],
                "sumber": "legacy",
            }
    return hasil


def baca_ringkasan_summary_shopee(file_storage):
    """Baca sheet 'Summary' laporan Income Shopee apa adanya (tidak diproses/dihitung ulang)
    supaya user bisa cross-check ke laporan resmi Shopee sendiri."""
    filename = file_storage.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return None
    try:
        raw = file_storage.read()
        wb = CalamineWorkbook.from_filelike(io.BytesIO(raw))
    except Exception:
        return None

    nama_sheet = next((s for s in wb.sheet_names if s.strip().lower() == "summary"), None)
    if not nama_sheet:
        return None
    try:
        baris_sheet = wb.get_sheet_by_name(nama_sheet).to_python()
    except Exception:
        return None

    hasil = []
    kosong_berturut = 0
    for row in baris_sheet[:80]:
        row = list(row) + [None] * (4 - len(row))
        label0, label1, nilai_tengah, nilai_kanan = row[:4]
        if all(c in (None, "") for c in row[:4]):
            kosong_berturut += 1
            if kosong_berturut >= 8 and hasil:
                break
            continue
        kosong_berturut = 0
        if label1 not in (None, ""):
            label, level = str(label1), 1
            nilai = nilai_tengah if nilai_tengah not in (None, "") else nilai_kanan
        elif label0 not in (None, ""):
            label, level = str(label0), 0
            nilai = nilai_kanan
        else:
            continue
        hasil.append({"label": label, "nilai": nilai, "level": level})
    return hasil or None


def baca_file_iklan(file_storage):
    filename = file_storage.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    raw = file_storage.read()

    if ext == "csv":
        text = raw.decode("utf-8-sig", errors="ignore")
        semua = _parse_csv_delimiter_terbaik(text)
    elif ext in ("xlsx", "xlsm"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception:
            return None, None, None, "File Excel tidak bisa dibaca. Pastikan file tidak rusak."
        ws = wb.active
        semua = []
        for row in ws.iter_rows(values_only=True):
            if any(cell not in (None, "") for cell in row):
                semua.append(["" if c is None else c for c in row])
    else:
        return None, None, None, "Format file tidak didukung. Gunakan file CSV atau XLSX (export laporan iklan dari marketplace)."

    if len(semua) < 2:
        return None, None, None, "File kosong atau tidak ada baris data."

    idx_header = _cari_baris_header(semua)
    if len(semua) - idx_header < 2:
        return None, None, None, "Tidak ditemukan baris data setelah baris header di file ini."

    headers = [str(h).strip() for h in semua[idx_header]]
    rows = semua[idx_header + 1:]
    return headers, rows, idx_header, None


def parse_angka_iklan(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if not s:
        return 0
    s = s.replace("Rp", "").replace("rp", "").replace("%", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        kanan = s.split(",")[-1]
        s = s.replace(",", ".") if len(kanan) <= 2 else s.replace(",", "")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    elif "." in s:
        kepala, ekor = s.split(".")
        if len(ekor) == 3 and len(kepala) <= 3:
            s = s.replace(".", "")
    cleaned = "".join(ch for ch in s if ch.isdigit() or ch in "-.")
    try:
        return float(cleaned) if cleaned not in ("", "-", ".") else 0
    except ValueError:
        return 0


SINGKATAN_BULAN_KHUSUS = {8: "ags"}  # Agustus umum disingkat "Ags", bukan "Agu"


def parse_tanggal_iklan(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # nilai tanggal Excel (serial number sejak 1899-12-30)
        try:
            return date(1899, 12, 30) + timedelta(days=int(value))
        except (OverflowError, ValueError):
            return None

    s = str(value).strip().replace("\xa0", " ")
    if not s:
        return None
    if "T" in s and len(s) >= 10 and s[:4].isdigit():
        s = s.split("T")[0]

    parts_nama_bulan = s.replace("-", " ").replace(",", " ").split()
    if len(parts_nama_bulan) == 3 and parts_nama_bulan[0].isdigit() and parts_nama_bulan[2].isdigit():
        try:
            hari = int(parts_nama_bulan[0])
            tahun = int(parts_nama_bulan[2])
            if tahun < 100:
                tahun += 2000
            bulan_txt = parts_nama_bulan[1].strip(".").lower()
            for idx, nama in enumerate(BULAN_NAMA):
                if idx == 0:
                    continue
                singkatan = SINGKATAN_BULAN_KHUSUS.get(idx, nama.lower()[:3])
                if bulan_txt in (nama.lower(), nama.lower()[:3], singkatan):
                    return date(tahun, idx, hari)
        except (ValueError, IndexError):
            pass

    formats = [
        "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
        "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y", "%Y.%m.%d",
        # Varian dengan jam -- dicoba lewat strptime eksplisit dulu (bukan lewat dateutil
        # fallback di bawah) karena dateutil's dayfirst=True ternyata bisa salah tukar
        # bulan/hari untuk string berformat tahun-dulu seperti "2026-06-01 00:00:00".
        "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    try:
        # Kalau string jelas diawali tahun 4 digit (mis. "2026-06-01 10:20:30"), pakai
        # yearfirst=True & dayfirst=False -- dateutil ternyata bisa salah tukar bulan/hari
        # kalau dayfirst=True tetap dipaksa walau tahun sudah jelas di depan.
        tahun_dulu = len(s) > 4 and s[:4].isdigit() and s[4] in "-/."
        if tahun_dulu:
            return date_parser.parse(s, yearfirst=True, dayfirst=False, fuzzy=False).date()
        return date_parser.parse(s, dayfirst=True, fuzzy=False).date()
    except (ValueError, OverflowError, TypeError):
        return None


_SATUAN = [
    "", "Satu", "Dua", "Tiga", "Empat", "Lima", "Enam", "Tujuh", "Delapan", "Sembilan",
    "Sepuluh", "Sebelas",
]


def _terbilang(n):
    n = int(n)
    if n < 12:
        return _SATUAN[n]
    if n < 20:
        return _terbilang(n - 10) + " Belas"
    if n < 100:
        return _terbilang(n // 10) + " Puluh " + _terbilang(n % 10)
    if n < 200:
        return "Seratus " + _terbilang(n - 100)
    if n < 1000:
        return _terbilang(n // 100) + " Ratus " + _terbilang(n % 100)
    if n < 2000:
        return "Seribu " + _terbilang(n - 1000)
    if n < 1000000:
        return _terbilang(n // 1000) + " Ribu " + _terbilang(n % 1000)
    if n < 1000000000:
        return _terbilang(n // 1000000) + " Juta " + _terbilang(n % 1000000)
    return _terbilang(n // 1000000000) + " Miliar " + _terbilang(n % 1000000000)


def terbilang_rupiah(n):
    n = int(n or 0)
    if n == 0:
        return "Nol Rupiah"
    words = _terbilang(n) + " Rupiah"
    return " ".join(words.split())


def haversine_meter(lat1, lng1, lat2, lng2):
    r = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "ganti-secret-key-ini")

    db_url = os.environ.get("DATABASE_URL", "sqlite:///absensi.db")
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = db_url
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

    upload_folder = os.path.join(app.static_folder, "uploads")
    os.makedirs(upload_folder, exist_ok=True)
    app.config["UPLOAD_FOLDER"] = upload_folder

    izin_upload_folder = os.path.join(upload_folder, "izin")
    os.makedirs(izin_upload_folder, exist_ok=True)
    app.config["IZIN_UPLOAD_FOLDER"] = izin_upload_folder

    laporan_upload_folder = os.path.join(upload_folder, "laporan")
    os.makedirs(laporan_upload_folder, exist_ok=True)
    app.config["LAPORAN_UPLOAD_FOLDER"] = laporan_upload_folder

    # Batas ukuran file upload di seluruh aplikasi, supaya server tidak kehabisan
    # memori/disk (paket hosting free tier) gara-gara file besar -- ini yang bikin
    # aplikasi lemot kalau tidak dibatasi.
    app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB

    EKSTENSI_DOKUMEN_IZIN = {"pdf", "doc", "docx", "jpg", "jpeg", "png"}
    EKSTENSI_LAMPIRAN_LAPORAN = {
        "pdf", "doc", "docx", "xls", "xlsx", "ppt", "pptx",
        "jpg", "jpeg", "png", "csv", "txt", "zip",
    }

    tmp_iklan_folder = os.path.join(app.instance_path, "tmp_iklan")
    os.makedirs(tmp_iklan_folder, exist_ok=True)
    app.config["TMP_IKLAN_FOLDER"] = tmp_iklan_folder

    db.init_app(app)

    login_manager = LoginManager()
    login_manager.login_view = "login"
    login_manager.login_message = "Silakan login terlebih dahulu."
    login_manager.init_app(app)

    @login_manager.user_loader
    def load_user(user_id):
        kind, _, raw_id = user_id.partition(":")
        if kind == "admin":
            return db.session.get(User, int(raw_id))
        if kind == "emp":
            return db.session.get(Employee, int(raw_id))
        return None

    def admin_required(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))
            if getattr(current_user, "role", None) != "admin":
                return redirect(url_for("pegawai_dashboard"))
            return f(*args, **kwargs)
        return wrapper

    def pegawai_required(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("pegawai_login"))
            if getattr(current_user, "role", None) != "pegawai":
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return wrapper

    def marketing_required(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))
            role = getattr(current_user, "role", None)
            if role == "admin":
                return f(*args, **kwargs)
            if role == "pegawai" and current_user.akses_marketing:
                return f(*args, **kwargs)
            if role == "pegawai":
                return redirect(url_for("pegawai_dashboard"))
            return redirect(url_for("login"))
        return wrapper

    def rupiah(value):
        try:
            value = int(value or 0)
        except (ValueError, TypeError):
            value = 0
        return f"Rp {value:,.0f}".replace(",", ".")

    app.jinja_env.filters["rupiah"] = rupiah
    app.jinja_env.filters["terbilang"] = terbilang_rupiah

    def buat_pesan_wa_slip(p, settings):
        label_total = "GAJI BRUTO" if p.tipe_pegawai == "Freelance" else "PENERIMAAN BERSIH"
        baris = [
            f"Halo {p.employee.nama},",
            "",
            f"Berikut slip gaji Anda periode {BULAN_NAMA[p.bulan]} {p.tahun} dari {settings.nama_perusahaan} "
            "(rincian lengkap ada di file PDF terlampir).",
            "",
            f"{label_total}: {rupiah(p.gaji_bersih)}",
            f"({terbilang_rupiah(p.gaji_bersih)})",
            "",
            "Status: SUDAH DIBAYAR",
            "",
            "Terima kasih atas kerja keras Anda selama ini. Mohon dicek kembali, jika ada pertanyaan silakan hubungi admin.",
        ]
        return "\n".join(baris)

    def normalisasi_no_hp_wa(no_hp):
        digit = "".join(ch for ch in (no_hp or "") if ch.isdigit())
        if not digit:
            return ""
        if digit.startswith("0"):
            digit = "62" + digit[1:]
        elif not digit.startswith("62"):
            digit = "62" + digit
        return digit

    def buat_wa_link(p, settings):
        if p.status != "Dibayar":
            return None
        nomor = normalisasi_no_hp_wa(p.employee.no_hp)
        if not nomor:
            return None
        pesan = buat_pesan_wa_slip(p, settings)
        return f"https://wa.me/{nomor}?text={quote(pesan)}"

    def format_tanggal_panjang(d):
        return f"{d.day:02d} {BULAN_NAMA[d.month]} {d.year}"

    app.jinja_env.filters["tanggal_panjang"] = format_tanggal_panjang

    def format_tanggal_lengkap(d):
        return f"{HARI_NAMA[d.weekday()]}, {d.day:02d} {BULAN_NAMA[d.month]} {d.year}"

    app.jinja_env.filters["tanggal_lengkap"] = format_tanggal_lengkap

    def apakah_hari_libur(tanggal):
        if tanggal is None:
            return None
        return HariLibur.query.filter_by(tanggal=tanggal).first()

    def get_settings():
        settings = Settings.query.first()
        if not settings:
            settings = Settings()
            db.session.add(settings)
            db.session.commit()
        return settings

    def parse_hhmm(value):
        try:
            parts = value.split(":")
            return int(parts[0]) * 60 + int(parts[1])
        except (ValueError, AttributeError, IndexError, TypeError):
            return None

    def standar_jam_pegawai(settings, tipe_pegawai):
        if tipe_pegawai == "Freelance":
            return settings.jam_masuk_standar_freelance, settings.jam_pulang_standar_freelance
        return settings.jam_masuk_standar, settings.jam_pulang_standar

    BATAS_MASUK_SHIFT_MALAM = 17 * 60  # 17:00 -- freelance yang absen masuk mulai jam ini dianggap shift malam

    def hitung_telat_lembur(jam_masuk, jam_pulang, settings, tipe_pegawai="Karyawan Tetap", tanggal=None, employee_id=None):
        telat = 0
        lembur = 0
        actual_masuk = parse_hhmm(jam_masuk)
        actual_pulang = parse_hhmm(jam_pulang)

        # Karyawan Tetap/Probation yang masuk di tanggal merah tanpa pernah mengajukan
        # lembur untuk tanggal itu: seluruh jam kerja hari itu dihitung lembur (bukan
        # dibandingkan ke jam standar, karena hari itu bukan hari kerja wajib). Freelance
        # tidak ikut aturan ini -- upahnya tetap dihitung harian seperti biasa.
        if (
            tipe_pegawai in ("Karyawan Tetap", "Probation")
            and tanggal is not None
            and employee_id is not None
            and apakah_hari_libur(tanggal)
            and not PengajuanLembur.query.filter_by(employee_id=employee_id, tanggal=tanggal).first()
        ):
            if actual_masuk is not None and actual_pulang is not None:
                lembur = max(actual_pulang - actual_masuk, 0)
            return 0, lembur

        jam_masuk_standar, jam_pulang_standar = standar_jam_pegawai(settings, tipe_pegawai)
        standar_masuk = parse_hhmm(jam_masuk_standar)
        standar_pulang = parse_hhmm(jam_pulang_standar)

        # Freelance yang absen masuk sore/malam (mis. shift co-host mulai ~18:00) punya jam
        # kerja yang sama sekali berbeda dari standar freelance siang (08:00-17:00) yang
        # dikonfigurasi di Pengaturan -- jangan hitung telat/lembur berdasarkan standar itu
        # untuk mereka, karena jam standarnya memang tidak relevan untuk shift malam.
        shift_malam = (
            tipe_pegawai == "Freelance"
            and actual_masuk is not None
            and actual_masuk >= BATAS_MASUK_SHIFT_MALAM
        )

        if not shift_malam and standar_masuk is not None and actual_masuk is not None:
            selisih = actual_masuk - standar_masuk - (settings.toleransi_telat_menit or 0)
            if selisih > 0:
                telat = selisih

        if not shift_malam and standar_pulang is not None and actual_pulang is not None:
            selisih = actual_pulang - standar_pulang
            if selisih >= (settings.toleransi_lembur_menit or 0):
                lembur = selisih

        return telat, lembur

    def cek_pulang_cepat(jam_pulang, settings, tipe_pegawai="Karyawan Tetap"):
        _, jam_pulang_standar = standar_jam_pegawai(settings, tipe_pegawai)
        standar_pulang = parse_hhmm(jam_pulang_standar)
        actual_pulang = parse_hhmm(jam_pulang)
        if standar_pulang is None or actual_pulang is None:
            return 0
        selisih = standar_pulang - actual_pulang
        return selisih if selisih > 0 else 0

    with app.app_context():
        db.create_all()
        kolom_employee = {c["name"] for c in db.inspect(db.engine).get_columns("employee")}
        if "akses_marketing" not in kolom_employee:
            db.session.execute(db.text(
                "ALTER TABLE employee ADD COLUMN akses_marketing BOOLEAN NOT NULL DEFAULT 0"
            ))
            db.session.commit()
        kolom_payroll = {c["name"] for c in db.inspect(db.engine).get_columns("payroll")}
        if "jumlah_lembur" not in kolom_payroll:
            db.session.execute(db.text("ALTER TABLE payroll ADD COLUMN jumlah_lembur INTEGER DEFAULT 0"))
            db.session.commit()
        if "tarif_lembur" not in kolom_payroll:
            db.session.execute(db.text("ALTER TABLE payroll ADD COLUMN tarif_lembur INTEGER DEFAULT 0"))
            db.session.commit()
        kolom_bahan_baku = {c["name"] for c in db.inspect(db.engine).get_columns("bahan_baku")}
        for kolom, tipe in [("warna", "VARCHAR(64)"), ("tinggi_meter", "FLOAT"), ("suplier", "VARCHAR(128)")]:
            if kolom not in kolom_bahan_baku:
                db.session.execute(db.text(f"ALTER TABLE bahan_baku ADD COLUMN {kolom} {tipe}"))
                db.session.commit()
        kolom_bb_transaksi = {c["name"] for c in db.inspect(db.engine).get_columns("bahan_baku_transaksi")}
        for kolom, tipe in [
            ("warna", "VARCHAR(64)"), ("suplier", "VARCHAR(128)"), ("tinggi_meter", "FLOAT"),
            ("harga_per_yard", "INTEGER"), ("total_dibayar", "INTEGER"), ("vendor", "VARCHAR(128)"),
            ("lebar_kain", "FLOAT"), ("produk_jadi_pcs", "INTEGER"), ("purchase_order_id", "INTEGER"),
        ]:
            if kolom not in kolom_bb_transaksi:
                db.session.execute(db.text(f"ALTER TABLE bahan_baku_transaksi ADD COLUMN {kolom} {tipe}"))
                db.session.commit()
        kolom_spek_ukuran = {c["name"] for c in db.inspect(db.engine).get_columns("produk_spek_ukuran")}
        if "kategori" not in kolom_spek_ukuran:
            db.session.execute(db.text("ALTER TABLE produk_spek_ukuran ADD COLUMN kategori VARCHAR(24)"))
            db.session.commit()
        kolom_po = {c["name"] for c in db.inspect(db.engine).get_columns("purchase_order")}
        if "status" not in kolom_po:
            db.session.execute(db.text(
                "ALTER TABLE purchase_order ADD COLUMN status VARCHAR(24) NOT NULL DEFAULT 'Menunggu Produksi'"
            ))
            db.session.commit()
        kolom_po_pembayaran = {c["name"] for c in db.inspect(db.engine).get_columns("purchase_order_pembayaran")}
        if "metode" not in kolom_po_pembayaran:
            db.session.execute(db.text(
                "ALTER TABLE purchase_order_pembayaran ADD COLUMN metode VARCHAR(16) NOT NULL DEFAULT 'Cicilan'"
            ))
            db.session.commit()
        if not User.query.first():
            admin = User(username="admin", nama="Administrator")
            admin.set_password("admin123")
            db.session.add(admin)
            db.session.commit()
        if not HariLibur.query.first():
            for tanggal_str, keterangan in HARI_LIBUR_2026:
                db.session.add(HariLibur(
                    tanggal=datetime.strptime(tanggal_str, "%Y-%m-%d").date(),
                    keterangan=keterangan,
                ))
            db.session.commit()
        get_settings()

    # ---------- AUTH ADMIN ----------
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            if getattr(current_user, "role", None) == "admin":
                return redirect(url_for("dashboard"))
            logout_user()
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user)
                return redirect(url_for("dashboard"))
            flash("Username atau password salah.", "danger")
        return render_template("login.html")

    @app.route("/logout")
    def logout():
        logout_user()
        return redirect(url_for("login"))

    # ---------- AUTH PEGAWAI ----------
    @app.route("/pegawai/login", methods=["GET", "POST"])
    def pegawai_login():
        if current_user.is_authenticated:
            if getattr(current_user, "role", None) == "pegawai":
                return redirect(url_for("pegawai_dashboard"))
            logout_user()
        if request.method == "POST":
            no_hp = request.form.get("no_hp", "").strip()
            password = request.form.get("password", "")
            emp = Employee.query.filter_by(no_hp=no_hp, status="Aktif").first()
            if emp and emp.check_password(password):
                login_user(emp)
                return redirect(url_for("pegawai_dashboard"))
            flash("No. HP atau password salah.", "danger")
        return render_template("pegawai/login.html")

    @app.route("/pegawai/logout")
    def pegawai_logout():
        logout_user()
        return redirect(url_for("pegawai_login"))

    # ---------- DASHBOARD ----------
    @app.route("/")
    @admin_required
    def dashboard():
        total_karyawan = Employee.query.filter_by(status="Aktif").count()
        hari_ini = today_wib()
        absen_hari_ini = Attendance.query.filter_by(tanggal=hari_ini).count()
        hadir_hari_ini = Attendance.query.filter_by(tanggal=hari_ini, status="Hadir").count()

        bulan_ini, tahun_ini = hari_ini.month, hari_ini.year
        payrolls_bulan_ini = Payroll.query.filter_by(bulan=bulan_ini, tahun=tahun_ini).all()
        total_gaji_bulan_ini = sum(p.gaji_bersih for p in payrolls_bulan_ini)

        return render_template(
            "dashboard.html",
            total_karyawan=total_karyawan,
            absen_hari_ini=absen_hari_ini,
            hadir_hari_ini=hadir_hari_ini,
            total_gaji_bulan_ini=total_gaji_bulan_ini,
            bulan_nama=BULAN_NAMA[bulan_ini],
            tahun_ini=tahun_ini,
        )

    # ---------- KARYAWAN ----------
    @app.route("/karyawan")
    @admin_required
    def karyawan_list():
        karyawan = Employee.query.order_by(Employee.nama).all()
        return render_template("employees_list.html", karyawan=karyawan)

    @app.route("/karyawan/tambah", methods=["GET", "POST"])
    @admin_required
    def karyawan_tambah():
        if request.method == "POST":
            no_hp = request.form.get("no_hp", "").strip()
            if no_hp and Employee.query.filter_by(no_hp=no_hp).first():
                flash("No. HP sudah dipakai karyawan lain. Gunakan nomor lain agar login tidak tertukar.", "danger")
                return render_template("employee_form.html", karyawan=None, jabatan_list=JABATAN_LIST)

            emp = Employee(
                nama=request.form["nama"].strip(),
                jabatan=request.form.get("jabatan", "").strip(),
                email=request.form.get("email", "").strip(),
                no_hp=no_hp,
                alamat=request.form.get("alamat", "").strip(),
                nomor_rekening=request.form.get("nomor_rekening", "").strip(),
                gaji_pokok=int(request.form.get("gaji_pokok") or 0),
                tunjangan_makan=int(request.form.get("tunjangan_makan") or 0),
                tunjangan_transport=int(request.form.get("tunjangan_transport") or 0),
                tipe_pegawai=request.form.get("tipe_pegawai", "Karyawan Tetap"),
                target_tercapai=request.form.get("target_tercapai", "Tidak Tercapai"),
                bpjs_jkk=int(request.form.get("bpjs_jkk") or 0),
                bpjs_jkm=int(request.form.get("bpjs_jkm") or 0),
                bpjs_jht=int(request.form.get("bpjs_jht") or 0),
                bpjs_kesehatan_terdaftar=request.form.get("bpjs_kesehatan_terdaftar") == "on",
                tarif_unit_freelance=int(request.form.get("tarif_unit_freelance") or 0),
                co_host_bulan_ini=request.form.get("co_host_bulan_ini", "Tidak"),
                status=request.form.get("status", "Aktif"),
                akses_marketing=request.form.get("akses_marketing") == "on",
            )
            password_baru = request.form.get("password_baru", "")
            if password_baru:
                emp.set_password(password_baru)
            db.session.add(emp)
            db.session.commit()
            flash(f"Karyawan {emp.nama} berhasil ditambahkan.", "success")
            return redirect(url_for("karyawan_list"))
        return render_template("employee_form.html", karyawan=None, jabatan_list=JABATAN_LIST)

    @app.route("/karyawan/<int:emp_id>/edit", methods=["GET", "POST"])
    @admin_required
    def karyawan_edit(emp_id):
        emp = db.session.get(Employee, emp_id) or abort_404()
        hari_ini = today_wib()
        hari_hadir_bulan_ini = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.status == "Hadir",
            db.extract("month", Attendance.tanggal) == hari_ini.month,
            db.extract("year", Attendance.tanggal) == hari_ini.year,
        ).count()
        if request.method == "POST":
            no_hp = request.form.get("no_hp", "").strip()
            if no_hp and Employee.query.filter(Employee.no_hp == no_hp, Employee.id != emp.id).first():
                flash("No. HP sudah dipakai karyawan lain. Gunakan nomor lain agar login tidak tertukar.", "danger")
                return render_template(
                    "employee_form.html", karyawan=emp, jabatan_list=JABATAN_LIST,
                    hari_hadir_bulan_ini=hari_hadir_bulan_ini,
                )

            emp.nama = request.form["nama"].strip()
            emp.jabatan = request.form.get("jabatan", "").strip()
            emp.email = request.form.get("email", "").strip()
            emp.no_hp = no_hp
            emp.alamat = request.form.get("alamat", "").strip()
            emp.nomor_rekening = request.form.get("nomor_rekening", "").strip()
            emp.gaji_pokok = int(request.form.get("gaji_pokok") or 0)
            emp.tunjangan_makan = int(request.form.get("tunjangan_makan") or 0)
            emp.tunjangan_transport = int(request.form.get("tunjangan_transport") or 0)
            emp.tipe_pegawai = request.form.get("tipe_pegawai", "Karyawan Tetap")
            emp.target_tercapai = request.form.get("target_tercapai", "Tidak Tercapai")
            emp.bpjs_jkk = int(request.form.get("bpjs_jkk") or 0)
            emp.bpjs_jkm = int(request.form.get("bpjs_jkm") or 0)
            emp.bpjs_jht = int(request.form.get("bpjs_jht") or 0)
            emp.bpjs_kesehatan_terdaftar = request.form.get("bpjs_kesehatan_terdaftar") == "on"
            emp.tarif_unit_freelance = int(request.form.get("tarif_unit_freelance") or 0)
            emp.co_host_bulan_ini = request.form.get("co_host_bulan_ini", "Tidak")
            emp.status = request.form.get("status", "Aktif")
            emp.akses_marketing = request.form.get("akses_marketing") == "on"
            password_baru = request.form.get("password_baru", "")
            if password_baru:
                emp.set_password(password_baru)
            db.session.commit()
            flash(f"Data {emp.nama} berhasil diperbarui.", "success")
            return redirect(url_for("karyawan_list"))
        return render_template(
            "employee_form.html", karyawan=emp, jabatan_list=JABATAN_LIST,
            hari_hadir_bulan_ini=hari_hadir_bulan_ini,
        )

    @app.route("/karyawan/<int:emp_id>/hapus", methods=["POST"])
    @admin_required
    def karyawan_hapus(emp_id):
        emp = db.session.get(Employee, emp_id) or abort_404()
        nama = emp.nama
        db.session.delete(emp)
        db.session.commit()
        flash(f"Karyawan {nama} beserta riwayatnya dihapus.", "info")
        return redirect(url_for("karyawan_list"))

    # ---------- DATA PRODUK & HARGA JUAL ----------
    def _hitung_margin_produk(p):
        def margin(harga):
            if not harga:
                return 0, 0
            nilai = harga - p.hpp
            persen = (nilai / harga * 100) if harga else 0
            return nilai, persen

        margin_hn, persen_hn = margin(p.harga_normal)
        margin_hfs, persen_hfs = margin(p.harga_flash_sale)
        margin_hc, persen_hc = margin(p.harga_big_campaign)
        return {
            "produk": p,
            "margin_hn": margin_hn, "persen_hn": persen_hn,
            "margin_hfs": margin_hfs, "persen_hfs": persen_hfs,
            "margin_hc": margin_hc, "persen_hc": persen_hc,
        }

    @app.route("/produk")
    @admin_required
    def produk_list():
        daftar = Produk.query.order_by(Produk.nama_produk).all()
        data = [_hitung_margin_produk(p) for p in daftar]
        return render_template("produk_list.html", data=data)

    @app.route("/produk/tambah", methods=["GET", "POST"])
    @admin_required
    def produk_tambah():
        if request.method == "POST":
            produk = Produk(
                nama_produk=request.form.get("nama_produk", "").strip(),
                modal=int(request.form.get("modal") or 0),
                hpp=int(request.form.get("hpp") or 0),
                harga_dasar=int(request.form.get("harga_dasar") or 0),
                harga_normal=int(request.form.get("harga_normal") or 0),
                harga_flash_sale=int(request.form.get("harga_flash_sale") or 0),
                harga_big_campaign=int(request.form.get("harga_big_campaign") or 0),
            )
            if not produk.nama_produk:
                flash("Nama produk wajib diisi.", "danger")
                return render_template("produk_form.html", produk=None)
            db.session.add(produk)
            db.session.commit()
            flash(f"Produk {produk.nama_produk} berhasil ditambahkan.", "success")
            return redirect(url_for("produk_list"))
        return render_template("produk_form.html", produk=None)

    @app.route("/produk/<int:produk_id>/edit", methods=["GET", "POST"])
    @admin_required
    def produk_edit(produk_id):
        produk = db.session.get(Produk, produk_id) or abort_404()
        if request.method == "POST":
            produk.nama_produk = request.form.get("nama_produk", "").strip()
            produk.modal = int(request.form.get("modal") or 0)
            produk.hpp = int(request.form.get("hpp") or 0)
            produk.harga_dasar = int(request.form.get("harga_dasar") or 0)
            produk.harga_normal = int(request.form.get("harga_normal") or 0)
            produk.harga_flash_sale = int(request.form.get("harga_flash_sale") or 0)
            produk.harga_big_campaign = int(request.form.get("harga_big_campaign") or 0)
            if not produk.nama_produk:
                flash("Nama produk wajib diisi.", "danger")
                return render_template("produk_form.html", produk=produk)
            db.session.commit()
            flash(f"Produk {produk.nama_produk} berhasil diperbarui.", "success")
            return redirect(url_for("produk_list"))
        return render_template("produk_form.html", produk=produk)

    @app.route("/produk/<int:produk_id>/hapus", methods=["POST"])
    @admin_required
    def produk_hapus(produk_id):
        produk = db.session.get(Produk, produk_id) or abort_404()
        nama = produk.nama_produk
        db.session.delete(produk)
        db.session.commit()
        flash(f"Produk {nama} dihapus.", "info")
        return redirect(url_for("produk_list"))

    @app.route("/produk/upload", methods=["GET", "POST"])
    @admin_required
    def produk_upload():
        bersihkan_tmp_iklan_lama()
        if request.method == "POST":
            file = request.files.get("file")
            if not file or not file.filename:
                flash("Pilih file data produk (CSV/XLSX) terlebih dahulu.", "danger")
                return redirect(url_for("produk_upload"))

            headers, rows, _idx_header, error = baca_file_iklan(file)
            if error:
                flash(error, "danger")
                return redirect(url_for("produk_upload"))

            mapping = deteksi_kolom_produk(headers)
            if mapping.get("nama_produk") is None or mapping.get("hpp") is None or mapping.get("harga_jual") is None:
                flash(
                    "Sistem tidak berhasil mengenali kolom Nama Produk/HPP/Harga Jual secara otomatis dari file "
                    f"ini. Header yang terbaca: {', '.join(str(h) for h in headers)}. Kirimkan daftar kolom ini "
                    "biar formatnya bisa didukung.",
                    "danger",
                )
                return redirect(url_for("produk_upload"))

            idx_nama = mapping["nama_produk"]
            idx_variasi = mapping.get("variasi")
            idx_hpp = mapping["hpp"]
            idx_harga = mapping["harga_jual"]

            preview = []
            dilewati = 0
            for row in rows:
                nama_raw = row[idx_nama] if idx_nama < len(row) else None
                nama = str(nama_raw).strip() if nama_raw is not None else ""
                if not nama:
                    dilewati += 1
                    continue
                variasi_raw = row[idx_variasi] if idx_variasi is not None and idx_variasi < len(row) else None
                variasi = str(variasi_raw).strip() if variasi_raw not in (None, "") else ""
                nama_final = (f"{nama} - {variasi}" if variasi else nama)[:128]

                hpp_val = round(parse_angka_iklan(row[idx_hpp])) if idx_hpp < len(row) else 0
                harga_val = round(parse_angka_iklan(row[idx_harga])) if idx_harga < len(row) else 0
                preview.append({"nama_produk": nama_final, "hpp": hpp_val, "harga_jual": harga_val})

            if not preview:
                flash("Tidak ada baris data produk yang valid di file ini (kolom Nama Produk kosong semua).", "danger")
                return redirect(url_for("produk_upload"))

            token = uuid.uuid4().hex
            path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"produkharga_{token}.json")
            with open(path_tmp, "w", encoding="utf-8") as f:
                json.dump({"preview": preview, "sumber_file": secure_filename(file.filename)}, f)

            def nama_kolom(idx):
                return str(headers[idx]).strip() if idx is not None and idx < len(headers) else None

            kolom_terdeteksi = [
                ("Nama Produk", nama_kolom(idx_nama)),
                ("Variasi", nama_kolom(idx_variasi)),
                ("HPP", nama_kolom(idx_hpp)),
                ("Harga Jual", nama_kolom(idx_harga)),
            ]

            return render_template(
                "produk_review.html",
                token=token,
                kolom_terdeteksi=kolom_terdeteksi,
                preview=preview,
                dilewati=dilewati,
            )

        return render_template("produk_upload.html")

    @app.route("/produk/upload/konfirmasi", methods=["POST"])
    @admin_required
    def produk_upload_konfirmasi():
        token = request.form.get("token", "")
        path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"produkharga_{token}.json")
        if not os.path.isfile(path_tmp):
            flash("Sesi upload sudah kedaluwarsa, silakan upload ulang file.", "danger")
            return redirect(url_for("produk_upload"))

        with open(path_tmp, "r", encoding="utf-8") as f:
            data_tmp = json.load(f)

        jumlah_baru = jumlah_update = 0
        for item in data_tmp["preview"]:
            nama = item["nama_produk"]
            existing = Produk.query.filter_by(nama_produk=nama).first()
            if not existing:
                existing = Produk(nama_produk=nama)
                db.session.add(existing)
                jumlah_baru += 1
            else:
                jumlah_update += 1
            existing.modal = item["hpp"]
            existing.hpp = item["hpp"]
            existing.harga_dasar = item["harga_jual"]
            existing.harga_normal = item["harga_jual"]
            existing.harga_flash_sale = item["harga_jual"]
            existing.harga_big_campaign = item["harga_jual"]
        db.session.commit()
        os.remove(path_tmp)

        flash(
            f"Berhasil impor {len(data_tmp['preview'])} produk ({jumlah_baru} baru, {jumlah_update} diperbarui).",
            "success",
        )
        return redirect(url_for("produk_list"))

    # ---------- INVENTORY: BAHAN BAKU ----------
    @app.route("/inventory/bahan-baku", methods=["GET", "POST"])
    @admin_required
    def bahan_baku_list():
        if request.method == "POST":
            nama_bahan = request.form.get("nama_bahan", "").strip()
            satuan = request.form.get("satuan", "Yard").strip() or "Yard"
            stok_awal = parse_angka_iklan(request.form.get("stok_awal"))
            harga_per_yard = round(parse_angka_iklan(request.form.get("harga_per_yard")))
            if not nama_bahan:
                flash("Nama bahan baku wajib diisi.", "danger")
                return redirect(url_for("bahan_baku_list"))
            bahan = BahanBaku(
                nama_bahan=nama_bahan, satuan=satuan, stok_saat_ini=stok_awal, harga_per_yard=harga_per_yard,
            )
            db.session.add(bahan)
            db.session.flush()
            if stok_awal:
                db.session.add(BahanBakuTransaksi(
                    bahan_baku_id=bahan.id, tanggal=today_wib(), jenis="Masuk",
                    jumlah_yard=stok_awal, keterangan="Stok awal",
                ))
            db.session.commit()
            flash(f"Bahan baku {bahan.nama_bahan} berhasil ditambahkan.", "success")
            return redirect(url_for("bahan_baku_detail", bahan_id=bahan.id))

        daftar = BahanBaku.query.order_by(BahanBaku.nama_bahan).all()
        return render_template("inventory/bahan_baku_list.html", daftar=daftar)

    @app.route("/inventory/bahan-baku/<int:bahan_id>")
    @admin_required
    def bahan_baku_detail(bahan_id):
        bahan = db.session.get(BahanBaku, bahan_id) or abort_404()
        produk_list_semua = Produk.query.order_by(Produk.nama_produk).all()
        produk_sudah_dipetakan = {k.produk_id for k in bahan.kebutuhan_list}
        produk_pilihan = [p for p in produk_list_semua if p.id not in produk_sudah_dipetakan]
        return render_template(
            "inventory/bahan_baku_detail.html",
            bahan=bahan, produk_pilihan=produk_pilihan, produk_list_semua=produk_list_semua,
            tanggal_hari_ini=today_wib().isoformat(),
        )

    @app.route("/inventory/bahan-baku/<int:bahan_id>/edit", methods=["POST"])
    @admin_required
    def bahan_baku_edit(bahan_id):
        bahan = db.session.get(BahanBaku, bahan_id) or abort_404()
        nama_bahan = request.form.get("nama_bahan", "").strip()
        if not nama_bahan:
            flash("Nama bahan baku wajib diisi.", "danger")
            return redirect(url_for("bahan_baku_detail", bahan_id=bahan.id))
        bahan.nama_bahan = nama_bahan
        bahan.satuan = request.form.get("satuan", "Yard").strip() or "Yard"
        bahan.harga_per_yard = round(parse_angka_iklan(request.form.get("harga_per_yard")))
        bahan.catatan = request.form.get("catatan", "").strip()
        db.session.commit()
        flash("Data bahan baku berhasil diperbarui.", "success")
        return redirect(url_for("bahan_baku_detail", bahan_id=bahan.id))

    @app.route("/inventory/bahan-baku/<int:bahan_id>/hapus", methods=["POST"])
    @admin_required
    def bahan_baku_hapus(bahan_id):
        bahan = db.session.get(BahanBaku, bahan_id) or abort_404()
        nama = bahan.nama_bahan
        db.session.delete(bahan)
        db.session.commit()
        flash(f"Bahan baku {nama} beserta seluruh riwayatnya dihapus.", "info")
        return redirect(url_for("bahan_baku_list"))

    @app.route("/inventory/bahan-baku/<int:bahan_id>/kebutuhan/tambah", methods=["POST"])
    @admin_required
    def bahan_baku_kebutuhan_tambah(bahan_id):
        bahan = db.session.get(BahanBaku, bahan_id) or abort_404()
        # Formnya juga dipakai dari halaman Cutting & Produksi (bukan cuma detail Bahan
        # Baku), jadi redirect-nya ikutin field "next" kalau dikirim.
        tujuan = request.form.get("next") or url_for("bahan_baku_detail", bahan_id=bahan.id)
        produk_id = request.form.get("produk_id", type=int)
        jumlah_yard = parse_angka_iklan(request.form.get("jumlah_yard"))
        produk = db.session.get(Produk, produk_id) if produk_id else None
        if not produk or jumlah_yard <= 0:
            flash("Pilih produk dan isi jumlah yard per produk (harus lebih dari 0).", "danger")
            return redirect(tujuan)
        sudah_ada = BahanBakuKebutuhan.query.filter_by(bahan_baku_id=bahan.id, produk_id=produk.id).first()
        if sudah_ada:
            sudah_ada.jumlah_yard = jumlah_yard
            pesan = f"Kebutuhan {bahan.nama_bahan} untuk {produk.nama_produk} diperbarui."
        else:
            db.session.add(BahanBakuKebutuhan(bahan_baku_id=bahan.id, produk_id=produk.id, jumlah_yard=jumlah_yard))
            pesan = f"{produk.nama_produk} butuh {jumlah_yard} {bahan.satuan} {bahan.nama_bahan} berhasil dicatat."

        # Begitu Kebutuhan-nya diisi/diubah, langsung hitung otomatis Produk Jadi
        # untuk transaksi Keluar yang cocok tapi masih kosong (misalnya transaksi
        # lama yang dicatat sebelum Kebutuhan per produk ini pernah diisi).
        transaksi_kosong = BahanBakuTransaksi.query.filter_by(
            bahan_baku_id=bahan.id, produk_id=produk.id, jenis="Keluar", produk_jadi_pcs=None,
        ).all()
        for t in transaksi_kosong:
            t.produk_jadi_pcs = round(t.jumlah_yard / jumlah_yard)
        if transaksi_kosong:
            pesan += f" {len(transaksi_kosong)} transaksi Keluar lama otomatis terisi Produk Jadi-nya."

        db.session.commit()
        flash(pesan, "success")
        return redirect(tujuan)

    @app.route("/inventory/bahan-baku/kebutuhan/<int:kebutuhan_id>/hapus", methods=["POST"])
    @admin_required
    def bahan_baku_kebutuhan_hapus(kebutuhan_id):
        k = db.session.get(BahanBakuKebutuhan, kebutuhan_id) or abort_404()
        bahan_id = k.bahan_baku_id
        tujuan = request.form.get("next") or url_for("bahan_baku_detail", bahan_id=bahan_id)
        db.session.delete(k)
        db.session.commit()
        flash("Kebutuhan bahan baku untuk produk itu dihapus.", "info")
        return redirect(tujuan)

    @app.route("/inventory/bahan-baku/<int:bahan_id>/transaksi/tambah", methods=["POST"])
    @admin_required
    def bahan_baku_transaksi_tambah(bahan_id):
        bahan = db.session.get(BahanBaku, bahan_id) or abort_404()
        jenis = request.form.get("jenis")
        if jenis not in ("Masuk", "Keluar"):
            flash("Jenis transaksi tidak valid.", "danger")
            return redirect(url_for("bahan_baku_detail", bahan_id=bahan.id))
        try:
            tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
        except ValueError:
            tanggal = today_wib()
        produk_id = request.form.get("produk_id", type=int)
        produk = db.session.get(Produk, produk_id) if produk_id else None
        jumlah_unit = parse_angka_iklan(request.form.get("jumlah_unit"))
        jumlah_yard_manual = parse_angka_iklan(request.form.get("jumlah_yard"))
        keterangan = request.form.get("keterangan", "").strip()

        jumlah_yard = jumlah_yard_manual
        # Kalau admin pilih produk & isi jumlah unit diproduksi (bukan langsung isi
        # jumlah yard), otomatis hitung dari kebutuhan yard/produk yang sudah dicatat --
        # jumlah yard yang ditulis manual tetap menang kalau diisi.
        if jenis == "Keluar" and produk and jumlah_unit > 0 and not jumlah_yard_manual:
            kebutuhan = BahanBakuKebutuhan.query.filter_by(bahan_baku_id=bahan.id, produk_id=produk.id).first()
            if kebutuhan:
                jumlah_yard = jumlah_unit * kebutuhan.jumlah_yard

        if jumlah_yard <= 0:
            flash(
                "Isi jumlah yard secara langsung, atau pilih produk + jumlah unit diproduksi "
                "(kalau kebutuhan yard/produk sudah dicatat).",
                "danger",
            )
            return redirect(url_for("bahan_baku_detail", bahan_id=bahan.id))

        db.session.add(BahanBakuTransaksi(
            bahan_baku_id=bahan.id, tanggal=tanggal, jenis=jenis, jumlah_yard=jumlah_yard,
            produk_id=produk.id if produk else None, keterangan=keterangan,
        ))
        bahan.stok_saat_ini = (bahan.stok_saat_ini or 0) + (jumlah_yard if jenis == "Masuk" else -jumlah_yard)
        db.session.commit()

        pesan = f"{jenis} {jumlah_yard:g} {bahan.satuan} {bahan.nama_bahan} berhasil dicatat."
        kategori = "success"
        if bahan.stok_saat_ini < 0:
            pesan += " Perhatian: stok sekarang minus, cek lagi catatan stok masuknya."
            kategori = "warning"
        flash(pesan, kategori)
        return redirect(url_for("bahan_baku_detail", bahan_id=bahan.id))

    @app.route("/inventory/bahan-baku/transaksi/<int:transaksi_id>/hapus", methods=["POST"])
    @admin_required
    def bahan_baku_transaksi_hapus(transaksi_id):
        t = db.session.get(BahanBakuTransaksi, transaksi_id) or abort_404()
        bahan = t.bahan_baku
        bahan.stok_saat_ini = (bahan.stok_saat_ini or 0) - (t.jumlah_yard if t.jenis == "Masuk" else -t.jumlah_yard)
        db.session.delete(t)
        db.session.commit()
        flash("Transaksi dihapus, stok disesuaikan kembali.", "info")
        return redirect(url_for("bahan_baku_detail", bahan_id=bahan.id))

    @app.route("/inventory/cutting/transaksi/<int:transaksi_id>/produk-jadi", methods=["POST"])
    @admin_required
    def bahan_baku_transaksi_isi_produk_jadi(transaksi_id):
        """Isi/koreksi Produk Jadi (pcs) langsung dari baris Riwayat Cutting di halaman
        Cutting & Produksi -- dipakai buat backfill transaksi lama yang belum kehitung
        otomatis (mis. sebelum Kebutuhan per Produk diisi)."""
        t = db.session.get(BahanBakuTransaksi, transaksi_id) or abort_404()
        pcs = request.form.get("produk_jadi_pcs", type=int)
        if pcs is None or pcs < 0:
            flash("Isi jumlah Produk Jadi dengan angka yang valid.", "danger")
            return redirect(url_for("bahan_baku_cutting"))
        t.produk_jadi_pcs = pcs
        db.session.commit()
        flash(f"Produk Jadi {t.bahan_baku.nama_bahan} diperbarui jadi {pcs} pcs.", "success")
        return redirect(url_for("bahan_baku_cutting"))

    @app.route("/inventory/transaksi")
    @admin_required
    def bahan_baku_transaksi_riwayat():
        """Kartu stok gabungan: semua pergerakan bahan baku (Masuk dari Pengadaan +
        Keluar dari Cutting & Produksi) dalam satu tabel, bisa difilter per bahan & jenis --
        sebelumnya riwayatnya kepisah di 2 halaman berbeda, susah lihat histori satu bahan
        dari awal sampai akhir."""
        bahan_id = request.args.get("bahan_id", type=int)
        jenis = request.args.get("jenis")
        if jenis not in ("Masuk", "Keluar"):
            jenis = None

        query = BahanBakuTransaksi.query
        if bahan_id:
            query = query.filter_by(bahan_baku_id=bahan_id)
        if jenis:
            query = query.filter_by(jenis=jenis)
        transaksi = query.order_by(BahanBakuTransaksi.tanggal.desc(), BahanBakuTransaksi.id.desc()).limit(300).all()

        total_masuk_yard = sum(t.jumlah_yard for t in transaksi if t.jenis == "Masuk")
        total_keluar_yard = sum(t.jumlah_yard for t in transaksi if t.jenis == "Keluar")
        total_dibayar = sum(t.total_dibayar or 0 for t in transaksi if t.jenis == "Masuk")

        daftar_bahan = BahanBaku.query.order_by(BahanBaku.nama_bahan).all()
        return render_template(
            "inventory/bahan_baku_transaksi_riwayat.html",
            transaksi=transaksi, daftar_bahan=daftar_bahan,
            bahan_id_filter=bahan_id, jenis_filter=jenis,
            total_masuk_yard=total_masuk_yard, total_keluar_yard=total_keluar_yard,
            total_dibayar=total_dibayar,
        )

    # ---------- INVENTORY: MASTER DATA (VENDOR / GUDANG / AKUN PEMBAYARAN) ----------
    @app.route("/inventory/master-data/vendor", methods=["GET", "POST"])
    @admin_required
    def vendor_list():
        if request.method == "POST":
            nama_vendor = request.form.get("nama_vendor", "").strip()
            if not nama_vendor:
                flash("Nama vendor wajib diisi.", "danger")
                return redirect(url_for("vendor_list"))
            db.session.add(Vendor(
                nama_vendor=nama_vendor,
                jenis=request.form.get("jenis", "Keduanya"),
                kontak=request.form.get("kontak", "").strip(),
                alamat=request.form.get("alamat", "").strip(),
                catatan=request.form.get("catatan", "").strip(),
            ))
            db.session.commit()
            flash(f"Vendor {nama_vendor} berhasil ditambahkan.", "success")
            return redirect(url_for("vendor_list"))
        daftar = Vendor.query.order_by(Vendor.nama_vendor).all()
        return render_template("inventory/vendor_list.html", daftar=daftar)

    @app.route("/inventory/master-data/vendor/<int:vendor_id>/edit", methods=["POST"])
    @admin_required
    def vendor_edit(vendor_id):
        v = db.session.get(Vendor, vendor_id) or abort_404()
        nama_vendor = request.form.get("nama_vendor", "").strip()
        if not nama_vendor:
            flash("Nama vendor wajib diisi.", "danger")
            return redirect(url_for("vendor_list"))
        v.nama_vendor = nama_vendor
        v.jenis = request.form.get("jenis", "Keduanya")
        v.kontak = request.form.get("kontak", "").strip()
        v.alamat = request.form.get("alamat", "").strip()
        v.catatan = request.form.get("catatan", "").strip()
        db.session.commit()
        flash("Data vendor berhasil diperbarui.", "success")
        return redirect(url_for("vendor_list"))

    @app.route("/inventory/master-data/vendor/<int:vendor_id>/hapus", methods=["POST"])
    @admin_required
    def vendor_hapus(vendor_id):
        v = db.session.get(Vendor, vendor_id) or abort_404()
        nama = v.nama_vendor
        db.session.delete(v)
        db.session.commit()
        flash(f"Vendor {nama} dihapus.", "info")
        return redirect(url_for("vendor_list"))

    @app.route("/inventory/master-data/gudang", methods=["GET", "POST"])
    @admin_required
    def gudang_list():
        if request.method == "POST":
            nama_gudang = request.form.get("nama_gudang", "").strip()
            if not nama_gudang:
                flash("Nama gudang wajib diisi.", "danger")
                return redirect(url_for("gudang_list"))
            db.session.add(Gudang(
                nama_gudang=nama_gudang,
                alamat=request.form.get("alamat", "").strip(),
                pic=request.form.get("pic", "").strip(),
                kontak=request.form.get("kontak", "").strip(),
                catatan=request.form.get("catatan", "").strip(),
            ))
            db.session.commit()
            flash(f"Gudang {nama_gudang} berhasil ditambahkan.", "success")
            return redirect(url_for("gudang_list"))
        daftar = Gudang.query.order_by(Gudang.nama_gudang).all()
        return render_template("inventory/gudang_list.html", daftar=daftar)

    @app.route("/inventory/master-data/gudang/<int:gudang_id>/edit", methods=["POST"])
    @admin_required
    def gudang_edit(gudang_id):
        g = db.session.get(Gudang, gudang_id) or abort_404()
        nama_gudang = request.form.get("nama_gudang", "").strip()
        if not nama_gudang:
            flash("Nama gudang wajib diisi.", "danger")
            return redirect(url_for("gudang_list"))
        g.nama_gudang = nama_gudang
        g.alamat = request.form.get("alamat", "").strip()
        g.pic = request.form.get("pic", "").strip()
        g.kontak = request.form.get("kontak", "").strip()
        g.catatan = request.form.get("catatan", "").strip()
        db.session.commit()
        flash("Data gudang berhasil diperbarui.", "success")
        return redirect(url_for("gudang_list"))

    @app.route("/inventory/master-data/gudang/<int:gudang_id>/hapus", methods=["POST"])
    @admin_required
    def gudang_hapus(gudang_id):
        g = db.session.get(Gudang, gudang_id) or abort_404()
        nama = g.nama_gudang
        db.session.delete(g)
        db.session.commit()
        flash(f"Gudang {nama} dihapus.", "info")
        return redirect(url_for("gudang_list"))

    @app.route("/inventory/master-data/akun-pembayaran", methods=["GET", "POST"])
    @admin_required
    def akun_pembayaran_list():
        if request.method == "POST":
            nama_akun = request.form.get("nama_akun", "").strip()
            if not nama_akun:
                flash("Nama akun wajib diisi.", "danger")
                return redirect(url_for("akun_pembayaran_list"))
            db.session.add(AkunPembayaran(
                nama_akun=nama_akun,
                jenis=request.form.get("jenis", "Bank"),
                nomor_rekening=request.form.get("nomor_rekening", "").strip(),
                catatan=request.form.get("catatan", "").strip(),
            ))
            db.session.commit()
            flash(f"Akun {nama_akun} berhasil ditambahkan.", "success")
            return redirect(url_for("akun_pembayaran_list"))
        daftar = AkunPembayaran.query.order_by(AkunPembayaran.nama_akun).all()
        return render_template("inventory/akun_pembayaran_list.html", daftar=daftar)

    @app.route("/inventory/master-data/akun-pembayaran/<int:akun_id>/edit", methods=["POST"])
    @admin_required
    def akun_pembayaran_edit(akun_id):
        a = db.session.get(AkunPembayaran, akun_id) or abort_404()
        nama_akun = request.form.get("nama_akun", "").strip()
        if not nama_akun:
            flash("Nama akun wajib diisi.", "danger")
            return redirect(url_for("akun_pembayaran_list"))
        a.nama_akun = nama_akun
        a.jenis = request.form.get("jenis", "Bank")
        a.nomor_rekening = request.form.get("nomor_rekening", "").strip()
        a.catatan = request.form.get("catatan", "").strip()
        db.session.commit()
        flash("Data akun berhasil diperbarui.", "success")
        return redirect(url_for("akun_pembayaran_list"))

    @app.route("/inventory/master-data/akun-pembayaran/<int:akun_id>/hapus", methods=["POST"])
    @admin_required
    def akun_pembayaran_hapus(akun_id):
        a = db.session.get(AkunPembayaran, akun_id) or abort_404()
        nama = a.nama_akun
        db.session.delete(a)
        db.session.commit()
        flash(f"Akun {nama} dihapus.", "info")
        return redirect(url_for("akun_pembayaran_list"))

    @app.route("/inventory/master-data/purchase-order", methods=["GET", "POST"])
    @admin_required
    def purchase_order_list():
        if request.method == "POST":
            vendor_id = request.form.get("vendor_id", type=int)
            vendor = db.session.get(Vendor, vendor_id) if vendor_id else None
            try:
                tanggal_order = datetime.strptime(request.form.get("tanggal_order", ""), "%Y-%m-%d").date()
            except ValueError:
                tanggal_order = None
            estimasi_selesai = None
            if request.form.get("estimasi_selesai"):
                try:
                    estimasi_selesai = datetime.strptime(request.form.get("estimasi_selesai"), "%Y-%m-%d").date()
                except ValueError:
                    estimasi_selesai = None

            if not vendor or not tanggal_order:
                flash("Vendor dan Tgl Order wajib diisi.", "danger")
                return redirect(url_for("purchase_order_list"))

            # -- baris Item Produk (opsional, boleh kosong semua) --
            item_produk_rows = []
            for produk_id_s, warna, size, qty_s in zip(
                request.form.getlist("ip_produk_id[]"), request.form.getlist("ip_warna[]"),
                request.form.getlist("ip_size[]"), request.form.getlist("ip_qty[]"),
            ):
                produk = db.session.get(Produk, int(produk_id_s)) if produk_id_s else None
                qty = parse_angka_iklan(qty_s)
                if produk and qty > 0:
                    item_produk_rows.append((produk, warna.strip(), size.strip(), int(qty)))

            # -- baris Pemakaian Bahan (wajib minimal 1) --
            bahan_pakai_rows = []
            for bahan_id_s, qty_s in zip(
                request.form.getlist("bp_bahan_id[]"), request.form.getlist("bp_qty[]"),
            ):
                bahan = db.session.get(BahanBaku, int(bahan_id_s)) if bahan_id_s else None
                qty = parse_angka_iklan(qty_s)
                if bahan and qty > 0:
                    bahan_pakai_rows.append((bahan, qty))

            if not bahan_pakai_rows:
                flash("Isi minimal 1 baris Pemakaian Bahan (bahan + qty pakai) sebelum disimpan.", "danger")
                return redirect(url_for("purchase_order_list"))

            total_biaya = sum(qty * (produk.modal or 0) for produk, _, _, qty in item_produk_rows)
            po = PurchaseOrder(
                # Nomor PO final (format dd/mm/yy/nomor-urut) baru bisa dipastikan setelah
                # baris ini punya id asli -- placeholder unik sementara di sini, ditimpa di
                # bawah sebelum commit. Pakai id (bukan hitungan jumlah PO) supaya nggak
                # bentrok kalau ada PO lama yang sudah dihapus.
                nomor_po=f"TEMP-{uuid.uuid4().hex}", vendor_id=vendor.id, tanggal_order=tanggal_order,
                estimasi_selesai=estimasi_selesai, total_biaya=total_biaya,
            )
            db.session.add(po)
            db.session.flush()
            nomor_po = f"{tanggal_order.strftime('%d/%m/%y')}/{po.id}"
            po.nomor_po = nomor_po

            for produk, warna, size, qty in item_produk_rows:
                db.session.add(PurchaseOrderItemProduk(
                    purchase_order_id=po.id, produk_id=produk.id, warna=warna, size=size,
                    qty=qty, total=qty * (produk.modal or 0),
                ))

            stok_minus = False
            for bahan, qty in bahan_pakai_rows:
                db.session.add(PurchaseOrderBahanPakai(
                    purchase_order_id=po.id, bahan_baku_id=bahan.id, qty_pakai=qty, satuan=bahan.satuan,
                ))
                bahan.stok_saat_ini = (bahan.stok_saat_ini or 0) - qty
                if bahan.stok_saat_ini < 0:
                    stok_minus = True
                db.session.add(BahanBakuTransaksi(
                    bahan_baku_id=bahan.id, tanggal=tanggal_order, jenis="Keluar", jumlah_yard=qty,
                    vendor=vendor.nama_vendor, keterangan=f"Purchase Order {nomor_po}",
                    purchase_order_id=po.id,
                ))

            db.session.commit()
            pesan = f"Purchase Order {nomor_po} berhasil disimpan, stok bahan sudah dikurangi."
            kategori = "success"
            if stok_minus:
                pesan += " Perhatian: ada stok bahan yang jadi minus, cek lagi catatan stok masuknya."
                kategori = "warning"
            flash(pesan, kategori)
            return redirect(url_for("purchase_order_list"))

        q = request.args.get("q", "").strip()
        query = PurchaseOrder.query
        if q:
            query = query.filter(PurchaseOrder.nomor_po.ilike(f"%{q}%"))
        daftar = query.order_by(PurchaseOrder.tanggal_order.desc(), PurchaseOrder.id.desc()).all()
        daftar_vendor = Vendor.query.order_by(Vendor.nama_vendor).all()
        daftar_produk = Produk.query.order_by(Produk.nama_produk).all()
        daftar_bahan = BahanBaku.query.order_by(BahanBaku.nama_bahan).all()
        daftar_akun = AkunPembayaran.query.order_by(AkunPembayaran.nama_akun).all()
        produk_json = {p.id: {"nama": p.nama_produk, "modal": p.modal or 0} for p in daftar_produk}
        bahan_json = {
            b.id: {"nama": b.nama_bahan, "stok": b.stok_saat_ini or 0, "satuan": b.satuan} for b in daftar_bahan
        }
        return render_template(
            "inventory/purchase_order_list.html",
            daftar=daftar, daftar_vendor=daftar_vendor, daftar_produk=daftar_produk,
            daftar_bahan=daftar_bahan, daftar_akun=daftar_akun, produk_json=produk_json, bahan_json=bahan_json,
            tanggal_hari_ini=today_wib().isoformat(), q=q,
            status_pilihan=["Menunggu Produksi", "Diproses", "Selesai Produksi", "Dibatalkan"],
        )

    @app.route("/inventory/master-data/purchase-order/<int:po_id>")
    @admin_required
    def purchase_order_detail(po_id):
        po = db.session.get(PurchaseOrder, po_id) or abort_404()
        return render_template("inventory/purchase_order_detail.html", po=po)

    @app.route("/inventory/master-data/purchase-order/<int:po_id>/pembayaran")
    @admin_required
    def purchase_order_pembayaran(po_id):
        po = db.session.get(PurchaseOrder, po_id) or abort_404()
        daftar_akun = AkunPembayaran.query.order_by(AkunPembayaran.nama_akun).all()
        return render_template(
            "inventory/purchase_order_pembayaran.html",
            po=po, daftar_akun=daftar_akun, tanggal_hari_ini=today_wib().isoformat(),
        )

    @app.route("/inventory/master-data/purchase-order/<int:po_id>/hapus", methods=["POST"])
    @admin_required
    def purchase_order_hapus(po_id):
        po = db.session.get(PurchaseOrder, po_id) or abort_404()
        nomor_po = po.nomor_po
        # Kembalikan stok bahan yg kepotong PO ini, & hapus transaksi Keluar yg
        # otomatis dibuat bareng PO -- biar catatan kartu stok gak nyisain data yatim.
        for bp in po.bahan_pakai_list:
            bp.bahan_baku.stok_saat_ini = (bp.bahan_baku.stok_saat_ini or 0) + bp.qty_pakai
        BahanBakuTransaksi.query.filter_by(purchase_order_id=po.id).delete()
        db.session.delete(po)
        db.session.commit()
        flash(f"Purchase Order {nomor_po} dihapus, stok bahan yg kepotong sudah dikembalikan.", "info")
        return redirect(url_for("purchase_order_list"))

    @app.route("/inventory/master-data/purchase-order/<int:po_id>/status", methods=["POST"])
    @admin_required
    def purchase_order_status_update(po_id):
        po = db.session.get(PurchaseOrder, po_id) or abort_404()
        status = request.form.get("status", "")
        if status not in ("Menunggu Produksi", "Diproses", "Selesai Produksi", "Dibatalkan"):
            flash("Status tidak valid.", "danger")
            return redirect(url_for("purchase_order_list"))
        po.status = status
        db.session.commit()
        flash(f"Status PO {po.nomor_po} diperbarui jadi {status}.", "success")
        return redirect(url_for("purchase_order_list"))

    @app.route("/inventory/master-data/purchase-order/<int:po_id>/pembayaran/tambah", methods=["POST"])
    @admin_required
    def purchase_order_pembayaran_tambah(po_id):
        po = db.session.get(PurchaseOrder, po_id) or abort_404()
        jumlah = round(parse_angka_iklan(request.form.get("jumlah")))
        try:
            tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
        except ValueError:
            tanggal = today_wib()
        metode = request.form.get("metode", "")
        if metode not in ("Kasbon", "Cicilan", "Pelunasan"):
            metode = "Cicilan"
        akun_id = request.form.get("akun_pembayaran_id", type=int)
        akun = db.session.get(AkunPembayaran, akun_id) if akun_id else None
        catatan = request.form.get("catatan", "").strip()

        if jumlah <= 0:
            flash("Isi jumlah pembayaran (harus lebih dari 0).", "danger")
            return redirect(url_for("purchase_order_pembayaran", po_id=po.id))

        db.session.add(PurchaseOrderPembayaran(
            purchase_order_id=po.id, tanggal=tanggal, jumlah=jumlah, metode=metode,
            akun_pembayaran_id=akun.id if akun else None, catatan=catatan,
        ))
        db.session.commit()
        sisa = po.sisa_bayar
        flash(
            f"{metode} Rp {jumlah:,.0f}".replace(",", ".") + f" utk PO {po.nomor_po} dicatat. "
            + (f"Sisa Rp {sisa:,.0f}".replace(",", ".") if sisa > 0 else "PO ini sudah lunas."),
            "success",
        )
        return redirect(url_for("purchase_order_pembayaran", po_id=po.id))

    @app.route("/inventory/master-data/purchase-order/pembayaran/<int:pembayaran_id>/hapus", methods=["POST"])
    @admin_required
    def purchase_order_pembayaran_hapus(pembayaran_id):
        p = db.session.get(PurchaseOrderPembayaran, pembayaran_id) or abort_404()
        po_id = p.purchase_order_id
        db.session.delete(p)
        db.session.commit()
        flash("Catatan pembayaran dihapus.", "info")
        return redirect(url_for("purchase_order_pembayaran", po_id=po_id))

    @app.route("/inventory/bahan-baku/input", methods=["GET", "POST"])
    @admin_required
    def bahan_baku_input():
        if request.method == "POST":
            jenis_bahan = request.form.get("jenis_bahan", "").strip()
            warna = request.form.get("warna", "").strip()
            panjang_yard = parse_angka_iklan(request.form.get("panjang_yard"))
            tinggi_meter = parse_angka_iklan(request.form.get("tinggi_meter")) or None
            suplier = request.form.get("suplier", "").strip()
            try:
                tanggal_datang = datetime.strptime(request.form.get("tanggal_datang", ""), "%Y-%m-%d").date()
            except ValueError:
                tanggal_datang = today_wib()
            produk_id = request.form.get("produk_id", type=int)
            harga_per_yard = round(parse_angka_iklan(request.form.get("harga_per_yard")))

            if not jenis_bahan or panjang_yard <= 0:
                flash("Jenis Bahan dan Panjang Bahan (yard) wajib diisi.", "danger")
                return redirect(url_for("bahan_baku_input"))

            # Digabung jadi 1 nama_bahan per kombinasi jenis+warna, biar stok per warna
            # ketahuan terpisah walau suplier-nya beda-beda tiap kali beli.
            nama_gabungan = f"{jenis_bahan} - {warna}" if warna else jenis_bahan
            bahan = BahanBaku.query.filter_by(nama_bahan=nama_gabungan).first()
            if not bahan:
                bahan = BahanBaku(nama_bahan=nama_gabungan, satuan="Yard", stok_saat_ini=0)
                db.session.add(bahan)
            bahan.warna = warna or bahan.warna
            bahan.tinggi_meter = tinggi_meter or bahan.tinggi_meter
            bahan.suplier = suplier or bahan.suplier
            if harga_per_yard:
                bahan.harga_per_yard = harga_per_yard
            bahan.stok_saat_ini = (bahan.stok_saat_ini or 0) + panjang_yard
            db.session.flush()

            db.session.add(BahanBakuTransaksi(
                bahan_baku_id=bahan.id, tanggal=tanggal_datang, jenis="Masuk", jumlah_yard=panjang_yard,
                produk_id=produk_id or None, warna=warna, suplier=suplier, tinggi_meter=tinggi_meter,
                harga_per_yard=harga_per_yard, total_dibayar=round(panjang_yard * harga_per_yard),
            ))
            db.session.commit()
            flash(f"{panjang_yard:g} Yard {jenis_bahan} berhasil dicatat masuk.", "success")
            return redirect(url_for("bahan_baku_input"))

        produk_list = Produk.query.order_by(Produk.nama_produk).all()
        riwayat = (
            BahanBakuTransaksi.query.filter_by(jenis="Masuk")
            .join(BahanBaku)
            .order_by(BahanBakuTransaksi.tanggal.desc(), BahanBakuTransaksi.id.desc())
            .limit(50).all()
        )
        total_yard = sum(b.stok_saat_ini or 0 for b in BahanBaku.query.all())
        total_dibayar_semua = sum(
            t.total_dibayar or 0 for t in BahanBakuTransaksi.query.filter_by(jenis="Masuk").all()
        )
        return render_template(
            "inventory/bahan_baku_input.html",
            produk_list=produk_list, riwayat=riwayat, total_yard=total_yard,
            total_dibayar_semua=total_dibayar_semua, tanggal_hari_ini=today_wib().isoformat(),
        )

    @app.route("/inventory/cutting", methods=["GET", "POST"])
    @admin_required
    def bahan_baku_cutting():
        if request.method == "POST":
            bahan_id = request.form.get("bahan_baku_id", type=int)
            bahan = db.session.get(BahanBaku, bahan_id) if bahan_id else None
            panjang_yard = parse_angka_iklan(request.form.get("panjang_yard"))
            lebar_kain = parse_angka_iklan(request.form.get("lebar_kain")) or None
            produk_id = request.form.get("produk_id", type=int)
            try:
                tanggal_ambil = datetime.strptime(request.form.get("tanggal_ambil", ""), "%Y-%m-%d").date()
            except ValueError:
                tanggal_ambil = today_wib()
            vendor = request.form.get("vendor", "").strip()
            produk_jadi_pcs = request.form.get("produk_jadi_pcs", type=int)

            if not bahan or panjang_yard <= 0:
                flash("Pilih bahan baku dan isi panjang bahan yang diambil (yard).", "danger")
                return redirect(url_for("bahan_baku_cutting"))

            db.session.add(BahanBakuTransaksi(
                bahan_baku_id=bahan.id, tanggal=tanggal_ambil, jenis="Keluar", jumlah_yard=panjang_yard,
                produk_id=produk_id or None, warna=bahan.warna, vendor=vendor, lebar_kain=lebar_kain,
                produk_jadi_pcs=produk_jadi_pcs,
            ))
            bahan.stok_saat_ini = (bahan.stok_saat_ini or 0) - panjang_yard
            db.session.commit()

            pesan = f"{panjang_yard:g} Yard {bahan.nama_bahan} berhasil dicatat keluar utk cutting."
            kategori = "success"
            if bahan.stok_saat_ini < 0:
                pesan += " Perhatian: stok sekarang minus, cek lagi catatan stok masuknya."
                kategori = "warning"
            flash(pesan, kategori)
            return redirect(url_for("bahan_baku_cutting"))

        bahan_list = BahanBaku.query.order_by(BahanBaku.nama_bahan).all()
        produk_list = Produk.query.order_by(Produk.nama_produk).all()
        spek_per_produk = {
            p.id: [
                {
                    "id": s.id, "size": s.size, "kategori": s.kategori or "",
                    "lingkar_dada": s.lingkar_dada, "panjang_atas": s.panjang_atas,
                    "lingkar_pinggang": s.lingkar_pinggang, "ld_lengan": s.ld_lengan, "pergelangan": s.pergelangan,
                }
                for s in p.spek_ukuran_list
            ]
            for p in produk_list
        }
        # Kelompokkan semua baris spek ukuran per kategori (bukan per produk) buat
        # ditampilkan sbg tabel terpisah-pisah di "Kelola Spek Ukuran per Produk" --
        # tiap kategori kolomnya beda (lihat KATEGORI_SPEK_FIELDS).
        spek_by_kategori = {}
        for p in produk_list:
            for s in p.spek_ukuran_list:
                spek_by_kategori.setdefault(s.kategori or "", []).append(s)
        # "bahanId_produkId" -> yard dibutuhkan per 1 pcs (dari menu Bahan Baku > Kebutuhan
        # per Produk) -- dipakai JS di halaman ini utk otomatis hitung Estimasi Produk Jadi
        # begitu Jenis Bahan + Peruntukan Produk + Panjang Bahan sudah diisi.
        daftar_kebutuhan = BahanBakuKebutuhan.query.join(BahanBaku).order_by(BahanBaku.nama_bahan).all()
        kebutuhan_map = {
            f"{k.bahan_baku_id}_{k.produk_id}": k.jumlah_yard
            for k in daftar_kebutuhan
        }
        riwayat = (
            BahanBakuTransaksi.query.filter_by(jenis="Keluar")
            .join(BahanBaku)
            .order_by(BahanBakuTransaksi.tanggal.desc(), BahanBakuTransaksi.id.desc())
            .limit(50).all()
        )
        return render_template(
            "inventory/bahan_baku_cutting.html",
            bahan_list=bahan_list, produk_list=produk_list, riwayat=riwayat,
            daftar_kebutuhan=daftar_kebutuhan,
            spek_per_produk_json=spek_per_produk, kebutuhan_map_json=kebutuhan_map,
            spek_by_kategori=spek_by_kategori, kategori_fields=KATEGORI_SPEK_FIELDS,
            tanggal_hari_ini=today_wib().isoformat(),
        )

    @app.route("/inventory/spek-ukuran/<int:produk_id>/tambah", methods=["POST"])
    @admin_required
    def produk_spek_ukuran_tambah(produk_id):
        produk = db.session.get(Produk, produk_id) or abort_404()
        size = request.form.get("size", "").strip() or "All Size"
        kategori = request.form.get("kategori", "").strip()
        if kategori not in KATEGORI_SPEK_FIELDS or not kategori:
            flash("Pilih kategori produk dulu (Cadar/Khimar/Pashmina/dsb).", "danger")
            return redirect(url_for("bahan_baku_cutting"))
        db.session.add(ProdukSpekUkuran(
            produk_id=produk.id, size=size, kategori=kategori,
            lingkar_dada=parse_angka_iklan(request.form.get("lingkar_dada")) or None,
            panjang_atas=parse_angka_iklan(request.form.get("panjang_atas")) or None,
            lingkar_pinggang=parse_angka_iklan(request.form.get("lingkar_pinggang")) or None,
            ld_lengan=parse_angka_iklan(request.form.get("ld_lengan")) or None,
            pergelangan=parse_angka_iklan(request.form.get("pergelangan")) or None,
        ))
        db.session.commit()
        flash(f"Spek ukuran {kategori} - {size} untuk {produk.nama_produk} disimpan.", "success")
        return redirect(url_for("bahan_baku_cutting"))

    @app.route("/inventory/spek-ukuran/<int:spek_id>/hapus", methods=["POST"])
    @admin_required
    def produk_spek_ukuran_hapus(spek_id):
        s = db.session.get(ProdukSpekUkuran, spek_id) or abort_404()
        db.session.delete(s)
        db.session.commit()
        flash("Baris spek ukuran dihapus.", "info")
        return redirect(url_for("bahan_baku_cutting"))

    # ---------- ABSENSI ----------
    @app.route("/absensi")
    @admin_required
    def absensi_list():
        tanggal_str = request.args.get("tanggal", today_wib().isoformat())
        try:
            tanggal = datetime.strptime(tanggal_str, "%Y-%m-%d").date()
        except ValueError:
            tanggal = today_wib()

        data = (
            Attendance.query.filter_by(tanggal=tanggal)
            .join(Employee)
            .order_by(Employee.nama)
            .all()
        )
        sudah_absen_ids = {a.employee_id for a in data}
        belum_absen = Employee.query.filter(
            Employee.status == "Aktif", ~Employee.id.in_(sudah_absen_ids) if sudah_absen_ids else True
        ).order_by(Employee.nama).all()

        return render_template(
            "attendance_list.html",
            data=data,
            tanggal=tanggal,
            belum_absen=belum_absen,
        )

    @app.route("/absensi/koreksi", methods=["GET", "POST"], defaults={"att_id": None})
    @app.route("/absensi/<int:att_id>/koreksi", methods=["GET", "POST"])
    @admin_required
    def absensi_koreksi(att_id):
        settings = get_settings()
        att = db.session.get(Attendance, att_id) if att_id else None

        if request.method == "POST":
            employee_id = int(request.form["employee_id"])
            tanggal = datetime.strptime(request.form["tanggal"], "%Y-%m-%d").date()
            status = request.form["status"]
            jam_masuk = request.form.get("jam_masuk") or None
            jam_pulang = request.form.get("jam_pulang") or None
            catatan = request.form.get("catatan", "").strip()

            telat, lembur = (0, 0)
            if status == "Hadir":
                emp_koreksi = db.session.get(Employee, employee_id)
                tipe_koreksi = emp_koreksi.tipe_pegawai if emp_koreksi else "Karyawan Tetap"
                telat, lembur = hitung_telat_lembur(
                    jam_masuk, jam_pulang, settings, tipe_koreksi, tanggal=tanggal, employee_id=employee_id
                )
            else:
                jam_masuk, jam_pulang = None, None

            target = att or Attendance.query.filter_by(
                employee_id=employee_id, tanggal=tanggal
            ).first()
            if not target:
                target = Attendance(employee_id=employee_id, tanggal=tanggal)
                db.session.add(target)

            target.employee_id = employee_id
            target.tanggal = tanggal
            target.status = status
            target.jam_masuk = jam_masuk
            target.jam_pulang = jam_pulang
            target.telat_menit = telat
            target.lembur_menit = lembur
            target.catatan = catatan
            db.session.commit()
            flash("Data absensi berhasil disimpan.", "success")
            return redirect(url_for("absensi_list", tanggal=tanggal.isoformat()))

        employees = Employee.query.filter_by(status="Aktif").order_by(Employee.nama).all()
        tanggal_default = request.args.get("tanggal", today_wib().isoformat())
        return render_template(
            "attendance_koreksi.html",
            att=att,
            employees=employees,
            tanggal_default=tanggal_default,
        )

    @app.route("/absensi/<int:att_id>/hapus", methods=["POST"])
    @admin_required
    def absensi_hapus(att_id):
        att = db.session.get(Attendance, att_id) or abort_404()
        tanggal = att.tanggal
        db.session.delete(att)
        db.session.commit()
        flash("Data absensi dihapus.", "info")
        return redirect(url_for("absensi_list", tanggal=tanggal.isoformat()))

    # ---------- HARI LIBUR ----------
    @app.route("/hari-libur")
    @admin_required
    def hari_libur_list():
        tahun = int(request.args.get("tahun", today_wib().year))
        daftar = (
            HariLibur.query.filter(db.extract("year", HariLibur.tanggal) == tahun)
            .order_by(HariLibur.tanggal)
            .all()
        )
        return render_template("hari_libur_list.html", daftar=daftar, tahun=tahun)

    @app.route("/hari-libur/tambah", methods=["POST"])
    @admin_required
    def hari_libur_tambah():
        try:
            tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
        except ValueError:
            tanggal = None
        keterangan = request.form.get("keterangan", "").strip()

        if not tanggal or not keterangan:
            flash("Tanggal dan keterangan wajib diisi.", "danger")
            return redirect(url_for("hari_libur_list"))

        if HariLibur.query.filter_by(tanggal=tanggal).first():
            flash(f"Tanggal {tanggal.strftime('%d-%m-%Y')} sudah ada di daftar hari libur.", "danger")
            return redirect(url_for("hari_libur_list", tahun=tanggal.year))

        db.session.add(HariLibur(tanggal=tanggal, keterangan=keterangan))
        db.session.commit()
        flash(f"Hari libur {tanggal.strftime('%d-%m-%Y')} berhasil ditambahkan.", "success")
        return redirect(url_for("hari_libur_list", tahun=tanggal.year))

    @app.route("/hari-libur/<int:hl_id>/hapus", methods=["POST"])
    @admin_required
    def hari_libur_hapus(hl_id):
        hl = db.session.get(HariLibur, hl_id) or abort_404()
        tahun = hl.tanggal.year
        db.session.delete(hl)
        db.session.commit()
        flash("Hari libur dihapus.", "info")
        return redirect(url_for("hari_libur_list", tahun=tahun))

    # ---------- PENGAJUAN IZIN (ADMIN) ----------
    @app.route("/pengajuan-izin")
    @admin_required
    def pengajuan_izin_list():
        status_filter = request.args.get("status", "Menunggu")
        q = PengajuanIzin.query.join(Employee)
        if status_filter in ("Menunggu", "Disetujui", "Ditolak"):
            q = q.filter(PengajuanIzin.status == status_filter)
        pengajuan = q.order_by(PengajuanIzin.tanggal_diajukan.desc()).all()
        jumlah_menunggu = PengajuanIzin.query.filter_by(status="Menunggu").count()
        settings = get_settings()
        wa_links = {p.id: buat_wa_link_notifikasi_izin(p, settings) for p in pengajuan}
        wa_links_karyawan = {p.id: buat_wa_link_hasil_izin(p) for p in pengajuan}
        return render_template(
            "pengajuan_izin_list.html",
            pengajuan=pengajuan,
            status_filter=status_filter,
            jumlah_menunggu=jumlah_menunggu,
            wa_links=wa_links,
            wa_links_karyawan=wa_links_karyawan,
        )

    @app.route("/pengajuan-izin/<int:pid>/setujui", methods=["POST"])
    @admin_required
    def pengajuan_izin_setujui(pid):
        p = db.session.get(PengajuanIzin, pid) or abort_404()
        p.status = "Disetujui"
        p.tanggal_diproses = now_wib()

        att = Attendance.query.filter_by(employee_id=p.employee_id, tanggal=p.tanggal).first()
        if not att:
            att = Attendance(employee_id=p.employee_id, tanggal=p.tanggal)
            db.session.add(att)
        att.status = p.jenis
        att.jam_masuk = None
        att.jam_pulang = None
        att.telat_menit = 0
        att.lembur_menit = 0
        att.catatan = p.alasan

        db.session.commit()

        diperbarui = _reconcile_payroll_draft(p.employee, p.tanggal)
        pesan = f"Pengajuan {p.employee.nama} disetujui & tercatat di absensi."
        kategori = "success"
        if diperbarui is True:
            pesan += " Slip gaji bulan ini otomatis ikut diperbarui."
        elif diperbarui is False:
            pesan += (
                f" Perhatian: slip gaji {p.employee.nama} bulan ini sudah berstatus Dibayar, "
                "jadi perubahan ini TIDAK otomatis masuk ke situ -- sesuaikan manual kalau perlu."
            )
            kategori = "warning"
        flash(pesan, kategori)
        return redirect(url_for("pengajuan_izin_list"))

    @app.route("/pengajuan-izin/<int:pid>/tolak", methods=["POST"])
    @admin_required
    def pengajuan_izin_tolak(pid):
        p = db.session.get(PengajuanIzin, pid) or abort_404()
        p.status = "Ditolak"
        p.catatan_admin = request.form.get("catatan_admin", "").strip()
        p.tanggal_diproses = now_wib()
        db.session.commit()
        flash(f"Pengajuan {p.employee.nama} ditolak.", "info")
        return redirect(url_for("pengajuan_izin_list"))

    # ---------- AREA PEGAWAI ----------
    @app.route("/pegawai")
    @pegawai_required
    def pegawai_dashboard():
        hari_ini = today_wib()
        absen = Attendance.query.filter_by(employee_id=current_user.id, tanggal=hari_ini).first()
        settings = get_settings()

        absensi_bulan_ini = Attendance.query.filter(
            Attendance.employee_id == current_user.id,
            db.extract("month", Attendance.tanggal) == hari_ini.month,
            db.extract("year", Attendance.tanggal) == hari_ini.year,
        ).all()
        rekap_bulan_ini = {
            "hadir": sum(1 for a in absensi_bulan_ini if a.status == "Hadir"),
            "sakit": sum(1 for a in absensi_bulan_ini if a.status == "Sakit"),
            "izin": sum(1 for a in absensi_bulan_ini if a.status == "Izin"),
            "cuti": sum(1 for a in absensi_bulan_ini if a.status == "Cuti"),
            "alpha": sum(1 for a in absensi_bulan_ini if a.status == "Alpha"),
        }

        return render_template(
            "pegawai/dashboard.html",
            absen=absen,
            settings=settings,
            hari_ini=hari_ini,
            rekap_bulan_ini=rekap_bulan_ini,
            bulan_nama=BULAN_NAMA[hari_ini.month],
        )

    @app.route("/pegawai/absen", methods=["POST"])
    @pegawai_required
    def pegawai_absen():
        data = request.get_json(silent=True) or {}
        lat = data.get("lat")
        lng = data.get("lng")
        if lat is None or lng is None:
            return jsonify(success=False, message="Lokasi tidak terdeteksi. Aktifkan GPS/izin lokasi di HP Anda."), 400

        settings = get_settings()
        if settings.kantor_lat is not None and settings.kantor_lng is not None:
            jarak = haversine_meter(float(lat), float(lng), settings.kantor_lat, settings.kantor_lng)
            radius = settings.radius_kantor_meter or 100
            if jarak > radius:
                return jsonify(
                    success=False,
                    message=f"Anda berjarak {int(jarak)} meter dari kantor (maksimal {radius} meter). Absen ditolak.",
                ), 400

        hari_ini = today_wib()
        jam_sekarang = now_wib().strftime("%H:%M:%S")
        tipe_pegawai = current_user.tipe_pegawai
        att = Attendance.query.filter_by(employee_id=current_user.id, tanggal=hari_ini).first()

        if not att or not att.jam_masuk:
            telat, _ = hitung_telat_lembur(
                jam_sekarang, None, settings, tipe_pegawai, tanggal=hari_ini, employee_id=current_user.id
            )
            if not att:
                att = Attendance(employee_id=current_user.id, tanggal=hari_ini)
                db.session.add(att)
            att.status = "Hadir"
            att.jam_masuk = jam_sekarang
            att.telat_menit = telat
            att.lokasi_masuk_lat = lat
            att.lokasi_masuk_lng = lng
            db.session.commit()
            pesan = f"Absen masuk berhasil pukul {jam_sekarang}."
            if telat > 0:
                pesan += f" Anda terlambat {telat} menit."
            return jsonify(success=True, aksi="masuk", message=pesan, telat=telat)

        if not att.jam_pulang:
            _, lembur = hitung_telat_lembur(
                att.jam_masuk, jam_sekarang, settings, tipe_pegawai, tanggal=hari_ini, employee_id=current_user.id
            )
            pulang_cepat = cek_pulang_cepat(jam_sekarang, settings, tipe_pegawai)
            att.jam_pulang = jam_sekarang
            att.lembur_menit = lembur
            att.lokasi_pulang_lat = lat
            att.lokasi_pulang_lng = lng
            db.session.commit()
            pesan = f"Absen pulang berhasil pukul {jam_sekarang}."
            if pulang_cepat > 0:
                pesan += f" ⚠️ Anda pulang {pulang_cepat} menit lebih awal dari jam standar."
            return jsonify(success=True, aksi="pulang", message=pesan, pulang_cepat=pulang_cepat)

        return jsonify(success=False, message="Anda sudah absen masuk & pulang hari ini."), 400

    @app.route("/pegawai/riwayat")
    @pegawai_required
    def pegawai_riwayat():
        riwayat = (
            Attendance.query.filter_by(employee_id=current_user.id)
            .order_by(Attendance.tanggal.desc())
            .limit(31)
            .all()
        )
        return render_template("pegawai/riwayat.html", riwayat=riwayat)

    @app.route("/pegawai/izin", methods=["GET", "POST"])
    @pegawai_required
    def pegawai_izin():
        if request.method == "POST":
            jenis = request.form.get("jenis")
            try:
                tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
            except ValueError:
                tanggal = None

            dokumen_file = request.files.get("dokumen")
            dokumen_ada = bool(dokumen_file and dokumen_file.filename)

            if jenis not in ("Sakit", "Izin", "Cuti") or not tanggal:
                flash("Lengkapi tanggal dan jenis pengajuan dengan benar.", "danger")
            elif jenis in ("Sakit", "Cuti") and not dokumen_ada:
                flash(
                    "Dokumen wajib dilampirkan untuk pengajuan Sakit atau Cuti "
                    "(surat izin cuti / surat keterangan dokter).",
                    "danger",
                )
            elif dokumen_ada and dokumen_file.filename.rsplit(".", 1)[-1].lower() not in EKSTENSI_DOKUMEN_IZIN:
                flash("Format dokumen harus PDF, DOC, DOCX, JPG, atau PNG.", "danger")
            else:
                sudah_ada = PengajuanIzin.query.filter_by(
                    employee_id=current_user.id, tanggal=tanggal, status="Menunggu"
                ).first()
                if sudah_ada:
                    flash("Sudah ada pengajuan untuk tanggal ini yang masih menunggu persetujuan.", "warning")
                else:
                    dokumen_filename = None
                    if dokumen_ada:
                        ext = dokumen_file.filename.rsplit(".", 1)[-1].lower()
                        dokumen_filename = secure_filename(
                            f"izin_{current_user.id}_{tanggal.isoformat()}_{now_wib().strftime('%H%M%S')}.{ext}"
                        )
                        dokumen_file.save(os.path.join(app.config["IZIN_UPLOAD_FOLDER"], dokumen_filename))

                    pengajuan_baru = PengajuanIzin(
                        employee_id=current_user.id,
                        tanggal=tanggal,
                        jenis=jenis,
                        alasan=request.form.get("alasan", "").strip(),
                        dokumen_filename=dokumen_filename,
                    )
                    db.session.add(pengajuan_baru)
                    db.session.commit()
                    flash("Pengajuan berhasil dikirim, menunggu persetujuan admin.", "success")
                    try:
                        kirim_notifikasi_pengajuan_izin(pengajuan_baru, get_settings())
                    except Exception:
                        pass
            return redirect(url_for("pegawai_izin"))

        riwayat = (
            PengajuanIzin.query.filter_by(employee_id=current_user.id)
            .order_by(PengajuanIzin.tanggal_diajukan.desc())
            .limit(20)
            .all()
        )
        return render_template("pegawai/izin.html", riwayat=riwayat)

    @app.route("/pegawai/akun", methods=["GET", "POST"])
    @pegawai_required
    def pegawai_akun():
        if request.method == "POST":
            password_baru = request.form.get("password_baru", "")
            password_konfirmasi = request.form.get("password_konfirmasi", "")
            if len(password_baru) < 6:
                flash("Password baru minimal 6 karakter.", "danger")
            elif password_baru != password_konfirmasi:
                flash("Konfirmasi password tidak cocok.", "danger")
            else:
                current_user.set_password(password_baru)
                db.session.commit()
                flash("Password berhasil diubah.", "success")
            return redirect(url_for("pegawai_akun"))
        return render_template("pegawai/akun.html")

    @app.route("/pegawai/laporan", methods=["GET", "POST"])
    @pegawai_required
    def pegawai_laporan():
        if request.method == "POST":
            try:
                tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
            except ValueError:
                tanggal = None
            isi_laporan = request.form.get("isi_laporan", "").strip()
            lampiran_file = request.files.get("lampiran")
            lampiran_ada = bool(lampiran_file and lampiran_file.filename)

            if not tanggal or not isi_laporan:
                flash("Lengkapi tanggal dan isi laporan.", "danger")
            elif lampiran_ada and lampiran_file.filename.rsplit(".", 1)[-1].lower() not in EKSTENSI_LAMPIRAN_LAPORAN:
                flash(
                    "Format lampiran tidak didukung. Gunakan PDF, Word, Excel, PowerPoint, gambar, CSV, TXT, atau ZIP.",
                    "danger",
                )
            else:
                lampiran_filename = None
                lampiran_nama_asli = None
                if lampiran_ada:
                    ext = lampiran_file.filename.rsplit(".", 1)[-1].lower()
                    lampiran_filename = secure_filename(
                        f"laporan_{current_user.id}_{tanggal.isoformat()}_{now_wib().strftime('%H%M%S')}.{ext}"
                    )
                    lampiran_file.save(os.path.join(app.config["LAPORAN_UPLOAD_FOLDER"], lampiran_filename))
                    lampiran_nama_asli = secure_filename(lampiran_file.filename)[:256]

                db.session.add(
                    LaporanPekerjaan(
                        employee_id=current_user.id,
                        tanggal=tanggal,
                        isi_laporan=isi_laporan,
                        lampiran_filename=lampiran_filename,
                        lampiran_nama_asli=lampiran_nama_asli,
                    )
                )
                db.session.commit()
                flash("Laporan pekerjaan berhasil dikirim.", "success")
            return redirect(url_for("pegawai_laporan"))

        riwayat = (
            LaporanPekerjaan.query.filter_by(employee_id=current_user.id)
            .order_by(LaporanPekerjaan.tanggal.desc())
            .limit(30)
            .all()
        )
        return render_template("pegawai/laporan.html", riwayat=riwayat, tanggal_default=today_wib().isoformat())

    @app.route("/pegawai/lembur", methods=["GET", "POST"])
    @pegawai_required
    def pegawai_lembur():
        if request.method == "POST":
            try:
                tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
            except ValueError:
                tanggal = None
            jam_mulai = request.form.get("jam_mulai") or None
            jam_selesai = request.form.get("jam_selesai") or None
            alasan = request.form.get("alasan", "").strip()

            durasi_valid = False
            if jam_mulai and jam_selesai:
                try:
                    hm, mm = jam_mulai.split(":")
                    hs, ms = jam_selesai.split(":")
                    durasi_valid = (int(hs) * 60 + int(ms)) > (int(hm) * 60 + int(mm))
                except ValueError:
                    durasi_valid = False

            if not tanggal or not jam_mulai or not jam_selesai:
                flash("Lengkapi tanggal, jam mulai, dan jam selesai.", "danger")
            elif not durasi_valid:
                flash("Jam selesai harus lebih besar dari jam mulai.", "danger")
            else:
                db.session.add(
                    PengajuanLembur(
                        employee_id=current_user.id,
                        tanggal=tanggal,
                        jam_mulai=jam_mulai,
                        jam_selesai=jam_selesai,
                        alasan=alasan,
                    )
                )
                db.session.commit()
                flash("Pengajuan lembur berhasil dikirim, menunggu persetujuan admin.", "success")
            return redirect(url_for("pegawai_lembur"))

        riwayat = (
            PengajuanLembur.query.filter_by(employee_id=current_user.id)
            .order_by(PengajuanLembur.tanggal_diajukan.desc())
            .limit(20)
            .all()
        )
        return render_template("pegawai/lembur.html", riwayat=riwayat)

    # ---------- LAPORAN PEKERJAAN (ADMIN) ----------
    @app.route("/laporan-pekerjaan")
    @admin_required
    def laporan_pekerjaan_list():
        employee_id = request.args.get("employee_id", type=int)
        query = LaporanPekerjaan.query.join(Employee)
        if employee_id:
            query = query.filter(LaporanPekerjaan.employee_id == employee_id)
        laporan = query.order_by(LaporanPekerjaan.tanggal.desc()).limit(200).all()
        employees = Employee.query.order_by(Employee.nama).all()
        return render_template(
            "laporan_pekerjaan_list.html", laporan=laporan, employees=employees, employee_id=employee_id
        )

    # ---------- PENGAJUAN LEMBUR (ADMIN) ----------
    @app.route("/pengajuan-lembur")
    @admin_required
    def pengajuan_lembur_list():
        status_filter = request.args.get("status", "Menunggu")
        query = PengajuanLembur.query.join(Employee)
        if status_filter in ("Menunggu", "Disetujui", "Ditolak"):
            query = query.filter(PengajuanLembur.status == status_filter)
        pengajuan = query.order_by(PengajuanLembur.tanggal_diajukan.desc()).all()
        jumlah_menunggu = PengajuanLembur.query.filter_by(status="Menunggu").count()
        wa_links_karyawan = {p.id: buat_wa_link_hasil_lembur(p) for p in pengajuan}
        return render_template(
            "pengajuan_lembur_list.html",
            pengajuan=pengajuan,
            status_filter=status_filter,
            jumlah_menunggu=jumlah_menunggu,
            wa_links_karyawan=wa_links_karyawan,
        )

    @app.route("/pengajuan-lembur/<int:pid>/setujui", methods=["POST"])
    @admin_required
    def pengajuan_lembur_setujui(pid):
        p = db.session.get(PengajuanLembur, pid) or abort_404()
        p.status = "Disetujui"
        p.tanggal_diproses = now_wib()

        hm, mm = p.jam_mulai.split(":")
        hs, ms = p.jam_selesai.split(":")
        durasi_menit = (int(hs) * 60 + int(ms)) - (int(hm) * 60 + int(mm))

        att = Attendance.query.filter_by(employee_id=p.employee_id, tanggal=p.tanggal).first()
        if not att:
            att = Attendance(employee_id=p.employee_id, tanggal=p.tanggal, status="Hadir")
            db.session.add(att)
        att.lembur_menit = (att.lembur_menit or 0) + durasi_menit

        db.session.commit()

        diperbarui = _reconcile_payroll_draft(p.employee, p.tanggal)
        pesan = f"Pengajuan lembur {p.employee.nama} disetujui & ditambahkan ke absensi."
        kategori = "success"
        if diperbarui is True:
            pesan += " Slip gaji bulan ini otomatis ikut diperbarui."
        elif diperbarui is False:
            pesan += (
                f" Perhatian: slip gaji {p.employee.nama} bulan ini sudah berstatus Dibayar, "
                "jadi lembur ini TIDAK otomatis masuk ke situ -- sesuaikan manual kalau perlu."
            )
            kategori = "warning"
        flash(pesan, kategori)
        return redirect(url_for("pengajuan_lembur_list"))

    @app.route("/pengajuan-lembur/<int:pid>/tolak", methods=["POST"])
    @admin_required
    def pengajuan_lembur_tolak(pid):
        p = db.session.get(PengajuanLembur, pid) or abort_404()
        p.status = "Ditolak"
        p.catatan_admin = request.form.get("catatan_admin", "").strip()
        p.tanggal_diproses = now_wib()
        db.session.commit()
        flash(f"Pengajuan lembur {p.employee.nama} ditolak.", "info")
        return redirect(url_for("pengajuan_lembur_list"))

    # ---------- PENGGAJIAN ----------
    def _hitung_simpan_payroll_karyawan(emp, bulan, tahun, settings):
        """Hitung ulang & simpan slip gaji satu karyawan Tetap/Probation untuk bulan
        tsb dari data absensi terkini. Dipakai baik saat generate massal di menu
        Penggajian maupun otomatis sesudah pengajuan lembur/izin disetujui, supaya
        slip Draft yang sudah dibuat lebih dulu ikut ter-update tanpa admin harus
        ingat klik Generate ulang. Return (payroll, diperbarui) -- diperbarui=False
        kalau slipnya sudah berstatus Dibayar (tidak boleh ditimpa)."""
        awal = date(tahun, bulan, 1)
        akhir = date(tahun, bulan, calendar.monthrange(tahun, bulan)[1])
        absensi = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.tanggal >= awal,
            Attendance.tanggal <= akhir,
        ).all()

        total_hadir = sum(1 for a in absensi if a.status == "Hadir")
        total_sakit = sum(1 for a in absensi if a.status == "Sakit")
        total_izin = sum(1 for a in absensi if a.status == "Izin")
        total_cuti = sum(1 for a in absensi if a.status == "Cuti")
        total_alpha = sum(1 for a in absensi if a.status == "Alpha")
        total_telat_menit = sum(a.telat_menit or 0 for a in absensi)
        total_lembur_menit = sum(a.lembur_menit or 0 for a in absensi)

        co_host_fee = 0
        upah_freelance = 0

        total_pokok = emp.gaji_pokok + emp.tunjangan_makan + emp.tunjangan_transport
        hari_kerja = settings.hari_kerja_per_bulan or 22
        gaji_harian = total_pokok / hari_kerja if hari_kerja else 0

        potongan_alpha = round(total_alpha * gaji_harian)
        # Pemotongan gaji akibat keterlambatan dinonaktifkan sementara atas permintaan
        # -- total_telat_menit tetap dihitung & disimpan supaya laporan keterlambatan
        # per karyawan tetap tampil, hanya saja tidak lagi mengurangi gaji_bersih.
        potongan_telat = 0

        bpjs_jkk = emp.bpjs_jkk or 0
        bpjs_jkm = emp.bpjs_jkm or 0
        bpjs_jht = emp.bpjs_jht or 0

        if emp.bpjs_kesehatan_terdaftar:
            basis_bpjs_kesehatan = min(total_pokok, BATAS_GAJI_BPJS_KESEHATAN)
            bpjs_kesehatan = round(basis_bpjs_kesehatan * PERSEN_BPJS_KESEHATAN_KARYAWAN)
            bpjs_kesehatan_perusahaan = round(basis_bpjs_kesehatan * PERSEN_BPJS_KESEHATAN_PERUSAHAAN)
        else:
            bpjs_kesehatan = 0
            bpjs_kesehatan_perusahaan = 0

        payroll = Payroll.query.filter_by(
            employee_id=emp.id, bulan=bulan, tahun=tahun
        ).first()
        if not payroll:
            payroll = Payroll(
                employee_id=emp.id, bulan=bulan, tahun=tahun,
                jumlah_lembur=0, tarif_lembur=0, uang_lembur=0,
            )
            db.session.add(payroll)

        if payroll.status == "Dibayar":
            return payroll, False  # jangan timpa slip yang sudah dibayar

        # Lembur & Bonus sekarang diisi manual oleh admin lewat form di halaman
        # detail slip (lembur: jumlah hari x tarif per hari; bonus: nominal
        # langsung) -- tidak lagi otomatis dari menit absensi / status target
        # tercapai. Nilai yang sudah admin isi sebelumnya dipertahankan tiap kali
        # slip di-generate ulang.
        jumlah_lembur = payroll.jumlah_lembur or 0
        tarif_lembur = payroll.tarif_lembur or 0
        uang_lembur = jumlah_lembur * tarif_lembur
        bonus_target = payroll.bonus_target or 0

        gaji_bersih = (
            total_pokok
            - potongan_alpha
            - potongan_telat
            - bpjs_jkk - bpjs_jkm - bpjs_jht - bpjs_kesehatan
            + uang_lembur
            + bonus_target
        )
        gaji_bersih = max(gaji_bersih, 0)

        payroll.gaji_pokok = emp.gaji_pokok
        payroll.tunjangan_makan = emp.tunjangan_makan
        payroll.tunjangan_transport = emp.tunjangan_transport
        payroll.tipe_pegawai = emp.tipe_pegawai
        payroll.total_hadir = total_hadir
        payroll.total_sakit = total_sakit
        payroll.total_izin = total_izin
        payroll.total_cuti = total_cuti
        payroll.total_alpha = total_alpha
        payroll.total_telat_menit = total_telat_menit
        payroll.total_lembur_menit = total_lembur_menit
        payroll.potongan_alpha = potongan_alpha
        payroll.potongan_telat = potongan_telat
        payroll.jumlah_lembur = jumlah_lembur
        payroll.tarif_lembur = tarif_lembur
        payroll.uang_lembur = uang_lembur
        payroll.upah_freelance = upah_freelance
        payroll.bonus_target = bonus_target
        payroll.co_host_fee = co_host_fee
        payroll.bpjs_jkk = bpjs_jkk
        payroll.bpjs_jkm = bpjs_jkm
        payroll.bpjs_jht = bpjs_jht
        payroll.bpjs_kesehatan = bpjs_kesehatan
        payroll.bpjs_kesehatan_perusahaan = bpjs_kesehatan_perusahaan
        payroll.gaji_bersih = gaji_bersih
        return payroll, True

    def _hitung_simpan_payroll_freelance(emp, bulan, tahun, settings):
        """Sama seperti _hitung_simpan_payroll_karyawan tapi pakai skema upah
        freelance (per hari kerja + lembur freelance). Return (payroll, diperbarui)."""
        awal = date(tahun, bulan, 1)
        akhir = date(tahun, bulan, calendar.monthrange(tahun, bulan)[1])
        absensi = Attendance.query.filter(
            Attendance.employee_id == emp.id,
            Attendance.tanggal >= awal,
            Attendance.tanggal <= akhir,
        ).all()
        total_hadir = sum(1 for a in absensi if a.status == "Hadir")
        hari_kerja = total_hadir
        total_sakit = sum(1 for a in absensi if a.status == "Sakit")
        total_izin = sum(1 for a in absensi if a.status == "Izin")
        total_cuti = sum(1 for a in absensi if a.status == "Cuti")
        total_alpha = sum(1 for a in absensi if a.status == "Alpha")
        total_telat_menit = sum(a.telat_menit or 0 for a in absensi)
        total_lembur_menit = sum(a.lembur_menit or 0 for a in absensi)

        tarif_unit = emp.tarif_unit_freelance or settings.tarif_harian_freelance or 0
        upah_freelance = hari_kerja * tarif_unit
        co_host_fee = (settings.tarif_co_host or 0) if emp.co_host_bulan_ini == "Ya" else 0

        payroll = Payroll.query.filter_by(
            employee_id=emp.id, bulan=bulan, tahun=tahun
        ).first()
        if not payroll:
            payroll = Payroll(
                employee_id=emp.id, bulan=bulan, tahun=tahun,
                jumlah_lembur=0, tarif_lembur=0, uang_lembur=0,
            )
            db.session.add(payroll)

        if payroll.status == "Dibayar":
            return payroll, False  # jangan timpa slip yang sudah dibayar

        # Lembur & Bonus diisi manual oleh admin (lihat catatan yang sama di
        # _hitung_simpan_payroll_karyawan) -- nilai lama dipertahankan saat regenerate.
        jumlah_lembur = payroll.jumlah_lembur or 0
        tarif_lembur = payroll.tarif_lembur or 0
        uang_lembur = jumlah_lembur * tarif_lembur
        bonus_target = payroll.bonus_target or 0

        gaji_bersih = max(upah_freelance + uang_lembur + bonus_target + co_host_fee, 0)

        payroll.gaji_pokok = 0
        payroll.tunjangan_makan = 0
        payroll.tunjangan_transport = 0
        payroll.tipe_pegawai = emp.tipe_pegawai
        payroll.total_hadir = hari_kerja
        payroll.total_sakit = total_sakit
        payroll.total_izin = total_izin
        payroll.total_cuti = total_cuti
        payroll.total_alpha = total_alpha
        payroll.total_telat_menit = total_telat_menit
        payroll.total_lembur_menit = total_lembur_menit
        payroll.potongan_alpha = 0
        payroll.potongan_telat = 0
        payroll.jumlah_lembur = jumlah_lembur
        payroll.tarif_lembur = tarif_lembur
        payroll.uang_lembur = uang_lembur
        payroll.upah_freelance = upah_freelance
        payroll.bonus_target = bonus_target
        payroll.co_host_fee = co_host_fee
        payroll.bpjs_jkk = 0
        payroll.bpjs_jkm = 0
        payroll.bpjs_jht = 0
        payroll.bpjs_kesehatan = 0
        payroll.bpjs_kesehatan_perusahaan = 0
        payroll.gaji_bersih = gaji_bersih
        return payroll, True

    def _reconcile_payroll_draft(employee, tanggal):
        """Kalau slip gaji bulan berjalan utk karyawan ini SUDAH pernah dibuat
        (Draft), langsung hitung ulang & simpan sekarang juga -- dipanggil sesudah
        pengajuan lembur/izin disetujui supaya datanya tidak ketinggalan sampai
        admin ingat klik Generate ulang di menu Penggajian. Return None kalau belum
        ada slip bulan itu sama sekali (tidak perlu tindakan, biar dibuat wajar
        saat admin generate); kalau sudah ada, return True/False (diperbarui atau
        tidak karena sudah Dibayar)."""
        bulan, tahun = tanggal.month, tanggal.year
        payroll_ada = Payroll.query.filter_by(
            employee_id=employee.id, bulan=bulan, tahun=tahun
        ).first()
        if not payroll_ada:
            return None
        settings = get_settings()
        if employee.tipe_pegawai == "Freelance":
            _, diperbarui = _hitung_simpan_payroll_freelance(employee, bulan, tahun, settings)
        else:
            _, diperbarui = _hitung_simpan_payroll_karyawan(employee, bulan, tahun, settings)
        db.session.commit()
        return diperbarui

    @app.route("/penggajian")
    @admin_required
    def penggajian_list():
        bulan = int(request.args.get("bulan", today_wib().month))
        tahun = int(request.args.get("tahun", today_wib().year))
        payrolls = (
            Payroll.query.filter_by(bulan=bulan, tahun=tahun)
            .join(Employee)
            .order_by(Employee.nama)
            .all()
        )
        settings = get_settings()
        for p in payrolls:
            p.wa_link = buat_wa_link(p, settings)
        total_gaji = sum(p.gaji_bersih for p in payrolls)
        return render_template(
            "payroll_list.html",
            payrolls=payrolls,
            bulan=bulan,
            tahun=tahun,
            bulan_nama=BULAN_NAMA,
            total_gaji=total_gaji,
        )

    @app.route("/penggajian/generate", methods=["POST"])
    @admin_required
    def penggajian_generate():
        bulan = int(request.form["bulan"])
        tahun = int(request.form["tahun"])
        settings = get_settings()

        employees = Employee.query.filter(
            Employee.status == "Aktif", Employee.tipe_pegawai != "Freelance"
        ).all()
        for emp in employees:
            _hitung_simpan_payroll_karyawan(emp, bulan, tahun, settings)

        db.session.commit()
        flash(f"Slip gaji Karyawan Tetap/Probation {BULAN_NAMA[bulan]} {tahun} berhasil dibuat/diperbarui.", "success")
        return redirect(url_for("penggajian_list", bulan=bulan, tahun=tahun))

    @app.route("/penggajian/freelance-review")
    @admin_required
    def penggajian_freelance_review():
        bulan = int(request.args.get("bulan", today_wib().month))
        tahun = int(request.args.get("tahun", today_wib().year))
        settings = get_settings()

        awal = date(tahun, bulan, 1)
        akhir = date(tahun, bulan, calendar.monthrange(tahun, bulan)[1])

        freelancer_rows = []
        employees = Employee.query.filter_by(status="Aktif", tipe_pegawai="Freelance").order_by(Employee.nama).all()
        for emp in employees:
            absensi = Attendance.query.filter(
                Attendance.employee_id == emp.id,
                Attendance.tanggal >= awal,
                Attendance.tanggal <= akhir,
            ).all()
            total_hadir = sum(1 for a in absensi if a.status == "Hadir")
            existing_payroll = Payroll.query.filter_by(employee_id=emp.id, bulan=bulan, tahun=tahun).first()
            tarif_unit = emp.tarif_unit_freelance or settings.tarif_harian_freelance or 0
            freelancer_rows.append({
                "employee": emp,
                "hari_kerja_absensi": total_hadir,
                "tarif_unit": tarif_unit,
                "sudah_dibayar": bool(existing_payroll and existing_payroll.status == "Dibayar"),
            })

        return render_template(
            "payroll_freelance_review.html",
            rows=freelancer_rows,
            bulan=bulan,
            tahun=tahun,
            bulan_nama=BULAN_NAMA,
        )

    @app.route("/penggajian/freelance-generate", methods=["POST"])
    @admin_required
    def penggajian_freelance_generate():
        bulan = int(request.form["bulan"])
        tahun = int(request.form["tahun"])
        settings = get_settings()

        employees = Employee.query.filter_by(status="Aktif", tipe_pegawai="Freelance").all()
        for emp in employees:
            _hitung_simpan_payroll_freelance(emp, bulan, tahun, settings)

        db.session.commit()
        flash(f"Slip gaji Freelance {BULAN_NAMA[bulan]} {tahun} berhasil dibuat/diperbarui.", "success")
        return redirect(url_for("penggajian_list", bulan=bulan, tahun=tahun))

    @app.route("/penggajian/<int:payroll_id>")
    @admin_required
    def penggajian_detail(payroll_id):
        payroll = db.session.get(Payroll, payroll_id) or abort_404()
        settings = get_settings()
        periode_awal = date(payroll.tahun, payroll.bulan, 1)
        periode_akhir = date(payroll.tahun, payroll.bulan, calendar.monthrange(payroll.tahun, payroll.bulan)[1])
        tarif_unit_efektif = (
            payroll.employee.tarif_unit_freelance or settings.tarif_harian_freelance or 0
        )
        return render_template(
            "payroll_detail.html",
            p=payroll,
            bulan_nama=BULAN_NAMA,
            settings=settings,
            periode_awal=periode_awal,
            periode_akhir=periode_akhir,
            tarif_unit_efektif=tarif_unit_efektif,
            wa_link=buat_wa_link(payroll, settings),
        )

    @app.route("/penggajian/<int:payroll_id>/lembur", methods=["POST"])
    @admin_required
    def penggajian_simpan_lembur(payroll_id):
        payroll = db.session.get(Payroll, payroll_id) or abort_404()
        if payroll.status == "Dibayar":
            flash("Slip ini sudah berstatus Dibayar, lembur & bonus tidak bisa diubah lagi.", "danger")
            return redirect(url_for("penggajian_detail", payroll_id=payroll_id))

        try:
            jumlah_lembur = max(int(request.form.get("jumlah_lembur") or 0), 0)
        except ValueError:
            jumlah_lembur = 0
        try:
            tarif_lembur = max(int(request.form.get("tarif_lembur") or 0), 0)
        except ValueError:
            tarif_lembur = 0
        try:
            bonus_target = max(int(request.form.get("bonus_target") or 0), 0)
        except ValueError:
            bonus_target = 0

        payroll.jumlah_lembur = jumlah_lembur
        payroll.tarif_lembur = tarif_lembur
        payroll.bonus_target = bonus_target
        db.session.commit()

        settings = get_settings()
        emp = payroll.employee
        if emp.tipe_pegawai == "Freelance":
            _hitung_simpan_payroll_freelance(emp, payroll.bulan, payroll.tahun, settings)
        else:
            _hitung_simpan_payroll_karyawan(emp, payroll.bulan, payroll.tahun, settings)
        db.session.commit()

        flash(
            f"Lembur & bonus {emp.nama} disimpan: lembur {jumlah_lembur} hari x {rupiah(tarif_lembur)} = "
            f"{rupiah(jumlah_lembur * tarif_lembur)}, bonus {rupiah(bonus_target)}. Gaji bersih ikut diperbarui.",
            "success",
        )
        return redirect(url_for("penggajian_detail", payroll_id=payroll_id))

    def pdf_link_callback(uri, rel):
        if uri.startswith("/static/"):
            return os.path.join(app.root_path, uri.lstrip("/"))
        return uri

    def nama_file_slip(payroll):
        return f"Slip_Gaji_{payroll.employee.nama.replace(' ', '_')}_{BULAN_NAMA[payroll.bulan]}_{payroll.tahun}.pdf"

    def buat_slip_pdf_bytes(payroll, settings):
        periode_awal = date(payroll.tahun, payroll.bulan, 1)
        periode_akhir = date(payroll.tahun, payroll.bulan, calendar.monthrange(payroll.tahun, payroll.bulan)[1])
        tarif_unit_efektif = (
            payroll.employee.tarif_unit_freelance or settings.tarif_harian_freelance or 0
        )
        logo_path = None
        if settings.logo_filename:
            candidate = os.path.join(app.config["UPLOAD_FOLDER"], settings.logo_filename)
            if os.path.exists(candidate):
                logo_path = candidate

        html = render_template(
            "payroll_slip_pdf.html",
            p=payroll,
            settings=settings,
            periode_awal=periode_awal,
            periode_akhir=periode_akhir,
            tarif_unit_efektif=tarif_unit_efektif,
            logo_path=logo_path,
        )

        buffer = BytesIO()
        pisa.CreatePDF(html, dest=buffer, link_callback=pdf_link_callback)
        return buffer.getvalue()

    def kirim_email_slip(payroll, settings):
        email_tujuan = (payroll.employee.email or "").strip()
        if not email_tujuan:
            return False, "Karyawan belum punya alamat email terdaftar."

        smtp_email = os.environ.get("SMTP_EMAIL")
        smtp_password = os.environ.get("SMTP_APP_PASSWORD")
        if not smtp_email or not smtp_password:
            return False, "Pengiriman email belum dikonfigurasi di server (SMTP_EMAIL/SMTP_APP_PASSWORD)."

        pdf_bytes = buat_slip_pdf_bytes(payroll, settings)

        msg = EmailMessage()
        msg["Subject"] = f"Slip Gaji {BULAN_NAMA[payroll.bulan]} {payroll.tahun} - {settings.nama_perusahaan}"
        msg["From"] = smtp_email
        msg["To"] = email_tujuan
        msg.set_content(
            f"Halo {payroll.employee.nama},\n\n"
            f"Berikut slip gaji Anda periode {BULAN_NAMA[payroll.bulan]} {payroll.tahun} dari "
            f"{settings.nama_perusahaan} (rincian lengkap ada di file PDF terlampir).\n\n"
            "Status: SUDAH DIBAYAR\n\n"
            "Terima kasih atas kerja keras Anda selama ini. Jika ada pertanyaan silakan hubungi admin."
        )
        msg.add_attachment(
            pdf_bytes, maintype="application", subtype="pdf", filename=nama_file_slip(payroll)
        )

        try:
            with smtplib.SMTP("smtp.gmail.com", 587) as server:
                server.starttls()
                server.login(smtp_email, smtp_password)
                server.send_message(msg)
            return True, None
        except Exception as e:
            return False, str(e)

    def kirim_notifikasi_pengajuan_izin(pengajuan, settings):
        smtp_email = os.environ.get("SMTP_EMAIL")
        smtp_password = os.environ.get("SMTP_APP_PASSWORD")
        if not smtp_email or not smtp_password:
            return False, "Pengiriman email belum dikonfigurasi di server (SMTP_EMAIL/SMTP_APP_PASSWORD)."

        penerima = {smtp_email.strip().lower(): smtp_email, "maslahaniqab@gmail.com": "maslahaniqab@gmail.com"}
        email_manager = (settings.email_perusahaan or "").strip()
        if email_manager:
            penerima[email_manager.lower()] = email_manager
        daftar_penerima = list(penerima.values())

        karyawan = pengajuan.employee
        msg = EmailMessage()
        msg["Subject"] = f"Pengajuan {pengajuan.jenis} - {karyawan.nama}"
        msg["From"] = smtp_email
        msg["To"] = ", ".join(daftar_penerima)
        msg.set_content(
            f"Ada pengajuan {pengajuan.jenis} baru yang perlu ditinjau.\n\n"
            f"Nama: {karyawan.nama}\n"
            f"Jabatan: {karyawan.jabatan or '-'}\n"
            f"Untuk tanggal: {pengajuan.tanggal.strftime('%d-%m-%Y')}\n"
            f"Alasan: {pengajuan.alasan or '-'}\n"
            f"Dokumen pendukung: {'Ada, cek di portal' if pengajuan.dokumen_filename else 'Tidak ada'}\n\n"
            "Silakan login ke Maslaha Portal (menu Karyawan > Pengajuan Izin) untuk menyetujui atau menolak."
        )

        try:
            with smtplib.SMTP("smtp.gmail.com", 587) as server:
                server.starttls()
                server.login(smtp_email, smtp_password)
                server.send_message(msg)
            return True, None
        except Exception as e:
            return False, str(e)

    def buat_wa_link_notifikasi_izin(pengajuan, settings):
        nomor = normalisasi_no_hp_wa(settings.no_hp_perusahaan)
        if not nomor:
            return None
        pesan = (
            f"Info: Ada pengajuan {pengajuan.jenis} baru dari {pengajuan.employee.nama} "
            f"untuk tanggal {pengajuan.tanggal.strftime('%d-%m-%Y')}. "
            f"Alasan: {pengajuan.alasan or '-'}. Mohon dicek & disetujui di Maslaha Portal ya."
        )
        return f"https://wa.me/{nomor}?text={quote(pesan)}"

    def buat_wa_link_hasil_izin(pengajuan):
        nomor = normalisasi_no_hp_wa(pengajuan.employee.no_hp)
        if not nomor:
            return None
        if pengajuan.status == "Disetujui":
            pesan = (
                f"Halo {pengajuan.employee.nama}, pengajuan {pengajuan.jenis} Anda untuk tanggal "
                f"{pengajuan.tanggal.strftime('%d-%m-%Y')} sudah *DISETUJUI*. Terima kasih."
            )
        elif pengajuan.status == "Ditolak":
            pesan = (
                f"Halo {pengajuan.employee.nama}, mohon maaf pengajuan {pengajuan.jenis} Anda untuk tanggal "
                f"{pengajuan.tanggal.strftime('%d-%m-%Y')} *DITOLAK*."
                + (f" Alasan: {pengajuan.catatan_admin}." if pengajuan.catatan_admin else "")
            )
        else:
            return None
        return f"https://wa.me/{nomor}?text={quote(pesan)}"

    def buat_wa_link_hasil_lembur(pengajuan):
        nomor = normalisasi_no_hp_wa(pengajuan.employee.no_hp)
        if not nomor:
            return None
        if pengajuan.status == "Disetujui":
            pesan = (
                f"Halo {pengajuan.employee.nama}, pengajuan lembur Anda tanggal "
                f"{pengajuan.tanggal.strftime('%d-%m-%Y')} pukul {pengajuan.jam_mulai}-{pengajuan.jam_selesai} "
                f"sudah *DISETUJUI* dan sudah masuk ke absensi. Terima kasih."
            )
        elif pengajuan.status == "Ditolak":
            pesan = (
                f"Halo {pengajuan.employee.nama}, mohon maaf pengajuan lembur Anda tanggal "
                f"{pengajuan.tanggal.strftime('%d-%m-%Y')} *DITOLAK*."
                + (f" Alasan: {pengajuan.catatan_admin}." if pengajuan.catatan_admin else "")
            )
        else:
            return None
        return f"https://wa.me/{nomor}?text={quote(pesan)}"

    @app.route("/penggajian/<int:payroll_id>/pdf")
    @admin_required
    def penggajian_pdf(payroll_id):
        payroll = db.session.get(Payroll, payroll_id) or abort_404()
        settings = get_settings()
        pdf_bytes = buat_slip_pdf_bytes(payroll, settings)
        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{nama_file_slip(payroll)}"'},
        )

    @app.route("/penggajian/<int:payroll_id>/kirim-email", methods=["POST"])
    @admin_required
    def penggajian_kirim_email(payroll_id):
        payroll = db.session.get(Payroll, payroll_id) or abort_404()
        settings = get_settings()
        berhasil, pesan_error = kirim_email_slip(payroll, settings)
        if berhasil:
            flash(f"Slip gaji berhasil dikirim ke email {payroll.employee.email}.", "success")
        else:
            flash(f"Gagal mengirim email: {pesan_error}", "danger")
        return redirect(url_for("penggajian_detail", payroll_id=payroll.id))

    @app.route("/penggajian/kirim-email-massal", methods=["POST"])
    @admin_required
    def penggajian_kirim_email_massal():
        bulan = int(request.form.get("bulan", today_wib().month))
        tahun = int(request.form.get("tahun", today_wib().year))
        payroll_ids = [int(x) for x in request.form.getlist("payroll_ids")]

        if not payroll_ids:
            flash("Pilih minimal satu slip untuk dikirim ke email.", "warning")
            return redirect(url_for("penggajian_list", bulan=bulan, tahun=tahun))

        settings = get_settings()
        berhasil = []
        gagal = []
        for payroll_id in payroll_ids:
            payroll = db.session.get(Payroll, payroll_id)
            if not payroll:
                continue
            ok, pesan_error = kirim_email_slip(payroll, settings)
            if ok:
                berhasil.append(payroll.employee.nama)
            else:
                gagal.append(f"{payroll.employee.nama} ({pesan_error})")

        if berhasil:
            flash(f"Slip berhasil dikirim ke email {len(berhasil)} karyawan: {', '.join(berhasil)}.", "success")
        if gagal:
            flash(f"Gagal mengirim ke {len(gagal)} karyawan: {'; '.join(gagal)}.", "danger")

        return redirect(url_for("penggajian_list", bulan=bulan, tahun=tahun))

    @app.route("/penggajian/<int:payroll_id>/bayar", methods=["POST"])
    @admin_required
    def penggajian_bayar(payroll_id):
        payroll = db.session.get(Payroll, payroll_id) or abort_404()
        payroll.status = "Dibayar"
        payroll.tanggal_dibayar = now_wib()
        db.session.commit()

        settings = get_settings()
        berhasil, pesan_error = kirim_email_slip(payroll, settings)
        if berhasil:
            flash(
                f"Gaji {payroll.employee.nama} ditandai sudah dibayar. "
                f"Slip terkirim otomatis ke email {payroll.employee.email}.",
                "success",
            )
        else:
            flash(
                f"Gaji {payroll.employee.nama} ditandai sudah dibayar. "
                f"Email slip tidak terkirim otomatis: {pesan_error}",
                "warning",
            )
        return redirect(url_for("penggajian_list", bulan=payroll.bulan, tahun=payroll.tahun))

    # ---------- KEUANGAN: LAPORAN PENGELUARAN ----------
    @app.route("/keuangan/pengeluaran", methods=["GET", "POST"])
    @admin_required
    def pengeluaran_list():
        if request.method == "POST":
            try:
                tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
            except ValueError:
                tanggal = None

            kategori = request.form.get("kategori", "").strip()
            if kategori == "__lainnya__":
                kategori = request.form.get("kategori_manual", "").strip()

            keterangan = request.form.get("keterangan", "").strip()
            try:
                jumlah = int(request.form.get("jumlah") or 0)
            except ValueError:
                jumlah = 0

            if not tanggal or not kategori or jumlah <= 0:
                flash("Lengkapi tanggal, kategori, dan jumlah (harus lebih dari 0).", "danger")
                redirect_bulan, redirect_tahun = today_wib().month, today_wib().year
            else:
                db.session.add(
                    PengeluaranOperasional(
                        tanggal=tanggal, kategori=kategori, keterangan=keterangan, jumlah=jumlah
                    )
                )
                db.session.commit()
                flash("Pengeluaran berhasil dicatat.", "success")
                redirect_bulan, redirect_tahun = tanggal.month, tanggal.year
            return redirect(url_for("pengeluaran_list", bulan=redirect_bulan, tahun=redirect_tahun))

        bulan = int(request.args.get("bulan", today_wib().month))
        tahun = int(request.args.get("tahun", today_wib().year))
        awal = date(tahun, bulan, 1)
        akhir = date(tahun, bulan, calendar.monthrange(tahun, bulan)[1])
        pengeluaran = (
            PengeluaranOperasional.query
            .filter(PengeluaranOperasional.tanggal >= awal, PengeluaranOperasional.tanggal <= akhir)
            .order_by(PengeluaranOperasional.tanggal.desc())
            .all()
        )
        total = sum(p.jumlah for p in pengeluaran)

        # Biaya iklan (otomatis dari menu Marketing, bukan input manual) -- ditampilkan di sini
        # supaya Laporan Pengeluaran mencerminkan SEMUA beban operasional dalam satu tempat,
        # sinkron dengan Laporan Laba/Rugi.
        labarugi = hitung_labarugi_periode(bulan, tahun)
        iklan_marketplace_list = [
            {"marketplace": mp, "biaya": labarugi["iklan_by_mp"].get(mp, 0)} for mp in MARKETPLACE_LIST
        ]
        total_iklan_marketplace = sum(x["biaya"] for x in iklan_marketplace_list)
        total_iklan = total_iklan_marketplace + labarugi["meta_biaya_kotor"]
        total_keseluruhan = total + total_iklan

        return render_template(
            "pengeluaran_list.html",
            pengeluaran=pengeluaran,
            bulan=bulan,
            tahun=tahun,
            total=total,
            kategori_rutin=KATEGORI_PENGELUARAN_RUTIN,
            tanggal_default=today_wib().isoformat(),
            iklan_marketplace_list=iklan_marketplace_list,
            iklan_meta_biaya=labarugi["meta_biaya_kotor"],
            total_iklan=total_iklan,
            total_keseluruhan=total_keseluruhan,
        )

    @app.route("/keuangan/pengeluaran/<int:pid>/hapus", methods=["POST"])
    @admin_required
    def pengeluaran_hapus(pid):
        p = db.session.get(PengeluaranOperasional, pid) or abort_404()
        bulan, tahun = p.tanggal.month, p.tanggal.year
        db.session.delete(p)
        db.session.commit()
        flash("Pengeluaran dihapus.", "info")
        return redirect(url_for("pengeluaran_list", bulan=bulan, tahun=tahun))

    # ---------- KEUANGAN: LABA/RUGI ----------
    def hitung_labarugi_periode(bulan, tahun):
        awal = date(tahun, bulan, 1)
        akhir = date(tahun, bulan, calendar.monthrange(tahun, bulan)[1])

        pendapatan_items = [
            (i.deskripsi, i.jumlah)
            for i in ItemLabaRugi.query.filter_by(bulan=bulan, tahun=tahun, kelompok="Pendapatan")
            .order_by(ItemLabaRugi.id).all()
        ]
        hpp_items = [
            (i.deskripsi, i.jumlah)
            for i in ItemLabaRugi.query.filter_by(bulan=bulan, tahun=tahun, kelompok="Beban Pokok Penjualan")
            .order_by(ItemLabaRugi.id).all()
        ]
        pendapatan_non_op_items = [
            (i.deskripsi, i.jumlah)
            for i in ItemLabaRugi.query.filter_by(bulan=bulan, tahun=tahun, kelompok="Pendapatan Non Operasional")
            .order_by(ItemLabaRugi.id).all()
        ]
        beban_non_op_items = [
            (i.deskripsi, i.jumlah)
            for i in ItemLabaRugi.query.filter_by(bulan=bulan, tahun=tahun, kelompok="Beban Non Operasional")
            .order_by(ItemLabaRugi.id).all()
        ]
        beban_operasional_manual_items = [
            (i.deskripsi, i.jumlah)
            for i in ItemLabaRugi.query.filter_by(bulan=bulan, tahun=tahun, kelompok="Beban Operasional")
            .order_by(ItemLabaRugi.id).all()
        ]

        pendapatan_marketplace = hitung_pendapatan_gross_marketplace(bulan, tahun)
        for mp, agg in pendapatan_marketplace.items():
            if agg["pendapatan"]:
                pendapatan_items.append((f"Penjualan {mp}", agg["pendapatan"]))

        total_pendapatan = sum(v for _, v in pendapatan_items)
        total_hpp = sum(v for _, v in hpp_items)
        laba_kotor = total_pendapatan - total_hpp

        iklan_by_mp = {}
        for r in IklanMarketplace.query.filter(
            IklanMarketplace.tanggal >= awal, IklanMarketplace.tanggal <= akhir
        ).all():
            iklan_by_mp[r.marketplace] = iklan_by_mp.get(r.marketplace, 0) + (r.biaya or 0)

        gaji_total = sum(p.gaji_bersih or 0 for p in Payroll.query.filter_by(bulan=bulan, tahun=tahun).all())

        opex_by_kategori = {}
        for r in PengeluaranOperasional.query.filter(
            PengeluaranOperasional.tanggal >= awal, PengeluaranOperasional.tanggal <= akhir
        ).all():
            opex_by_kategori[r.kategori] = opex_by_kategori.get(r.kategori, 0) + (r.jumlah or 0)

        meta_rows = IklanMeta.query.filter(IklanMeta.tanggal >= awal, IklanMeta.tanggal <= akhir).all()
        # "biaya" di laporan Iklan Meta adalah total yang sudah termasuk pajak (PPN), jadi "Pajak Iklan"
        # bukan beban tambahan di luar biaya itu -- melainkan pecahan dari biaya yang sama. iklan_meta_total
        # yang ditampilkan/dihitung di beban operasional dibuat NET (biaya - pajak) supaya tidak dobel hitung;
        # pajak_iklan_total tetap ditampilkan terpisah sebagai rincian, dan keduanya berjumlah sama dengan
        # meta_biaya_kotor (total riil yang dibayar ke Meta).
        meta_biaya_kotor = sum(r.biaya or 0 for r in meta_rows)
        pajak_iklan_total = sum(r.pajak or 0 for r in meta_rows)
        iklan_meta_total = meta_biaya_kotor - pajak_iklan_total

        meta_omzet_total = sum(r.omzet or 0 for r in meta_rows)
        meta_klik_total = sum(r.klik or 0 for r in meta_rows)
        meta_impresi_total = sum(r.impresi or 0 for r in meta_rows)
        meta_pesanan_total = sum(r.pesanan or 0 for r in meta_rows)
        meta_roas = (meta_omzet_total / meta_biaya_kotor) if meta_biaya_kotor else 0
        meta_ctr = (meta_klik_total / meta_impresi_total * 100) if meta_impresi_total else 0
        meta_cpa = (meta_biaya_kotor / meta_pesanan_total) if meta_pesanan_total else 0

        total_beban_operasional_manual = sum(v for _, v in beban_operasional_manual_items)

        total_beban_operasional = (
            sum(iklan_by_mp.values()) + gaji_total + sum(opex_by_kategori.values())
            + meta_biaya_kotor + total_beban_operasional_manual
        )
        pendapatan_operasional = laba_kotor - total_beban_operasional

        total_pendapatan_non_op = sum(v for _, v in pendapatan_non_op_items)
        total_beban_non_op = sum(v for _, v in beban_non_op_items)
        pendapatan_beban_non_op = total_pendapatan_non_op - total_beban_non_op
        laba_bersih = pendapatan_operasional + total_pendapatan_non_op - total_beban_non_op

        return {
            "bulan": bulan,
            "tahun": tahun,
            "pendapatan_items": pendapatan_items,
            "total_pendapatan": total_pendapatan,
            "hpp_items": hpp_items,
            "total_hpp": total_hpp,
            "laba_kotor": laba_kotor,
            "iklan_by_mp": iklan_by_mp,
            "gaji_total": gaji_total,
            "opex_by_kategori": opex_by_kategori,
            "iklan_meta_total": iklan_meta_total,
            "pajak_iklan_total": pajak_iklan_total,
            "meta_biaya_kotor": meta_biaya_kotor,
            "meta_omzet_total": meta_omzet_total,
            "meta_klik_total": meta_klik_total,
            "meta_impresi_total": meta_impresi_total,
            "meta_pesanan_total": meta_pesanan_total,
            "meta_roas": meta_roas,
            "meta_ctr": meta_ctr,
            "meta_cpa": meta_cpa,
            "beban_operasional_manual_items": beban_operasional_manual_items,
            "total_beban_operasional": total_beban_operasional,
            "pendapatan_operasional": pendapatan_operasional,
            "pendapatan_non_op_items": pendapatan_non_op_items,
            "total_pendapatan_non_op": total_pendapatan_non_op,
            "beban_non_op_items": beban_non_op_items,
            "total_beban_non_op": total_beban_non_op,
            "pendapatan_beban_non_op": pendapatan_beban_non_op,
            "laba_bersih": laba_bersih,
        }

    def _nilai_item(items, label):
        for l, v in items:
            if l == label:
                return v
        return 0

    @app.route("/keuangan/laba-rugi")
    @admin_required
    def laba_rugi():
        dari_bulan = int(request.args.get("dari_bulan", today_wib().month))
        dari_tahun = int(request.args.get("dari_tahun", today_wib().year))
        sampai_bulan = int(request.args.get("sampai_bulan", today_wib().month))
        sampai_tahun = int(request.args.get("sampai_tahun", today_wib().year))

        periode_list = []
        b, t = dari_bulan, dari_tahun
        while (t, b) <= (sampai_tahun, sampai_bulan) and len(periode_list) < 12:
            periode_list.append((b, t))
            b += 1
            if b > 12:
                b = 1
                t += 1
        if not periode_list:
            periode_list = [(sampai_bulan, sampai_tahun)]

        data_per_periode = [hitung_labarugi_periode(b, t) for b, t in periode_list]

        def kumpulkan_label(kunci):
            labels = []
            for d in data_per_periode:
                for label, _ in d[kunci]:
                    if label not in labels:
                        labels.append(label)
            return labels

        pendapatan_labels = kumpulkan_label("pendapatan_items")
        hpp_labels = kumpulkan_label("hpp_items")
        non_op_pendapatan_labels = kumpulkan_label("pendapatan_non_op_items")
        non_op_beban_labels = kumpulkan_label("beban_non_op_items")
        beban_operasional_manual_labels = kumpulkan_label("beban_operasional_manual_items")

        marketplace_labels = []
        for d in data_per_periode:
            for mp in d["iklan_by_mp"]:
                if mp not in marketplace_labels:
                    marketplace_labels.append(mp)

        opex_labels = []
        for d in data_per_periode:
            for kat in d["opex_by_kategori"]:
                if kat not in opex_labels:
                    opex_labels.append(kat)

        return render_template(
            "laba_rugi.html",
            periode_list=periode_list,
            data_per_periode=data_per_periode,
            pendapatan_labels=pendapatan_labels,
            hpp_labels=hpp_labels,
            marketplace_labels=marketplace_labels,
            opex_labels=opex_labels,
            non_op_pendapatan_labels=non_op_pendapatan_labels,
            non_op_beban_labels=non_op_beban_labels,
            beban_operasional_manual_labels=beban_operasional_manual_labels,
            nilai=_nilai_item,
            dari_bulan=dari_bulan,
            dari_tahun=dari_tahun,
            sampai_bulan=sampai_bulan,
            sampai_tahun=sampai_tahun,
        )

    @app.route("/keuangan/laba-rugi/input", methods=["GET", "POST"])
    @admin_required
    def laba_rugi_input():
        if request.method == "POST":
            try:
                bulan = int(request.form.get("bulan"))
                tahun = int(request.form.get("tahun"))
            except (TypeError, ValueError):
                bulan = tahun = None

            kelompok = request.form.get("kelompok", "")
            deskripsi = request.form.get("deskripsi", "").strip()
            if deskripsi == "__lainnya__":
                deskripsi = request.form.get("deskripsi_manual", "").strip()

            try:
                jumlah = int(request.form.get("jumlah") or 0)
            except ValueError:
                jumlah = 0

            if not bulan or not tahun or kelompok not in KELOMPOK_LABA_RUGI or not deskripsi or jumlah == 0:
                flash("Lengkapi bulan, tahun, kelompok, deskripsi, dan jumlah (tidak boleh 0).", "danger")
            else:
                db.session.add(
                    ItemLabaRugi(bulan=bulan, tahun=tahun, kelompok=kelompok, deskripsi=deskripsi, jumlah=jumlah)
                )
                db.session.commit()
                flash("Item berhasil dicatat.", "success")
                return redirect(url_for("laba_rugi_input", bulan=bulan, tahun=tahun))
            return redirect(url_for("laba_rugi_input"))

        bulan = int(request.args.get("bulan", today_wib().month))
        tahun = int(request.args.get("tahun", today_wib().year))
        items = (
            ItemLabaRugi.query.filter_by(bulan=bulan, tahun=tahun)
            .order_by(ItemLabaRugi.kelompok, ItemLabaRugi.id)
            .all()
        )
        return render_template(
            "laba_rugi_input.html",
            items=items,
            bulan=bulan,
            tahun=tahun,
            kelompok_list=KELOMPOK_LABA_RUGI,
            preset=PRESET_LABA_RUGI,
        )

    @app.route("/keuangan/laba-rugi/input/<int:iid>/hapus", methods=["POST"])
    @admin_required
    def laba_rugi_hapus(iid):
        i = db.session.get(ItemLabaRugi, iid) or abort_404()
        bulan, tahun = i.bulan, i.tahun
        db.session.delete(i)
        db.session.commit()
        flash("Item dihapus.", "info")
        return redirect(url_for("laba_rugi_input", bulan=bulan, tahun=tahun))

    # ---------- PENDAPATAN: PENJUALAN PER MARKETPLACE ----------
    def proses_baris_penjualan(rows, mapping):
        # File laporan marketplace umumnya 1 baris = 1 SKU/produk, bukan 1 pesanan —
        # order yang berisi beberapa produk akan muncul di beberapa baris dengan
        # No. Pesanan yang sama dan Total Pembayaran/Total Diskon yang berulang.
        # Maka nilai per-pesanan hanya diambil sekali (dedup) berdasarkan No. Pesanan
        # supaya tidak terhitung dobel.
        pakai_no_pesanan = mapping.get("no_pesanan") is not None
        order_map = {}
        dilewati = 0
        contoh_gagal_tanggal = []
        dilewati_angka_tidak_wajar = 0
        contoh_angka_tidak_wajar = []
        for i, row in enumerate(rows):
            def ambil(key):
                idx = mapping.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            tanggal_raw = ambil("tanggal")
            tanggal = parse_tanggal_iklan(tanggal_raw)
            if not tanggal:
                dilewati += 1
                if tanggal_raw not in (None, ""):
                    nilai_str = str(tanggal_raw).strip()
                    if nilai_str and nilai_str not in contoh_gagal_tanggal and len(contoh_gagal_tanggal) < 5:
                        contoh_gagal_tanggal.append(nilai_str)
                continue

            total_penjualan_raw = ambil("total_penjualan")
            total_diskon_raw = ambil("total_diskon")
            total_penjualan_val = parse_angka_iklan(total_penjualan_raw) if total_penjualan_raw is not None else 0
            total_diskon_val = parse_angka_iklan(total_diskon_raw) if total_diskon_raw is not None else 0

            if abs(total_penjualan_val) > BATAS_ANGKA_WAJAR_PENJUALAN or abs(total_diskon_val) > BATAS_ANGKA_WAJAR_PENJUALAN:
                dilewati_angka_tidak_wajar += 1
                if len(contoh_angka_tidak_wajar) < 3:
                    contoh_angka_tidak_wajar.append(str(total_penjualan_raw if abs(total_penjualan_val) > BATAS_ANGKA_WAJAR_PENJUALAN else total_diskon_raw))
                continue

            entri = {
                "tanggal": tanggal,
                "total_penjualan": total_penjualan_val,
                "total_diskon": total_diskon_val,
            }

            no_pesanan = str(ambil("no_pesanan") or "").strip() if pakai_no_pesanan else ""
            kunci = no_pesanan if no_pesanan else f"__baris_{i}"
            if kunci not in order_map:
                order_map[kunci] = entri

        agregat = {}
        for order in order_map.values():
            t = order["tanggal"]
            if t not in agregat:
                agregat[t] = {"jumlah_pesanan": 0, "total_penjualan": 0, "total_diskon": 0}
            agregat[t]["jumlah_pesanan"] += 1
            agregat[t]["total_penjualan"] += order["total_penjualan"]
            agregat[t]["total_diskon"] += order["total_diskon"]

        return agregat, dilewati, contoh_gagal_tanggal, dilewati_angka_tidak_wajar, contoh_angka_tidak_wajar

    def hitung_ringkasan_penjualan(bulan, tahun):
        pendapatan_marketplace = hitung_pendapatan_gross_marketplace(bulan, tahun)
        total = {
            "jumlah_pesanan": sum(a["jumlah_pesanan"] for a in pendapatan_marketplace.values()),
            "pendapatan": sum(a["pendapatan"] for a in pendapatan_marketplace.values()),
        }
        breakdown = [
            {"marketplace": mp, **pendapatan_marketplace[mp]}
            for mp in MARKETPLACE_LIST if mp in pendapatan_marketplace
        ]
        return total, breakdown

    @app.route("/pendapatan/penjualan")
    @admin_required
    def pendapatan_penjualan_dashboard():
        bulan = int(request.args.get("bulan", today_wib().month))
        tahun = int(request.args.get("tahun", today_wib().year))

        total, breakdown = hitung_ringkasan_penjualan(bulan, tahun)
        labarugi = hitung_labarugi_periode(bulan, tahun)

        laba_bersih_operasional = total["pendapatan"] - labarugi["total_beban_operasional"]

        return render_template(
            "pendapatan/penjualan_dashboard.html",
            marketplace_list=MARKETPLACE_LIST,
            bulan=bulan,
            tahun=tahun,
            total=total,
            breakdown=breakdown,
            iklan_by_mp=labarugi["iklan_by_mp"],
            gaji_total=labarugi["gaji_total"],
            opex_by_kategori=labarugi["opex_by_kategori"],
            total_beban_operasional=labarugi["total_beban_operasional"],
            laba_bersih_operasional=laba_bersih_operasional,
        )

    def _finalisasi_import_penjualan(data_tmp, mapping):
        """Proses+simpan data penjualan. Mengembalikan (berhasil, pesan)."""
        jumlah_baris_asli = len(data_tmp["rows"])
        agregat, dilewati, contoh_gagal, dilewati_tidak_wajar, contoh_tidak_wajar = proses_baris_penjualan(
            data_tmp["rows"], mapping
        )

        if not agregat:
            if dilewati_tidak_wajar and not dilewati:
                pesan = (
                    "Tidak ada baris data yang berhasil diproses — nilai di kolom Total Penjualan/Total Diskon yang "
                    "terdeteksi tidak wajar (kemungkinan salah kenali kolom, misalnya ikut mengenali kolom nomor "
                    "pesanan/ID sebagai jumlah uang)."
                )
                if contoh_tidak_wajar:
                    pesan += " Contoh nilai yang terbaca: " + ", ".join(f"'{c}'" for c in contoh_tidak_wajar) + "."
            else:
                pesan = "Tidak ada baris data yang berhasil diproses — kolom Tanggal yang terdeteksi sepertinya bukan tanggal yang valid."
                if contoh_gagal:
                    contoh = ", ".join(f"'{c}'" for c in contoh_gagal)
                    pesan += f" Contoh nilai di kolom itu: {contoh}."
            pesan += " Kirimkan daftar nama kolom di file ini ke admin supaya format file ini bisa didukung."
            return False, pesan

        marketplace = data_tmp["marketplace"]
        sumber_file = data_tmp.get("sumber_file", "")
        jumlah_pesanan_unik = sum(v["jumlah_pesanan"] for v in agregat.values())
        for tanggal, nilai in agregat.items():
            existing = PenjualanMarketplace.query.filter_by(marketplace=marketplace, tanggal=tanggal).first()
            if not existing:
                existing = PenjualanMarketplace(marketplace=marketplace, tanggal=tanggal)
                db.session.add(existing)
            existing.jumlah_pesanan = round(nilai["jumlah_pesanan"])
            existing.total_penjualan = round(nilai["total_penjualan"])
            existing.total_diskon = round(nilai["total_diskon"])
            existing.sumber_file = sumber_file
            existing.dibuat_pada = now_wib()
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            return False, (
                "Gagal menyimpan — ada angka yang tidak wajar di data. Kemungkinan sistem salah mengenali kolom "
                "(mis. kolom nomor pesanan/ID terbaca sebagai Total Penjualan). Kirimkan daftar nama kolom di file "
                "ini ke admin supaya format file ini bisa didukung."
            )

        tgl_min = min(agregat.keys())
        tgl_max = max(agregat.keys())

        def nama_kolom(key):
            idx = mapping.get(key)
            if idx is None or idx >= len(data_tmp["headers"]):
                return None
            return str(data_tmp["headers"][idx]).strip()

        keterangan_kolom = ", ".join(
            f"{label}='{nama_kolom(key)}'"
            for key, label, _wajib, _kk in KOLOM_TARGET_PENJUALAN
            if nama_kolom(key)
        )

        total_dilewati = dilewati + dilewati_tidak_wajar
        if total_dilewati == 0:
            pesan = (
                f"Berhasil impor semua {jumlah_baris_asli} baris dari file — seluruhnya berhasil diproses, "
                f"tidak ada yang terlewat. Tersimpan {jumlah_pesanan_unik} pesanan unik pada {len(agregat)} hari "
                f"({tgl_min.strftime('%d/%m/%Y')} - {tgl_max.strftime('%d/%m/%Y')}) untuk {marketplace}."
            )
        else:
            pesan = (
                f"Impor selesai: {jumlah_baris_asli} baris dibaca dari file, {jumlah_pesanan_unik} pesanan unik "
                f"berhasil disimpan pada {len(agregat)} hari ({tgl_min.strftime('%d/%m/%Y')} - "
                f"{tgl_max.strftime('%d/%m/%Y')}) untuk {marketplace}."
            )
            if dilewati:
                pesan += f" {dilewati} baris dilewati karena tanggal tidak terbaca."
                if contoh_gagal:
                    pesan += " Contoh: " + ", ".join(f"'{c}'" for c in contoh_gagal) + "."
            if dilewati_tidak_wajar:
                pesan += f" {dilewati_tidak_wajar} baris dilewati karena angkanya tidak wajar (kemungkinan salah kolom)."
                if contoh_tidak_wajar:
                    pesan += " Contoh: " + ", ".join(f"'{c}'" for c in contoh_tidak_wajar) + "."
        if keterangan_kolom:
            pesan += f" Kolom yang dipakai: {keterangan_kolom}."

        return True, (pesan, tgl_min)

    @app.route("/pendapatan/penjualan/upload", methods=["GET", "POST"])
    @admin_required
    def pendapatan_penjualan_upload():
        bersihkan_tmp_iklan_lama()
        if request.method == "POST":
            marketplace = request.form.get("marketplace", "")
            file = request.files.get("file")
            if marketplace not in MARKETPLACE_LIST:
                flash("Pilih marketplace terlebih dahulu.", "danger")
                return redirect(url_for("pendapatan_penjualan_upload"))
            if not file or not file.filename:
                flash("Pilih file laporan penjualan (CSV/XLSX) terlebih dahulu.", "danger")
                return redirect(url_for("pendapatan_penjualan_upload"))

            token, headers, rows_bersih, idx_header, error = simpan_tmp_upload("penjualan", marketplace, file)
            if error:
                flash(error, "danger")
                return redirect(url_for("pendapatan_penjualan_upload"))

            path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"penjualan_{token}.json")
            data_tmp = {
                "marketplace": marketplace,
                "headers": headers,
                "rows": rows_bersih,
                "sumber_file": secure_filename(file.filename),
            }

            nama_profil, mapping = cocokkan_profil_header_penjualan(headers)
            if not nama_profil:
                mapping = deteksi_otomatis_kolom_penjualan(headers, rows_bersih)
                if mapping.get("tanggal") is not None and mapping.get("total_penjualan") is not None:
                    nama_profil = "deteksi otomatis dari isi data"

            if not nama_profil:
                os.remove(path_tmp)
                flash(
                    "Sistem tidak berhasil mengenali kolom Tanggal dan Total Penjualan secara otomatis dari file "
                    f"ini. Header yang terbaca: {', '.join(str(h) for h in headers)}. Kirimkan daftar kolom ini ke "
                    "admin supaya format file ini bisa didukung.",
                    "danger",
                )
                return redirect(url_for("pendapatan_penjualan_upload"))

            berhasil, hasil = _finalisasi_import_penjualan(data_tmp, mapping)
            os.remove(path_tmp)
            if berhasil:
                pesan, tgl_min = hasil
                flash(f"Format kolom dikenali otomatis ({nama_profil}). {pesan}", "success")
                return redirect(url_for("pendapatan_penjualan_dashboard", bulan=tgl_min.month, tahun=tgl_min.year))

            flash(hasil, "danger")
            return redirect(url_for("pendapatan_penjualan_upload"))

        return render_template("pendapatan/penjualan_upload.html", marketplace_list=MARKETPLACE_LIST)

    @app.route("/pendapatan/penjualan/manual", methods=["POST"])
    @admin_required
    def pendapatan_penjualan_manual():
        marketplace = request.form.get("marketplace", "")
        try:
            tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
        except ValueError:
            tanggal = None

        if marketplace not in MARKETPLACE_LIST or not tanggal:
            flash("Marketplace dan tanggal wajib diisi dengan benar.", "danger")
            return redirect(url_for("pendapatan_penjualan_dashboard"))

        existing = PenjualanMarketplace.query.filter_by(marketplace=marketplace, tanggal=tanggal).first()
        if not existing:
            existing = PenjualanMarketplace(marketplace=marketplace, tanggal=tanggal)
            db.session.add(existing)
        existing.jumlah_pesanan = round(parse_angka_iklan(request.form.get("jumlah_pesanan", "0")))
        existing.total_penjualan = round(parse_angka_iklan(request.form.get("total_penjualan", "0")))
        existing.total_diskon = round(parse_angka_iklan(request.form.get("total_diskon", "0")))
        existing.sumber_file = "Input manual"
        existing.dibuat_pada = now_wib()
        db.session.commit()
        flash(f"Data penjualan {marketplace} tanggal {tanggal.strftime('%d/%m/%Y')} berhasil disimpan.", "success")
        return redirect(url_for("pendapatan_penjualan_dashboard", bulan=tanggal.month, tahun=tanggal.year))

    @app.route("/pendapatan/penjualan/hapus/<int:pid>", methods=["POST"])
    @admin_required
    def pendapatan_penjualan_hapus(pid):
        data = db.session.get(PenjualanMarketplace, pid)
        if data:
            bulan, tahun = data.tanggal.month, data.tanggal.year
            db.session.delete(data)
            db.session.commit()
            flash("Data penjualan berhasil dihapus.", "success")
            return redirect(url_for("pendapatan_penjualan_dashboard", bulan=bulan, tahun=tahun))
        flash("Data tidak ditemukan.", "danger")
        return redirect(url_for("pendapatan_penjualan_dashboard"))

    # ---------- PENGATURAN ----------
    @app.route("/pengaturan", methods=["GET", "POST"])
    @admin_required
    def pengaturan():
        settings = get_settings()
        if request.method == "POST":
            settings.nama_perusahaan = request.form.get("nama_perusahaan", "").strip()
            settings.jam_masuk_standar = request.form.get("jam_masuk_standar")
            settings.jam_pulang_standar = request.form.get("jam_pulang_standar")
            settings.jam_masuk_standar_freelance = request.form.get("jam_masuk_standar_freelance")
            settings.jam_pulang_standar_freelance = request.form.get("jam_pulang_standar_freelance")
            settings.hari_kerja_per_bulan = int(request.form.get("hari_kerja_per_bulan") or 22)
            settings.toleransi_telat_menit = int(request.form.get("toleransi_telat_menit") or 0)
            settings.toleransi_lembur_menit = int(request.form.get("toleransi_lembur_menit") or 30)
            settings.denda_telat_per_menit = int(request.form.get("denda_telat_per_menit") or 0)
            settings.upah_lembur_per_jam = int(request.form.get("upah_lembur_per_jam") or 0)
            settings.tarif_harian_freelance = int(request.form.get("tarif_harian_freelance") or 0)
            settings.upah_lembur_freelance_per_jam = int(request.form.get("upah_lembur_freelance_per_jam") or 0)
            settings.tarif_co_host = int(request.form.get("tarif_co_host") or 0)
            settings.bonus_target_tercapai = int(request.form.get("bonus_target_tercapai") or 0)
            settings.alamat_perusahaan = request.form.get("alamat_perusahaan", "").strip()
            settings.no_hp_perusahaan = request.form.get("no_hp_perusahaan", "").strip()
            settings.email_perusahaan = request.form.get("email_perusahaan", "").strip()
            settings.instagram_perusahaan = request.form.get("instagram_perusahaan", "").strip()
            settings.kota_perusahaan = request.form.get("kota_perusahaan", "").strip()

            kantor_lat = request.form.get("kantor_lat", "").strip()
            kantor_lng = request.form.get("kantor_lng", "").strip()
            settings.kantor_lat = float(kantor_lat) if kantor_lat else None
            settings.kantor_lng = float(kantor_lng) if kantor_lng else None
            settings.radius_kantor_meter = int(request.form.get("radius_kantor_meter") or 100)

            logo_file = request.files.get("logo")
            if logo_file and logo_file.filename:
                ext = logo_file.filename.rsplit(".", 1)[-1].lower()
                if ext in ("png", "jpg", "jpeg", "webp"):
                    filename = secure_filename(f"logo.{ext}")
                    logo_file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
                    settings.logo_filename = filename
                else:
                    flash("Format logo harus PNG, JPG, atau WEBP.", "danger")

            db.session.commit()
            flash("Pengaturan berhasil disimpan.", "success")
            return redirect(url_for("pengaturan"))
        return render_template("settings.html", settings=settings)

    # ---------- MARKETING: ANALISA IKLAN PER MARKETPLACE ----------
    def bersihkan_tmp_iklan_lama():
        batas = time.time() - 24 * 3600
        try:
            for nama in os.listdir(app.config["TMP_IKLAN_FOLDER"]):
                path_file = os.path.join(app.config["TMP_IKLAN_FOLDER"], nama)
                if os.path.isfile(path_file) and os.path.getmtime(path_file) < batas:
                    os.remove(path_file)
        except OSError:
            pass

    @app.route("/marketing/iklan")
    @marketing_required
    def marketing_iklan_dashboard():
        marketplace_filter = request.args.get("marketplace", "Semua")
        hari_ini = today_wib()
        default_dari = hari_ini.replace(day=1)
        try:
            dari = datetime.strptime(request.args.get("dari", ""), "%Y-%m-%d").date()
        except ValueError:
            dari = default_dari
        try:
            sampai = datetime.strptime(request.args.get("sampai", ""), "%Y-%m-%d").date()
        except ValueError:
            sampai = hari_ini
        try:
            threshold_roas = float(request.args.get("roas_min", 3))
        except ValueError:
            threshold_roas = 3.0
        try:
            threshold_ctr = float(request.args.get("ctr_min", 1))
        except ValueError:
            threshold_ctr = 1.0

        query = IklanMarketplace.query.filter(
            IklanMarketplace.tanggal >= dari, IklanMarketplace.tanggal <= sampai
        )
        semua_data = query.order_by(IklanMarketplace.tanggal).all()

        if marketplace_filter != "Semua":
            data_terfilter = [d for d in semua_data if d.marketplace == marketplace_filter]
        else:
            data_terfilter = semua_data

        def totalkan(items):
            return {
                "biaya": sum(d.biaya or 0 for d in items),
                "impresi": sum(d.impresi or 0 for d in items),
                "klik": sum(d.klik or 0 for d in items),
                "pesanan": sum(d.pesanan or 0 for d in items),
                "omzet": sum(d.omzet or 0 for d in items),
            }

        total = totalkan(data_terfilter)
        ringkasan = {
            **total,
            "roas": (total["omzet"] / total["biaya"]) if total["biaya"] else 0,
            "ctr": (total["klik"] / total["impresi"] * 100) if total["impresi"] else 0,
            "cpc": (total["biaya"] / total["klik"]) if total["klik"] else 0,
            "cpa": (total["biaya"] / total["pesanan"]) if total["pesanan"] else 0,
            "konversi_rate": (total["pesanan"] / total["klik"] * 100) if total["klik"] else 0,
        }

        def hitung_tren_arah(item_mp):
            tanggal_unik = sorted(set(d.tanggal for d in item_mp))
            if len(tanggal_unik) < 4:
                return "Data belum cukup"
            tengah = len(tanggal_unik) // 2
            tanggal_awal = set(tanggal_unik[:tengah])
            tanggal_akhir = set(tanggal_unik[tengah:])
            t_awal = totalkan([d for d in item_mp if d.tanggal in tanggal_awal])
            t_akhir = totalkan([d for d in item_mp if d.tanggal in tanggal_akhir])
            if not t_awal["biaya"] or not t_akhir["biaya"]:
                return "Data belum cukup"
            roas_awal = t_awal["omzet"] / t_awal["biaya"]
            roas_akhir = t_akhir["omzet"] / t_akhir["biaya"]
            if roas_akhir >= roas_awal * 1.15:
                return "naik"
            if roas_akhir <= roas_awal * 0.85:
                return "turun"
            return "stabil"

        def buat_insight(t, ctr, tren_arah):
            roas = (t["omzet"] / t["biaya"]) if t["biaya"] else 0
            if roas < threshold_roas * 0.5:
                verdict, badge = "Sangat Kurang Efektif", "danger"
                rekomendasi = (
                    f"ROAS {roas:.2f}x jauh di bawah target {threshold_roas:.1f}x — iklan cenderung merugi. "
                    "Pertimbangkan hentikan sementara atau turunkan drastis budgetnya, lalu evaluasi ulang "
                    "produk, harga, atau target audiens sebelum lanjut beriklan."
                )
            elif roas < threshold_roas:
                verdict, badge = "Kurang Efektif", "warning"
                rekomendasi = (
                    f"ROAS {roas:.2f}x masih di bawah target {threshold_roas:.1f}x. Perbaiki dulu materi iklan "
                    "atau targeting sebelum menambah budget — jangan naikkan budget selama ROAS belum sesuai target."
                )
            elif tren_arah == "turun":
                verdict, badge = "Waspada — Tren Menurun", "warning"
                rekomendasi = (
                    f"ROAS {roas:.2f}x masih di atas target, tapi cenderung menurun dibanding paruh awal periode ini. "
                    "Pantau ketat dan tunda penambahan budget sampai tren membaik lagi."
                )
            else:
                verdict, badge = "Bagus", "success"
                tren_txt = " dan tren membaik" if tren_arah == "naik" else ""
                rekomendasi = (
                    f"ROAS {roas:.2f}x sudah di atas target {threshold_roas:.1f}x{tren_txt}. Pertimbangkan naikkan "
                    "budget bertahap (+20-30%) untuk memaksimalkan momentum penjualan."
                )
            if ctr and ctr < threshold_ctr:
                rekomendasi += (
                    f" Catatan tambahan: CTR {ctr:.2f}% tergolong rendah (di bawah {threshold_ctr:.1f}%), "
                    "materi iklan mungkin kurang menarik untuk diklik."
                )
            return verdict, badge, rekomendasi

        breakdown = []
        for mp in MARKETPLACE_LIST:
            item_mp = [d for d in semua_data if d.marketplace == mp]
            if not item_mp:
                continue
            t = totalkan(item_mp)
            ctr = (t["klik"] / t["impresi"] * 100) if t["impresi"] else 0
            entri = {
                "marketplace": mp,
                **t,
                "roas": (t["omzet"] / t["biaya"]) if t["biaya"] else 0,
                "ctr": ctr,
                "cpa": (t["biaya"] / t["pesanan"]) if t["pesanan"] else 0,
            }
            if t["biaya"] > 0:
                tren_arah = hitung_tren_arah(item_mp)
                verdict, badge, rekomendasi = buat_insight(t, ctr, tren_arah)
                entri.update(tren_arah=tren_arah, verdict=verdict, badge=badge, rekomendasi=rekomendasi)
            else:
                entri.update(tren_arah=None, verdict=None, badge=None, rekomendasi=None)
            breakdown.append(entri)

        tren_map = {}
        for d in data_terfilter:
            key = d.tanggal.isoformat()
            if key not in tren_map:
                tren_map[key] = {"biaya": 0, "omzet": 0, "klik": 0}
            tren_map[key]["biaya"] += d.biaya or 0
            tren_map[key]["omzet"] += d.omzet or 0
            tren_map[key]["klik"] += d.klik or 0
        tren_tanggal = sorted(tren_map.keys())
        tren_biaya = [tren_map[k]["biaya"] for k in tren_tanggal]
        tren_omzet = [tren_map[k]["omzet"] for k in tren_tanggal]

        return render_template(
            "marketing/iklan_dashboard.html",
            marketplace_list=MARKETPLACE_LIST,
            marketplace_filter=marketplace_filter,
            dari=dari,
            sampai=sampai,
            threshold_roas=threshold_roas,
            threshold_ctr=threshold_ctr,
            ringkasan=ringkasan,
            breakdown=breakdown,
            data_list=list(reversed(data_terfilter)),
            tren_tanggal=tren_tanggal,
            tren_biaya=tren_biaya,
            tren_omzet=tren_omzet,
        )

    def simpan_tmp_upload(prefix, marketplace, file):
        headers, rows, idx_header, error = baca_file_iklan(file)
        if error:
            return None, None, None, None, error
        rows_bersih = []
        for row in rows:
            row_lengkap = list(row) + [""] * (len(headers) - len(row))
            row_serial = [
                c.isoformat() if isinstance(c, (datetime, date)) else c
                for c in row_lengkap[: len(headers)]
            ]
            rows_bersih.append(row_serial)
        token = uuid.uuid4().hex
        path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"{prefix}_{token}.json")
        with open(path_tmp, "w", encoding="utf-8") as f:
            json.dump({
                "marketplace": marketplace,
                "headers": headers,
                "rows": rows_bersih,
                "sumber_file": secure_filename(file.filename),
            }, f)
        return token, headers, rows_bersih, idx_header, None

    def proses_baris_upload(rows, mapping, butuh_produk):
        agregat = {}
        dilewati = 0
        contoh_gagal_tanggal = []
        for row in rows:
            def ambil(key):
                idx = mapping.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            nama_produk = None
            if butuh_produk:
                raw_produk = ambil("nama_produk")
                nama_produk = str(raw_produk).strip() if raw_produk is not None else ""

            tanggal_raw = ambil("tanggal")
            tanggal = parse_tanggal_iklan(tanggal_raw)
            if not tanggal or (butuh_produk and not nama_produk):
                dilewati += 1
                if not tanggal and tanggal_raw not in (None, ""):
                    nilai_str = str(tanggal_raw).strip()
                    if nilai_str and nilai_str not in contoh_gagal_tanggal and len(contoh_gagal_tanggal) < 5:
                        contoh_gagal_tanggal.append(nilai_str)
                continue

            kunci = (nama_produk, tanggal) if butuh_produk else tanggal
            if kunci not in agregat:
                agregat[kunci] = {"biaya": 0, "impresi": 0, "klik": 0, "pesanan": 0, "omzet": 0, "pajak": 0}
            for k in ("biaya", "impresi", "klik", "pesanan", "omzet", "pajak"):
                nilai_mentah = ambil(k)
                if nilai_mentah is not None:
                    agregat[kunci][k] += parse_angka_iklan(nilai_mentah)

        return agregat, dilewati, contoh_gagal_tanggal

    @app.route("/marketing/iklan/upload", methods=["GET", "POST"])
    @marketing_required
    def marketing_iklan_upload():
        bersihkan_tmp_iklan_lama()
        if request.method == "POST":
            marketplace = request.form.get("marketplace", "")
            file = request.files.get("file")
            if marketplace not in MARKETPLACE_LIST:
                flash("Pilih marketplace terlebih dahulu.", "danger")
                return redirect(url_for("marketing_iklan_upload"))
            if not file or not file.filename:
                flash("Pilih file laporan iklan (CSV/XLSX) terlebih dahulu.", "danger")
                return redirect(url_for("marketing_iklan_upload"))

            token, headers, rows_bersih, idx_header, error = simpan_tmp_upload("iklan", marketplace, file)
            if error:
                flash(error, "danger")
                return redirect(url_for("marketing_iklan_upload"))
            path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"iklan_{token}.json")

            mapping = deteksi_otomatis_kolom_iklan(headers, rows_bersih, KOLOM_TARGET_IKLAN)
            if mapping.get("tanggal") is None or mapping.get("biaya") is None:
                os.remove(path_tmp)
                flash(
                    "Sistem tidak berhasil mengenali kolom Tanggal dan/atau Biaya Iklan secara otomatis dari file "
                    f"ini. Header yang terbaca: {', '.join(str(h) for h in headers)}. Kirimkan daftar kolom ini "
                    "biar formatnya bisa didukung.",
                    "danger",
                )
                return redirect(url_for("marketing_iklan_upload"))

            agregat, dilewati, contoh_gagal = proses_baris_upload(rows_bersih, mapping, butuh_produk=False)
            if not agregat:
                os.remove(path_tmp)
                pesan = "Kolom Tanggal terdeteksi tapi isinya sepertinya bukan tanggal yang valid."
                if contoh_gagal:
                    pesan += " Contoh nilai: " + ", ".join(f"'{c}'" for c in contoh_gagal) + "."
                flash(pesan, "danger")
                return redirect(url_for("marketing_iklan_upload"))

            def nama_kolom(idx):
                return str(headers[idx]).strip() if idx is not None and idx < len(headers) else None

            kolom_terdeteksi = [(label, nama_kolom(mapping.get(key))) for key, label, _w, _kk in KOLOM_TARGET_IKLAN]
            preview = []
            for tanggal in sorted(agregat.keys()):
                nilai = agregat[tanggal]
                preview.append({
                    "tanggal": tanggal, "biaya": round(nilai["biaya"]), "impresi": round(nilai["impresi"]),
                    "klik": round(nilai["klik"]), "pesanan": round(nilai["pesanan"]), "omzet": round(nilai["omzet"]),
                })

            return render_template(
                "marketing/review_kolom.html",
                judul=f"Review Data Iklan — {marketplace}",
                token=token,
                mapping=mapping,
                kolom_terdeteksi=kolom_terdeteksi,
                preview=preview,
                tampilkan_produk=False,
                tampilkan_pajak=False,
                dilewati=dilewati,
                contoh_gagal=contoh_gagal,
                konfirmasi_url=url_for("marketing_iklan_konfirmasi"),
                upload_url=url_for("marketing_iklan_upload"),
            )

        return render_template("marketing/iklan_upload.html", marketplace_list=MARKETPLACE_LIST)

    @app.route("/marketing/iklan/konfirmasi", methods=["POST"])
    @marketing_required
    def marketing_iklan_konfirmasi():
        token = request.form.get("token", "")
        path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"iklan_{token}.json")
        if not os.path.isfile(path_tmp):
            flash("Sesi upload sudah kedaluwarsa, silakan upload ulang file.", "danger")
            return redirect(url_for("marketing_iklan_upload"))

        with open(path_tmp, "r", encoding="utf-8") as f:
            data_tmp = json.load(f)

        mapping = {}
        for key, label, wajib, _kk in KOLOM_TARGET_IKLAN:
            nilai = request.form.get(f"map_{key}", "")
            mapping[key] = int(nilai) if nilai != "" else None
            if wajib and mapping[key] is None:
                flash(f"Kolom '{label}' wajib dipilih.", "danger")
                os.remove(path_tmp)
                return redirect(url_for("marketing_iklan_upload"))

        agregat, dilewati, contoh_gagal = proses_baris_upload(data_tmp["rows"], mapping, butuh_produk=False)

        if not agregat:
            os.remove(path_tmp)
            pesan = "Tidak ada baris data yang berhasil diproses — kolom Tanggal sepertinya bukan tanggal yang valid."
            if contoh_gagal:
                contoh = ", ".join(f"'{c}'" for c in contoh_gagal)
                pesan += f" Contoh nilai di kolom itu: {contoh}."
            flash(pesan, "danger")
            return redirect(url_for("marketing_iklan_upload"))

        marketplace = data_tmp["marketplace"]
        sumber_file = data_tmp.get("sumber_file", "")
        for tanggal, nilai in agregat.items():
            existing = IklanMarketplace.query.filter_by(marketplace=marketplace, tanggal=tanggal).first()
            if not existing:
                existing = IklanMarketplace(marketplace=marketplace, tanggal=tanggal)
                db.session.add(existing)
            existing.biaya = round(nilai["biaya"])
            existing.impresi = round(nilai["impresi"])
            existing.klik = round(nilai["klik"])
            existing.pesanan = round(nilai["pesanan"])
            existing.omzet = round(nilai["omzet"])
            existing.sumber_file = sumber_file
            existing.dibuat_pada = now_wib()
        db.session.commit()
        os.remove(path_tmp)

        tgl_min = min(agregat.keys())
        tgl_max = max(agregat.keys())
        pesan = f"Berhasil impor {len(agregat)} hari data iklan {marketplace} ({tgl_min.strftime('%d/%m/%Y')} - {tgl_max.strftime('%d/%m/%Y')})."
        if dilewati:
            pesan += f" {dilewati} baris dilewati karena tanggal tidak terbaca."
            if contoh_gagal:
                pesan += " Contoh: " + ", ".join(f"'{c}'" for c in contoh_gagal) + "."
        flash(pesan, "success")

        return redirect(url_for("marketing_iklan_dashboard", marketplace=marketplace))

    @app.route("/marketing/iklan/manual", methods=["POST"])
    @marketing_required
    def marketing_iklan_manual():
        marketplace = request.form.get("marketplace", "")
        try:
            tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
        except ValueError:
            tanggal = None

        if marketplace not in MARKETPLACE_LIST or not tanggal:
            flash("Marketplace dan tanggal wajib diisi dengan benar.", "danger")
            return redirect(url_for("marketing_iklan_dashboard"))

        existing = IklanMarketplace.query.filter_by(marketplace=marketplace, tanggal=tanggal).first()
        if not existing:
            existing = IklanMarketplace(marketplace=marketplace, tanggal=tanggal)
            db.session.add(existing)
        existing.biaya = round(parse_angka_iklan(request.form.get("biaya", "0")))
        existing.impresi = round(parse_angka_iklan(request.form.get("impresi", "0")))
        existing.klik = round(parse_angka_iklan(request.form.get("klik", "0")))
        existing.pesanan = round(parse_angka_iklan(request.form.get("pesanan", "0")))
        existing.omzet = round(parse_angka_iklan(request.form.get("omzet", "0")))
        existing.sumber_file = "Input manual"
        existing.dibuat_pada = now_wib()
        db.session.commit()
        flash(f"Data iklan {marketplace} tanggal {tanggal.strftime('%d/%m/%Y')} berhasil disimpan.", "success")
        return redirect(url_for("marketing_iklan_dashboard", marketplace=marketplace))

    @app.route("/marketing/iklan/hapus/<int:iklan_id>", methods=["POST"])
    @marketing_required
    def marketing_iklan_hapus(iklan_id):
        data = db.session.get(IklanMarketplace, iklan_id)
        if data:
            marketplace = data.marketplace
            db.session.delete(data)
            db.session.commit()
            flash("Data iklan berhasil dihapus.", "success")
            return redirect(url_for("marketing_iklan_dashboard", marketplace=marketplace))
        flash("Data tidak ditemukan.", "danger")
        return redirect(url_for("marketing_iklan_dashboard"))

    # ---------- MARKETING: IKLAN META ----------
    @app.route("/marketing/meta")
    @marketing_required
    def marketing_meta_dashboard():
        # Sengaja tanpa filter tanggal/threshold apa pun -- tampilkan semua data yang
        # pernah diupload/diinput user apa adanya, sebagai data pendukung Laba Rugi.
        semua_data = IklanMeta.query.order_by(IklanMeta.tanggal).all()

        total = {
            "biaya": sum(d.biaya or 0 for d in semua_data),
            "pajak": sum(d.pajak or 0 for d in semua_data),
            "impresi": sum(d.impresi or 0 for d in semua_data),
            "klik": sum(d.klik or 0 for d in semua_data),
            "pesanan": sum(d.pesanan or 0 for d in semua_data),
            "omzet": sum(d.omzet or 0 for d in semua_data),
        }
        ringkasan = {
            **total,
            "roas": (total["omzet"] / total["biaya"]) if total["biaya"] else 0,
            "ctr": (total["klik"] / total["impresi"] * 100) if total["impresi"] else 0,
            "cpc": (total["biaya"] / total["klik"]) if total["klik"] else 0,
            "cpa": (total["biaya"] / total["pesanan"]) if total["pesanan"] else 0,
            "konversi_rate": (total["pesanan"] / total["klik"] * 100) if total["klik"] else 0,
        }

        tren_map = {}
        for d in semua_data:
            key = d.tanggal.isoformat()
            if key not in tren_map:
                tren_map[key] = {"biaya": 0, "omzet": 0}
            tren_map[key]["biaya"] += d.biaya or 0
            tren_map[key]["omzet"] += d.omzet or 0
        tren_tanggal = sorted(tren_map.keys())
        tren_biaya = [tren_map[k]["biaya"] for k in tren_tanggal]
        tren_omzet = [tren_map[k]["omzet"] for k in tren_tanggal]

        return render_template(
            "marketing/meta_dashboard.html",
            ringkasan=ringkasan,
            data_list=list(reversed(semua_data)),
            tren_tanggal=tren_tanggal,
            tren_biaya=tren_biaya,
            tren_omzet=tren_omzet,
            tanggal_default=today_wib().isoformat(),
        )

    @app.route("/marketing/meta/upload", methods=["GET", "POST"])
    @marketing_required
    def marketing_meta_upload():
        bersihkan_tmp_iklan_lama()
        if request.method == "POST":
            file = request.files.get("file")
            if not file or not file.filename:
                flash("Pilih file laporan Iklan Meta (CSV/XLSX) terlebih dahulu.", "danger")
                return redirect(url_for("marketing_meta_upload"))

            token, headers, rows_bersih, idx_header, error = simpan_tmp_upload("meta", "Meta", file)
            if error:
                flash(error, "danger")
                return redirect(url_for("marketing_meta_upload"))
            path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"meta_{token}.json")

            mapping = deteksi_otomatis_kolom_iklan(headers, rows_bersih, KOLOM_TARGET_META)
            if mapping.get("tanggal") is None or mapping.get("biaya") is None:
                os.remove(path_tmp)
                flash(
                    "Sistem tidak berhasil mengenali kolom Tanggal dan/atau Biaya Iklan secara otomatis dari file "
                    f"ini. Header yang terbaca: {', '.join(str(h) for h in headers)}. Kirimkan daftar kolom ini "
                    "biar formatnya bisa didukung.",
                    "danger",
                )
                return redirect(url_for("marketing_meta_upload"))

            agregat, dilewati, contoh_gagal = proses_baris_upload(rows_bersih, mapping, butuh_produk=False)
            if not agregat:
                os.remove(path_tmp)
                pesan = "Kolom Tanggal terdeteksi tapi isinya sepertinya bukan tanggal yang valid."
                if contoh_gagal:
                    pesan += " Contoh nilai: " + ", ".join(f"'{c}'" for c in contoh_gagal) + "."
                flash(pesan, "danger")
                return redirect(url_for("marketing_meta_upload"))

            def nama_kolom(key):
                idx = mapping.get(key)
                return str(headers[idx]).strip() if idx is not None and idx < len(headers) else None

            kolom_terdeteksi = [
                (label, nama_kolom(key)) for key, label, _wajib, _kk in KOLOM_TARGET_META
            ]
            preview = []
            for tanggal in sorted(agregat.keys()):
                nilai = agregat[tanggal]
                biaya = round(nilai["biaya"])
                pajak = round(nilai["pajak"]) or hitung_pajak_meta(biaya)
                preview.append({
                    "tanggal": tanggal, "biaya": biaya, "pajak": pajak,
                    "impresi": round(nilai["impresi"]), "klik": round(nilai["klik"]),
                    "pesanan": round(nilai["pesanan"]), "omzet": round(nilai["omzet"]),
                })

            return render_template(
                "marketing/review_kolom.html",
                judul="Review Data Iklan Meta",
                token=token,
                mapping=mapping,
                kolom_terdeteksi=kolom_terdeteksi,
                preview=preview,
                tampilkan_produk=False,
                tampilkan_pajak=True,
                dilewati=dilewati,
                contoh_gagal=contoh_gagal,
                konfirmasi_url=url_for("marketing_meta_konfirmasi"),
                upload_url=url_for("marketing_meta_upload"),
            )

        return render_template("marketing/meta_upload.html")

    @app.route("/marketing/meta/konfirmasi", methods=["POST"])
    @marketing_required
    def marketing_meta_konfirmasi():
        token = request.form.get("token", "")
        path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"meta_{token}.json")
        if not os.path.isfile(path_tmp):
            flash("Sesi upload sudah kedaluwarsa, silakan upload ulang file.", "danger")
            return redirect(url_for("marketing_meta_upload"))

        with open(path_tmp, "r", encoding="utf-8") as f:
            data_tmp = json.load(f)

        mapping = {}
        for key, label, wajib, _kk in KOLOM_TARGET_META:
            nilai = request.form.get(f"map_{key}", "")
            mapping[key] = int(nilai) if nilai != "" else None
            if wajib and mapping[key] is None:
                flash(f"Kolom '{label}' wajib dipilih.", "danger")
                os.remove(path_tmp)
                return redirect(url_for("marketing_meta_upload"))

        agregat, dilewati, contoh_gagal = proses_baris_upload(data_tmp["rows"], mapping, butuh_produk=False)

        if not agregat:
            os.remove(path_tmp)
            pesan = "Tidak ada baris data yang berhasil diproses — kolom Tanggal sepertinya bukan tanggal yang valid."
            if contoh_gagal:
                contoh = ", ".join(f"'{c}'" for c in contoh_gagal)
                pesan += f" Contoh nilai di kolom itu: {contoh}."
            flash(pesan, "danger")
            return redirect(url_for("marketing_meta_upload"))

        sumber_file = data_tmp.get("sumber_file", "")
        for tanggal, nilai in agregat.items():
            existing = IklanMeta.query.filter_by(tanggal=tanggal).first()
            if not existing:
                existing = IklanMeta(tanggal=tanggal)
                db.session.add(existing)
            existing.biaya = round(nilai["biaya"])
            existing.pajak = round(nilai["pajak"]) or hitung_pajak_meta(nilai["biaya"])
            existing.impresi = round(nilai["impresi"])
            existing.klik = round(nilai["klik"])
            existing.pesanan = round(nilai["pesanan"])
            existing.omzet = round(nilai["omzet"])
            existing.sumber_file = sumber_file
            existing.dibuat_pada = now_wib()
        db.session.commit()
        os.remove(path_tmp)

        tgl_min = min(agregat.keys())
        tgl_max = max(agregat.keys())
        pesan = f"Berhasil impor {len(agregat)} hari data Iklan Meta ({tgl_min.strftime('%d/%m/%Y')} - {tgl_max.strftime('%d/%m/%Y')})."
        if dilewati:
            pesan += f" {dilewati} baris dilewati karena tanggal tidak terbaca."
            if contoh_gagal:
                pesan += " Contoh: " + ", ".join(f"'{c}'" for c in contoh_gagal) + "."
        flash(pesan, "success")

        return redirect(url_for("marketing_meta_dashboard"))

    @app.route("/marketing/meta/manual", methods=["POST"])
    @marketing_required
    def marketing_meta_manual():
        try:
            tanggal = datetime.strptime(request.form.get("tanggal", ""), "%Y-%m-%d").date()
        except ValueError:
            tanggal = None

        if not tanggal:
            flash("Tanggal wajib diisi dengan benar.", "danger")
            return redirect(url_for("marketing_meta_dashboard"))

        existing = IklanMeta.query.filter_by(tanggal=tanggal).first()
        if not existing:
            existing = IklanMeta(tanggal=tanggal)
            db.session.add(existing)
        existing.biaya = round(parse_angka_iklan(request.form.get("biaya", "0")))
        existing.pajak = round(parse_angka_iklan(request.form.get("pajak", "0"))) or hitung_pajak_meta(existing.biaya)
        existing.impresi = round(parse_angka_iklan(request.form.get("impresi", "0")))
        existing.klik = round(parse_angka_iklan(request.form.get("klik", "0")))
        existing.pesanan = round(parse_angka_iklan(request.form.get("pesanan", "0")))
        existing.omzet = round(parse_angka_iklan(request.form.get("omzet", "0")))
        existing.sumber_file = "Input manual"
        existing.dibuat_pada = now_wib()
        db.session.commit()
        flash(f"Data Iklan Meta tanggal {tanggal.strftime('%d/%m/%Y')} berhasil disimpan.", "success")
        return redirect(url_for("marketing_meta_dashboard"))

    @app.route("/marketing/meta/hapus/<int:meta_id>", methods=["POST"])
    @marketing_required
    def marketing_meta_hapus(meta_id):
        data = db.session.get(IklanMeta, meta_id)
        if data:
            db.session.delete(data)
            db.session.commit()
            flash("Data Iklan Meta berhasil dihapus.", "success")
        else:
            flash("Data tidak ditemukan.", "danger")
        return redirect(url_for("marketing_meta_dashboard"))

    @app.route("/marketing/meta/hitung-ulang-pajak", methods=["POST"])
    @marketing_required
    def marketing_meta_hitung_ulang_pajak():
        semua = IklanMeta.query.all()
        jumlah = 0
        for r in semua:
            pajak_baru = hitung_pajak_meta(r.biaya)
            if r.pajak != pajak_baru:
                r.pajak = pajak_baru
                jumlah += 1
        db.session.commit()
        flash(
            f"Pajak Iklan dihitung ulang ({int(TARIF_PAJAK_META * 100)}% dari Biaya) untuk {jumlah} dari {len(semua)} baris data.",
            "success",
        )
        return redirect(url_for("marketing_meta_dashboard"))

    @app.route("/marketing/produk/upload", methods=["GET", "POST"])
    @marketing_required
    def marketing_produk_upload():
        bersihkan_tmp_iklan_lama()
        if request.method == "POST":
            marketplace = request.form.get("marketplace", "")
            file = request.files.get("file")
            if marketplace not in MARKETPLACE_LIST:
                flash("Pilih marketplace terlebih dahulu.", "danger")
                return redirect(url_for("marketing_produk_upload"))
            if not file or not file.filename:
                flash("Pilih file laporan iklan per produk (CSV/XLSX) terlebih dahulu.", "danger")
                return redirect(url_for("marketing_produk_upload"))

            token, headers, rows_bersih, idx_header, error = simpan_tmp_upload("produk", marketplace, file)
            if error:
                flash(error, "danger")
                return redirect(url_for("marketing_produk_upload"))
            path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"produk_{token}.json")

            mapping = deteksi_otomatis_kolom_iklan(headers, rows_bersih, KOLOM_TARGET_PRODUK)
            if mapping.get("nama_produk") is None or mapping.get("tanggal") is None or mapping.get("biaya") is None:
                os.remove(path_tmp)
                flash(
                    "Sistem tidak berhasil mengenali kolom Nama Produk/Tanggal/Biaya Iklan secara otomatis dari "
                    f"file ini. Header yang terbaca: {', '.join(str(h) for h in headers)}. Kirimkan daftar kolom "
                    "ini biar formatnya bisa didukung.",
                    "danger",
                )
                return redirect(url_for("marketing_produk_upload"))

            agregat, dilewati, contoh_gagal = proses_baris_upload(rows_bersih, mapping, butuh_produk=True)
            if not agregat:
                os.remove(path_tmp)
                pesan = "Kolom Nama Produk/Tanggal terdeteksi tapi isinya sepertinya tidak valid."
                if contoh_gagal:
                    pesan += " Contoh nilai Tanggal: " + ", ".join(f"'{c}'" for c in contoh_gagal) + "."
                flash(pesan, "danger")
                return redirect(url_for("marketing_produk_upload"))

            def nama_kolom(idx):
                return str(headers[idx]).strip() if idx is not None and idx < len(headers) else None

            kolom_terdeteksi = [(label, nama_kolom(mapping.get(key))) for key, label, _w, _kk in KOLOM_TARGET_PRODUK]
            preview = []
            for (nama_produk, tanggal) in sorted(agregat.keys(), key=lambda k: (k[1], k[0])):
                nilai = agregat[(nama_produk, tanggal)]
                preview.append({
                    "nama_produk": nama_produk, "tanggal": tanggal, "biaya": round(nilai["biaya"]),
                    "impresi": round(nilai["impresi"]), "klik": round(nilai["klik"]),
                    "pesanan": round(nilai["pesanan"]), "omzet": round(nilai["omzet"]),
                })

            return render_template(
                "marketing/review_kolom.html",
                judul=f"Review Data Performa Produk — {marketplace}",
                token=token,
                mapping=mapping,
                kolom_terdeteksi=kolom_terdeteksi,
                preview=preview,
                tampilkan_produk=True,
                tampilkan_pajak=False,
                dilewati=dilewati,
                contoh_gagal=contoh_gagal,
                konfirmasi_url=url_for("marketing_produk_konfirmasi"),
                upload_url=url_for("marketing_produk_upload"),
            )

        return render_template("marketing/produk_upload.html", marketplace_list=MARKETPLACE_LIST)

    @app.route("/marketing/produk/konfirmasi", methods=["POST"])
    @marketing_required
    def marketing_produk_konfirmasi():
        token = request.form.get("token", "")
        path_tmp = os.path.join(app.config["TMP_IKLAN_FOLDER"], f"produk_{token}.json")
        if not os.path.isfile(path_tmp):
            flash("Sesi upload sudah kedaluwarsa, silakan upload ulang file.", "danger")
            return redirect(url_for("marketing_produk_upload"))

        with open(path_tmp, "r", encoding="utf-8") as f:
            data_tmp = json.load(f)

        mapping = {}
        for key, label, wajib, _kk in KOLOM_TARGET_PRODUK:
            nilai = request.form.get(f"map_{key}", "")
            mapping[key] = int(nilai) if nilai != "" else None
            if wajib and mapping[key] is None:
                flash(f"Kolom '{label}' wajib dipilih.", "danger")
                os.remove(path_tmp)
                return redirect(url_for("marketing_produk_upload"))

        agregat, dilewati, contoh_gagal = proses_baris_upload(data_tmp["rows"], mapping, butuh_produk=True)

        if not agregat:
            os.remove(path_tmp)
            pesan = "Tidak ada baris data yang berhasil diproses — kolom Nama Produk kosong, atau kolom Tanggal bukan tanggal yang valid."
            if contoh_gagal:
                contoh = ", ".join(f"'{c}'" for c in contoh_gagal)
                pesan += f" Contoh nilai di kolom Tanggal: {contoh}."
            flash(pesan, "danger")
            return redirect(url_for("marketing_produk_upload"))

        marketplace = data_tmp["marketplace"]
        sumber_file = data_tmp.get("sumber_file", "")
        for (nama_produk, tanggal), nilai in agregat.items():
            existing = ProdukIklan.query.filter_by(
                marketplace=marketplace, nama_produk=nama_produk, tanggal=tanggal
            ).first()
            if not existing:
                existing = ProdukIklan(marketplace=marketplace, nama_produk=nama_produk, tanggal=tanggal)
                db.session.add(existing)
            existing.biaya = round(nilai["biaya"])
            existing.impresi = round(nilai["impresi"])
            existing.klik = round(nilai["klik"])
            existing.pesanan = round(nilai["pesanan"])
            existing.omzet = round(nilai["omzet"])
            existing.sumber_file = sumber_file
            existing.dibuat_pada = now_wib()
        db.session.commit()
        os.remove(path_tmp)

        jumlah_produk = len(set(k[0] for k in agregat.keys()))
        pesan = f"Berhasil impor {len(agregat)} baris data ({jumlah_produk} produk) untuk {marketplace}."
        if dilewati:
            pesan += f" {dilewati} baris dilewati karena nama produk/tanggal tidak terbaca."
            if contoh_gagal:
                pesan += " Contoh: " + ", ".join(f"'{c}'" for c in contoh_gagal) + "."
        flash(pesan, "success")

        return redirect(url_for("marketing_produk_dashboard", marketplace=marketplace))

    @app.route("/marketing/produk")
    @marketing_required
    def marketing_produk_dashboard():
        hari_ini = today_wib()
        marketplace_filter = request.args.get("marketplace", "Semua")
        try:
            periode_hari = max(int(request.args.get("periode", 30)), 1)
        except ValueError:
            periode_hari = 30
        try:
            threshold_ctr = float(request.args.get("ctr_min", 1))
        except ValueError:
            threshold_ctr = 1.0
        try:
            threshold_roas = float(request.args.get("roas_min", 3))
        except ValueError:
            threshold_roas = 3.0
        try:
            threshold_pesanan14 = int(request.args.get("pesanan14_min", 3))
        except ValueError:
            threshold_pesanan14 = 3

        dari_periode = hari_ini - timedelta(days=periode_hari - 1)
        dari_14 = hari_ini - timedelta(days=13)

        query_periode = ProdukIklan.query.filter(
            ProdukIklan.tanggal >= dari_periode, ProdukIklan.tanggal <= hari_ini
        )
        query_14 = ProdukIklan.query.filter(
            ProdukIklan.tanggal >= dari_14, ProdukIklan.tanggal <= hari_ini
        )
        if marketplace_filter != "Semua":
            query_periode = query_periode.filter(ProdukIklan.marketplace == marketplace_filter)
            query_14 = query_14.filter(ProdukIklan.marketplace == marketplace_filter)

        pesanan14_map = {}
        for d in query_14.all():
            kunci = (d.marketplace, d.nama_produk)
            pesanan14_map[kunci] = pesanan14_map.get(kunci, 0) + (d.pesanan or 0)

        produk_map = {}
        for d in query_periode.all():
            kunci = (d.marketplace, d.nama_produk)
            if kunci not in produk_map:
                produk_map[kunci] = {
                    "marketplace": d.marketplace, "nama_produk": d.nama_produk,
                    "biaya": 0, "impresi": 0, "klik": 0, "pesanan": 0, "omzet": 0,
                }
            produk_map[kunci]["biaya"] += d.biaya or 0
            produk_map[kunci]["impresi"] += d.impresi or 0
            produk_map[kunci]["klik"] += d.klik or 0
            produk_map[kunci]["pesanan"] += d.pesanan or 0
            produk_map[kunci]["omzet"] += d.omzet or 0

        tidak_perform = []
        perform = []
        for kunci, p in produk_map.items():
            if p["biaya"] <= 0:
                continue
            ctr = (p["klik"] / p["impresi"] * 100) if p["impresi"] else 0
            roas = (p["omzet"] / p["biaya"]) if p["biaya"] else 0
            pesanan14 = pesanan14_map.get(kunci, 0)

            alasan = []
            rekomendasi = []
            if p["impresi"] > 0 and ctr < threshold_ctr:
                alasan.append(f"CTR rendah ({ctr:.2f}% < {threshold_ctr:.2f}%)")
                rekomendasi.append("Ganti foto utama & judul produk, uji materi iklan baru agar lebih menarik diklik.")
            if roas < threshold_roas:
                alasan.append(f"ROAS rendah ({roas:.2f}x < {threshold_roas:.2f}x)")
                rekomendasi.append("Klik sudah ada tapi konversi kurang — cek harga, deskripsi, ulasan & foto produk, atau turunkan bid iklan.")
            if pesanan14 < threshold_pesanan14:
                alasan.append(f"Penjualan 14 hari terakhir rendah ({pesanan14} < {threshold_pesanan14})")
                rekomendasi.append("Penjualan masih minim — naikkan budget bertahap sambil dipantau, atau alihkan budget bila 1-2 minggu tidak membaik.")

            item = {**p, "ctr": ctr, "roas": roas, "pesanan14": pesanan14, "alasan": alasan, "rekomendasi": rekomendasi}
            if alasan:
                tidak_perform.append(item)
            else:
                if roas >= threshold_roas * 1.5:
                    item["rekomendasi"] = ["Performa baik, pertimbangkan naikkan budget iklan bertahap (+20-30%) untuk maksimalkan momentum penjualan."]
                else:
                    item["rekomendasi"] = ["Performa sudah sesuai target, pertahankan strategi saat ini."]
                perform.append(item)

        tidak_perform.sort(key=lambda x: x["roas"])
        perform.sort(key=lambda x: -x["roas"])

        return render_template(
            "marketing/produk_dashboard.html",
            marketplace_list=MARKETPLACE_LIST,
            marketplace_filter=marketplace_filter,
            periode_hari=periode_hari,
            threshold_ctr=threshold_ctr,
            threshold_roas=threshold_roas,
            threshold_pesanan14=threshold_pesanan14,
            tidak_perform=tidak_perform,
            perform=perform,
            total_produk=len(tidak_perform) + len(perform),
        )

    @app.route("/marketing/proyeksi")
    @marketing_required
    def marketing_proyeksi():
        hari_ini = today_wib()
        periode_hari = 30
        dari = hari_ini - timedelta(days=periode_hari - 1)
        data = IklanMarketplace.query.filter(
            IklanMarketplace.tanggal >= dari, IklanMarketplace.tanggal <= hari_ini
        ).all()

        bulan_depan = hari_ini.month + 1
        tahun_depan = hari_ini.year
        if bulan_depan > 12:
            bulan_depan = 1
            tahun_depan += 1
        jumlah_hari_bulan_depan = calendar.monthrange(tahun_depan, bulan_depan)[1]

        proyeksi = []
        for mp in MARKETPLACE_LIST:
            item = [d for d in data if d.marketplace == mp]
            if not item:
                continue
            jumlah_hari_data = len(set(d.tanggal for d in item))
            total_biaya = sum(d.biaya or 0 for d in item)
            total_omzet = sum(d.omzet or 0 for d in item)
            avg_biaya_harian = (total_biaya / jumlah_hari_data) if jumlah_hari_data else 0
            roas_historis = (total_omzet / total_biaya) if total_biaya else 0
            proyeksi_budget = avg_biaya_harian * jumlah_hari_bulan_depan
            proyeksi.append({
                "marketplace": mp,
                "avg_biaya_harian": avg_biaya_harian,
                "roas_historis": roas_historis,
                "proyeksi_budget": proyeksi_budget,
                "proyeksi_omzet": proyeksi_budget * roas_historis,
            })

        target_marketplace = request.args.get("target_marketplace", "")
        try:
            target_omzet = float(request.args.get("target_omzet", "") or 0)
        except ValueError:
            target_omzet = 0

        hasil_target = None
        if target_marketplace and target_omzet > 0:
            if target_marketplace == "Semua":
                item = data
            else:
                item = [d for d in data if d.marketplace == target_marketplace]
            total_biaya_t = sum(d.biaya or 0 for d in item)
            total_omzet_t = sum(d.omzet or 0 for d in item)
            roas_pakai = (total_omzet_t / total_biaya_t) if total_biaya_t else 0
            if roas_pakai > 0:
                hasil_target = {
                    "marketplace": target_marketplace,
                    "target_omzet": target_omzet,
                    "roas_pakai": roas_pakai,
                    "budget_dibutuhkan": target_omzet / roas_pakai,
                }
            else:
                hasil_target = {"marketplace": target_marketplace, "error": True}

        return render_template(
            "marketing/proyeksi.html",
            marketplace_list=MARKETPLACE_LIST,
            proyeksi=proyeksi,
            periode_hari=periode_hari,
            bulan_depan_label=f"{BULAN_NAMA[bulan_depan]} {tahun_depan}",
            target_marketplace=target_marketplace,
            target_omzet=target_omzet,
            hasil_target=hasil_target,
        )

    # ---------- MARKETING: PROFITABILITAS (ORDER + INCOME + HPP) ----------
    def _simpan_order_marketplace(marketplace, file_storage, headers, rows_data):
        nama_file_aman = secure_filename(file_storage.filename)
        waktu_impor = now_wib()
        item_list = PARSER_ORDER_MARKETPLACE[marketplace](headers, rows_data)
        no_pesanan_set = {item["no_pesanan"] for item in item_list}
        peta_existing = {
            (p.no_pesanan, p.sku, p.nama_produk): p
            for p in PesananMarketplace.query.filter(
                PesananMarketplace.marketplace == marketplace, PesananMarketplace.no_pesanan.in_(no_pesanan_set),
            ).all()
        } if no_pesanan_set else {}
        for item in item_list:
            kunci = (item["no_pesanan"], item["sku"], item["nama_produk"])
            existing = peta_existing.get(kunci)
            if not existing:
                existing = PesananMarketplace(
                    marketplace=marketplace, no_pesanan=item["no_pesanan"],
                    sku=item["sku"], nama_produk=item["nama_produk"],
                )
                db.session.add(existing)
                peta_existing[kunci] = existing
            existing.tanggal_pesanan = item["tanggal_pesanan"]
            existing.status_pesanan = item["status_pesanan"]
            existing.jumlah = item["jumlah"]
            existing.subtotal = item["subtotal"]
            existing.sumber_file = nama_file_aman
            existing.dibuat_pada = waktu_impor
        return item_list

    def _simpan_income_marketplace(marketplace, file_storage, headers, rows_data):
        nama_file_aman = secure_filename(file_storage.filename)
        waktu_impor = now_wib()
        item_list = PARSER_INCOME_MARKETPLACE[marketplace](headers, rows_data)
        no_pesanan_set = {item["no_pesanan"] for item in item_list}
        peta_existing = {
            p.no_pesanan: p
            for p in PendapatanPesanan.query.filter(
                PendapatanPesanan.marketplace == marketplace, PendapatanPesanan.no_pesanan.in_(no_pesanan_set),
            ).all()
        } if no_pesanan_set else {}
        for item in item_list:
            existing = peta_existing.get(item["no_pesanan"])
            if not existing:
                existing = PendapatanPesanan(marketplace=marketplace, no_pesanan=item["no_pesanan"])
                db.session.add(existing)
                peta_existing[item["no_pesanan"]] = existing
            existing.tanggal_dana_dilepas = item["tanggal_dana_dilepas"]
            existing.total_penghasilan = item["total_penghasilan"]
            existing.biaya_admin = item["biaya_admin"]
            existing.biaya_layanan = item["biaya_layanan"]
            existing.biaya_lainnya = item["biaya_lainnya"]
            existing.sumber_file = nama_file_aman
            existing.dibuat_pada = waktu_impor
        return item_list

    def _buat_income_otomatis_manual(order_item_list, nama_file_aman):
        """Pesanan Manual (WA/COD/langsung, di luar marketplace) tidak punya file Income
        terpisah karena memang tidak ada potongan biaya admin/layanan platform -- jadi
        begitu Order Manual diupload, data Income-nya langsung dibuat otomatis di sini
        (total_penghasilan = subtotal, semua biaya 0) supaya Order & Income dan Dashboard
        Profit langsung nyambung tanpa perlu upload apa pun lagi. Pesanan yang statusnya
        Batal/Dibatalkan dilewati, sama seperti pesanan marketplace lain yang batal
        memang tidak pernah ada dana yang cair."""
        waktu_impor = now_wib()
        subtotal_per_pesanan = {}
        tanggal_per_pesanan = {}
        for item in order_item_list:
            if item["status_pesanan"] in STATUS_BATAL_MARKETPLACE:
                continue
            no = item["no_pesanan"]
            subtotal_per_pesanan[no] = subtotal_per_pesanan.get(no, 0) + item["subtotal"]
            tanggal_per_pesanan.setdefault(no, item["tanggal_pesanan"])

        income_item_list = []
        if subtotal_per_pesanan:
            peta_existing = {
                p.no_pesanan: p
                for p in PendapatanPesanan.query.filter(
                    PendapatanPesanan.marketplace == "Manual",
                    PendapatanPesanan.no_pesanan.in_(subtotal_per_pesanan.keys()),
                ).all()
            }
            for no, total in subtotal_per_pesanan.items():
                existing = peta_existing.get(no)
                if not existing:
                    existing = PendapatanPesanan(marketplace="Manual", no_pesanan=no)
                    db.session.add(existing)
                    peta_existing[no] = existing
                existing.tanggal_dana_dilepas = tanggal_per_pesanan[no]
                existing.total_penghasilan = total
                existing.biaya_admin = 0
                existing.biaya_layanan = 0
                existing.biaya_lainnya = 0
                existing.sumber_file = nama_file_aman
                existing.dibuat_pada = waktu_impor
                income_item_list.append({
                    "no_pesanan": no, "tanggal_dana_dilepas": tanggal_per_pesanan[no],
                    "total_penghasilan": total, "biaya_admin": 0, "biaya_layanan": 0, "biaya_lainnya": 0,
                })
        return income_item_list

    @app.route("/marketing/profit/upload", methods=["GET", "POST"])
    @marketing_required
    def profit_upload():
        hasil = {
            "order": None, "income": None, "iklan": None, "order_manual": None,
            "ringkasan_resmi": None, "ringkasan_gabungan": None,
        }

        if request.method == "POST":
            file_order = request.files.get("file_order")
            file_income = request.files.get("file_income")
            file_iklan = request.files.get("file_iklan")
            file_order_manual = request.files.get("file_order_manual")

            if (
                not (file_order and file_order.filename) and not (file_income and file_income.filename)
                and not (file_iklan and file_iklan.filename)
                and not (file_order_manual and file_order_manual.filename)
            ):
                flash("Pilih minimal satu file terlebih dahulu.", "danger")
                return redirect(url_for("profit_upload"))

            marketplace_terdeteksi = None
            order_item_list = None
            income_item_list = None

            if file_order and file_order.filename:
                marketplace, tipe, headers, rows_data, error = baca_laporan_marketplace(file_order)
                if error or tipe != "order":
                    hasil["order"] = {"ok": False, "nama": file_order.filename, "pesan": error or (
                        f"File ini terbaca sebagai laporan {tipe or 'tidak dikenali'}, bukan Laporan Order. "
                        "Coba cek lagi filenya."
                    )}
                else:
                    order_item_list = _simpan_order_marketplace(marketplace, file_order, headers, rows_data)
                    marketplace_terdeteksi = marketplace
                    hasil["order"] = {
                        "ok": True, "nama": file_order.filename, "jumlah": len(order_item_list), "marketplace": marketplace,
                        "preview_items": order_item_list[:8],
                    }

            if file_income and file_income.filename:
                marketplace, tipe, headers, rows_data, error = baca_laporan_marketplace(file_income)
                if error or tipe != "income":
                    hasil["income"] = {"ok": False, "nama": file_income.filename, "pesan": error or (
                        f"File ini terbaca sebagai laporan {tipe or 'tidak dikenali'}, bukan Laporan Income. "
                        "Coba cek lagi filenya."
                    )}
                else:
                    income_item_list = _simpan_income_marketplace(marketplace, file_income, headers, rows_data)
                    marketplace_terdeteksi = marketplace_terdeteksi or marketplace
                    hasil["income"] = {
                        "ok": True, "nama": file_income.filename, "jumlah": len(income_item_list), "marketplace": marketplace,
                        "preview_items": income_item_list[:8],
                    }
                    if marketplace == "Shopee":
                        file_income.seek(0)
                        hasil["ringkasan_resmi"] = baca_ringkasan_summary_shopee(file_income)

            if (
                order_item_list is not None and income_item_list is not None
                and hasil["order"]["marketplace"] == hasil["income"]["marketplace"]
            ):
                hasil["ringkasan_gabungan"] = {
                    "marketplace": hasil["order"]["marketplace"],
                    **hitung_ringkasan_gabungan(order_item_list, income_item_list),
                }

            if file_order_manual and file_order_manual.filename:
                headers_m, rows_m, _idx_header_m, error_m = baca_file_iklan(file_order_manual)
                if error_m:
                    hasil["order_manual"] = {"ok": False, "nama": file_order_manual.filename, "pesan": error_m}
                else:
                    order_manual_item_list = _simpan_order_marketplace("Manual", file_order_manual, headers_m, rows_m)
                    if not order_manual_item_list:
                        hasil["order_manual"] = {
                            "ok": False, "nama": file_order_manual.filename,
                            "pesan": "Tidak ada baris pesanan yang terbaca dari file ini. Pastikan kolom 'No. Pesanan' "
                                     "dan 'Waktu Pesanan Dibuat' terisi, sesuai template.",
                        }
                    else:
                        nama_file_manual_aman = secure_filename(file_order_manual.filename)
                        income_manual_item_list = _buat_income_otomatis_manual(order_manual_item_list, nama_file_manual_aman)
                        hasil["order_manual"] = {
                            "ok": True, "nama": file_order_manual.filename, "jumlah": len(order_manual_item_list),
                            "jumlah_income_otomatis": len(income_manual_item_list),
                            "preview_items": order_manual_item_list[:8],
                        }

            if file_iklan and file_iklan.filename:
                headers_i, rows_i, _idx_header, error_i = baca_file_iklan(file_iklan)
                if error_i:
                    hasil["iklan"] = {"ok": False, "nama": file_iklan.filename, "pesan": error_i}
                else:
                    mapping_i = deteksi_otomatis_kolom_iklan(headers_i, rows_i, KOLOM_TARGET_IKLAN)
                    if mapping_i.get("tanggal") is None or mapping_i.get("biaya") is None:
                        hasil["iklan"] = {
                            "ok": False, "nama": file_iklan.filename,
                            "pesan": "Kolom Tanggal/Biaya Iklan tidak berhasil dikenali otomatis dari file ini.",
                        }
                    else:
                        marketplace_iklan = marketplace_terdeteksi or "Shopee"
                        agregat_i, _dilewati_i, _contoh_i = proses_baris_upload(rows_i, mapping_i, butuh_produk=False)
                        waktu_impor = now_wib()
                        nama_file_iklan_aman = secure_filename(file_iklan.filename)
                        for tanggal, nilai in agregat_i.items():
                            existing = IklanMarketplace.query.filter_by(marketplace=marketplace_iklan, tanggal=tanggal).first()
                            if not existing:
                                existing = IklanMarketplace(marketplace=marketplace_iklan, tanggal=tanggal)
                                db.session.add(existing)
                            existing.biaya = round(nilai["biaya"])
                            existing.impresi = round(nilai["impresi"])
                            existing.klik = round(nilai["klik"])
                            existing.pesanan = round(nilai["pesanan"])
                            existing.omzet = round(nilai["omzet"])
                            existing.sumber_file = nama_file_iklan_aman
                            existing.dibuat_pada = waktu_impor
                        iklan_preview_items = sorted(
                            ({"tanggal": t, **v} for t, v in agregat_i.items()),
                            key=lambda x: x["tanggal"], reverse=True,
                        )[:8]
                        hasil["iklan"] = {
                            "ok": True, "nama": file_iklan.filename, "jumlah": len(agregat_i),
                            "marketplace": marketplace_iklan, "preview_items": iklan_preview_items,
                        }

            db.session.commit()

            upload_sukses = (
                (hasil["order"] and hasil["order"]["ok"])
                or (hasil["income"] and hasil["income"]["ok"])
                or (hasil["iklan"] and hasil["iklan"]["ok"])
                or (hasil["order_manual"] and hasil["order_manual"]["ok"])
            )
            jumlah_produk_belum_hpp = None
            if upload_sukses:
                jumlah_produk_belum_hpp = sum(1 for p in _daftar_produk_untuk_hpp() if not p["hpp"])
                flash("Data berhasil diimpor. Lihat pratinjau di bawah, atau lanjut ke langkah berikutnya.", "success")
            else:
                flash("Upload gagal diproses, lihat detail di bawah tiap file.", "danger")

            return render_template(
                "marketing/profit_upload.html", aktif="upload", hasil=hasil,
                jumlah_produk_belum_hpp=jumlah_produk_belum_hpp,
            )

        return render_template("marketing/profit_upload.html", aktif="upload", hasil=hasil, jumlah_produk_belum_hpp=None)

    @app.route("/marketing/profit/order-manual/template")
    @marketing_required
    def profit_order_manual_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Order Manual"
        header = [
            "No. Pesanan", "Status Pesanan", "No. Resi", "Opsi Pengiriman", "Antar ke counter/ pick-up",
            "Waktu Pengiriman Diatur", "Waktu Pesanan Dibuat", "Waktu Pembayaran Dilakukan", "Metode Pembayaran",
            "SKU Induk", "Nama Produk", "Nomor Referensi SKU", "Nama Variasi", "Harga Awal",
            "Harga Setelah Diskon", "Jumlah", "Subtotal Pesanan", "Total Diskon", "Diskon Dari Penjual",
            "Berat Produk", "Jumlah Produk di Pesan", "Total Berat", "Total Pembayaran", "Perkiraan Ongkos Kirim",
            "Username (Pembeli)", "Nama Penerima", "No. Telepon", "Alamat Pengiriman", "Kota/Kabupaten",
            "Provinsi", "Waktu Pesanan Selesai",
        ]
        ws.append(header)
        contoh = [
            "WA-01-08-2026", "Selesai", "", "COD/Diantar Langsung", "", "", "2026-08-01 10:00", "", "Transfer",
            "", "Nama Produk Contoh", "SKU-001", "Warna/Ukuran", 100000, 90000, 1, 90000, 10000, 10000,
            "", "1", "", 90000, "", "", "Nama Pembeli", "", "", "", "", "2026-08-01 10:00",
        ]
        ws.append(contoh)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, as_attachment=True, download_name="template_order_manual.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/marketing/profit/data")
    @marketing_required
    def profit_order_income():
        cari = request.args.get("cari", "").strip()
        bulan_filter = request.args.get("bulan", "")
        marketplace_filter = request.args.get("marketplace", "")
        status_filter = request.args.get("status", "")
        income_status_filter = request.args.get("income_status", "")  # "" / cocok / belum

        total_pesanan = PesananMarketplace.query.count()
        total_income = PendapatanPesanan.query.count()
        kunci_pesanan = {(p.marketplace, p.no_pesanan) for p in PesananMarketplace.query.with_entities(PesananMarketplace.marketplace, PesananMarketplace.no_pesanan).distinct()}
        kunci_income = {(p.marketplace, p.no_pesanan) for p in PendapatanPesanan.query.with_entities(PendapatanPesanan.marketplace, PendapatanPesanan.no_pesanan).all()}
        jumlah_sudah_cocok = len(kunci_pesanan & kunci_income)
        jumlah_belum_ada_income = len(kunci_pesanan - kunci_income)

        daftar_marketplace = sorted({m for m, _ in kunci_pesanan} | {m for m, _ in kunci_income})
        daftar_status = sorted({
            r[0] for r in PesananMarketplace.query.with_entities(PesananMarketplace.status_pesanan).distinct().all() if r[0]
        })

        subq_ada_income = db.exists().where(db.and_(
            PendapatanPesanan.marketplace == PesananMarketplace.marketplace,
            PendapatanPesanan.no_pesanan == PesananMarketplace.no_pesanan,
        ))

        def terapkan_filter_pesanan(query):
            if cari:
                like = f"%{cari}%"
                query = query.filter(db.or_(PesananMarketplace.no_pesanan.ilike(like), PesananMarketplace.nama_produk.ilike(like)))
            if marketplace_filter:
                query = query.filter(PesananMarketplace.marketplace == marketplace_filter)
            if status_filter:
                query = query.filter(PesananMarketplace.status_pesanan == status_filter)
            if bulan_filter:
                try:
                    tahun_f, bulan_f = (int(x) for x in bulan_filter.split("-"))
                    query = query.filter(
                        db.extract("year", PesananMarketplace.tanggal_pesanan) == tahun_f,
                        db.extract("month", PesananMarketplace.tanggal_pesanan) == bulan_f,
                    )
                except ValueError:
                    pass
            if income_status_filter == "belum":
                query = query.filter(~subq_ada_income)
            elif income_status_filter == "cocok":
                query = query.filter(subq_ada_income)
            return query

        item_pesanan = (
            terapkan_filter_pesanan(PesananMarketplace.query)
            .order_by(PesananMarketplace.tanggal_pesanan.desc()).limit(500).all()
        )

        kunci_terlihat = {(p.marketplace, p.no_pesanan) for p in item_pesanan}
        income_map = {
            (p.marketplace, p.no_pesanan): p
            for p in PendapatanPesanan.query.filter(
                db.tuple_(PendapatanPesanan.marketplace, PendapatanPesanan.no_pesanan).in_(kunci_terlihat)
            ).all()
        } if kunci_terlihat else {}

        # Ringkasan omset dari SELURUH baris yang cocok filter di atas (bukan cuma 500
        # baris yang ditampilkan) -- omset bersih mengecualikan pesanan Batal, dan nilai
        # pesanan Batal dilaporkan terpisah, dipecah per marketplace.
        omzet_per_marketplace = {}
        baris_untuk_omzet = terapkan_filter_pesanan(PesananMarketplace.query).with_entities(
            PesananMarketplace.marketplace, PesananMarketplace.status_pesanan, PesananMarketplace.subtotal,
        ).all()
        for mp, status, subtotal in baris_untuk_omzet:
            d = omzet_per_marketplace.setdefault(mp, {
                "marketplace": mp, "omzet_bersih": 0, "jumlah_pesanan": 0,
                "nilai_batal": 0, "jumlah_batal": 0,
            })
            d["jumlah_pesanan"] += 1
            if status in STATUS_BATAL_MARKETPLACE:
                d["nilai_batal"] += subtotal or 0
                d["jumlah_batal"] += 1
            else:
                d["omzet_bersih"] += subtotal or 0
        omzet_list = sorted(omzet_per_marketplace.values(), key=lambda x: -x["omzet_bersih"])
        total_omzet_bersih = sum(v["omzet_bersih"] for v in omzet_list)
        total_nilai_batal = sum(v["nilai_batal"] for v in omzet_list)
        total_jumlah_batal = sum(v["jumlah_batal"] for v in omzet_list)

        qi = PendapatanPesanan.query
        if marketplace_filter:
            qi = qi.filter(PendapatanPesanan.marketplace == marketplace_filter)
        if cari:
            qi = qi.filter(PendapatanPesanan.no_pesanan.ilike(f"%{cari}%"))
        if bulan_filter:
            try:
                tahun_f, bulan_f = (int(x) for x in bulan_filter.split("-"))
                qi = qi.filter(
                    db.extract("year", PendapatanPesanan.tanggal_dana_dilepas) == tahun_f,
                    db.extract("month", PendapatanPesanan.tanggal_dana_dilepas) == bulan_f,
                )
            except ValueError:
                pass
        item_income = qi.order_by(PendapatanPesanan.tanggal_dana_dilepas.desc()).limit(500).all()

        return render_template(
            "marketing/profit_data.html",
            aktif="data",
            item_pesanan=item_pesanan,
            item_income=item_income,
            income_map=income_map,
            cari=cari,
            bulan_filter=bulan_filter,
            marketplace_filter=marketplace_filter,
            daftar_marketplace=daftar_marketplace,
            total_pesanan=total_pesanan,
            total_income=total_income,
            jumlah_order_unik=len(kunci_pesanan),
            jumlah_sudah_cocok=jumlah_sudah_cocok,
            jumlah_belum_ada_income=jumlah_belum_ada_income,
            status_filter=status_filter,
            income_status_filter=income_status_filter,
            daftar_status=daftar_status,
            omzet_list=omzet_list,
            total_omzet_bersih=total_omzet_bersih,
            total_nilai_batal=total_nilai_batal,
            total_jumlah_batal=total_jumlah_batal,
        )

    def _daftar_produk_untuk_hpp():
        nama_produk_list = [
            r[0] for r in PesananMarketplace.query.with_entities(PesananMarketplace.nama_produk).distinct().all()
        ]
        qty_map = {}
        for it in PesananMarketplace.query.all():
            qty_map[it.nama_produk] = qty_map.get(it.nama_produk, 0) + it.jumlah
        produk_map = {p.nama_produk: p for p in Produk.query.filter(Produk.nama_produk.in_(nama_produk_list)).all()}
        daftar = []
        for nama in sorted(nama_produk_list):
            p = produk_map.get(nama)
            daftar.append({"nama_produk": nama, "hpp": (p.hpp if p else 0), "terjual": qty_map.get(nama, 0)})
        daftar.sort(key=lambda x: (x["hpp"] > 0, -x["terjual"]))
        return daftar

    @app.route("/marketing/profit/hpp", methods=["GET", "POST"])
    @marketing_required
    def profit_hpp():
        if request.method == "POST":
            nama_list = request.form.getlist("nama_produk")
            hpp_list = request.form.getlist("hpp")
            jumlah_disimpan = 0
            for nama, hpp_raw in zip(nama_list, hpp_list):
                if not nama:
                    continue
                hpp_val = round(parse_angka_iklan(hpp_raw))
                produk = Produk.query.filter_by(nama_produk=nama).first()
                if not produk:
                    produk = Produk(nama_produk=nama)
                    db.session.add(produk)
                produk.hpp = hpp_val
                produk.modal = hpp_val
                jumlah_disimpan += 1
            db.session.commit()
            flash(f"HPP {jumlah_disimpan} produk berhasil disimpan.", "success")
            return redirect(url_for("profit_hpp"))

        return render_template("marketing/profit_hpp.html", aktif="hpp", daftar=_daftar_produk_untuk_hpp())

    @app.route("/marketing/profit/hpp/template")
    @marketing_required
    def profit_hpp_template():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template HPP"
        ws.append(["Nama Produk", "HPP"])
        for item in _daftar_produk_untuk_hpp():
            ws.append([item["nama_produk"], item["hpp"] or ""])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, as_attachment=True, download_name="template_hpp.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/marketing/profit/hpp/upload", methods=["POST"])
    @marketing_required
    def profit_hpp_upload():
        file = request.files.get("file")
        if not file or not file.filename:
            flash("Pilih file template HPP yang sudah diisi terlebih dahulu.", "danger")
            return redirect(url_for("profit_hpp"))

        headers, rows, _idx_header, error = baca_file_iklan(file)
        if error:
            flash(error, "danger")
            return redirect(url_for("profit_hpp"))

        headers_lower = [str(h).strip().lower() for h in headers]
        try:
            idx_nama = headers_lower.index("nama produk")
        except ValueError:
            idx_nama = 0
        idx_hpp = None
        for i, h in enumerate(headers_lower):
            if "hpp" in h:
                idx_hpp = i
                break
        if idx_hpp is None:
            idx_hpp = 1 if len(headers) > 1 else None

        if idx_hpp is None:
            flash("Kolom Nama Produk/HPP tidak ditemukan di file ini. Gunakan template yang sudah didownload.", "danger")
            return redirect(url_for("profit_hpp"))

        jumlah = 0
        for row in rows:
            nama = str(row[idx_nama]).strip() if idx_nama < len(row) and row[idx_nama] is not None else ""
            if not nama:
                continue
            hpp_raw = row[idx_hpp] if idx_hpp < len(row) else None
            if hpp_raw in (None, ""):
                continue
            hpp_val = round(parse_angka_iklan(hpp_raw))
            produk = Produk.query.filter_by(nama_produk=nama).first()
            if not produk:
                produk = Produk(nama_produk=nama)
                db.session.add(produk)
            produk.hpp = hpp_val
            produk.modal = hpp_val
            jumlah += 1
        db.session.commit()
        flash(f"HPP {jumlah} produk berhasil diperbarui dari file.", "success")
        return redirect(url_for("profit_hpp"))

    @app.route("/marketing/profit/dashboard")
    @marketing_required
    def profit_dashboard():
        bulan_filter = request.args.get("bulan", "")
        data = hitung_profit_agregat(bulan_filter or None)
        data["produk"].sort(key=lambda p: p["profit"])

        bulan_tersedia = sorted({
            r[0].strftime("%Y-%m")
            for r in PesananMarketplace.query.with_entities(PesananMarketplace.tanggal_pesanan).all()
        }, reverse=True)

        bulan_a = request.args.get("bulan_a", "")
        bulan_b = request.args.get("bulan_b", "")
        perbandingan = None
        if bulan_a and bulan_b:
            perbandingan = {
                "a": {"label": bulan_a, **hitung_profit_agregat(bulan_a)["ringkasan"]},
                "b": {"label": bulan_b, **hitung_profit_agregat(bulan_b)["ringkasan"]},
            }

        # ---- Health Score, Diagnosa Singkat, Sorotan Produk (buat panel ringkasan
        # "sekali lihat langsung ngerti" di atas halaman) ----
        r = data["ringkasan"]
        hpp_lengkap = r["jumlah_produk"] > 0 and r["jumlah_produk_ada_hpp"] == r["jumlah_produk"]
        ad_ratio = (r["total_biaya_iklan"] / r["total_omzet"] * 100) if r["total_omzet"] else 0

        health_score = None
        status_profit = None
        if hpp_lengkap:
            margin_score = max(0, min(100, r["margin_real"] / 40 * 100))
            ad_penalty = max(0, min(30, ad_ratio))
            health_score = round(max(0, min(100, margin_score - ad_penalty)))
            if r["total_profit_real"] < 0:
                status_profit = "rugi"
            elif r["margin_real"] < 10:
                status_profit = "tipis"
            else:
                status_profit = "sehat"

        diagnosa = []
        if ad_ratio > 10:
            diagnosa.append({"level": "danger", "pesan": f"Rasio biaya iklan {ad_ratio:.0f}% terlalu tinggi terhadap omzet"})
        if not hpp_lengkap and r["jumlah_produk"] > 0:
            kurang = r["jumlah_produk"] - r["jumlah_produk_ada_hpp"]
            diagnosa.append({"level": "warning", "pesan": f"{kurang} produk belum ada HPP — profitnya belum ikut terhitung"})
        if hpp_lengkap and status_profit == "rugi":
            diagnosa.append({"level": "danger", "pesan": "Profit real minus bulan ini — segera evaluasi harga jual & biaya"})
        elif hpp_lengkap and status_profit == "tipis":
            diagnosa.append({"level": "warning", "pesan": f"Margin real cuma {r['margin_real']:.1f}% dari sales — tergolong tipis"})
        if not diagnosa and r["jumlah_produk"] > 0:
            diagnosa.append({"level": "success", "pesan": "Tidak ada masalah besar terdeteksi pada periode ini"})

        produk_ada_income = [p for p in data["produk"] if p["ada_income"]]
        produk_ada_hpp = [p for p in produk_ada_income if p["ada_hpp"]]
        terlaris = max(produk_ada_income, key=lambda p: p["qty"]) if produk_ada_income else None
        paling_profit = max(produk_ada_hpp, key=lambda p: p["profit"]) if produk_ada_hpp else None
        paling_rugi = min(produk_ada_hpp, key=lambda p: p["profit"]) if produk_ada_hpp else None
        if paling_rugi and paling_profit and paling_rugi["nama_produk"] == paling_profit["nama_produk"]:
            paling_rugi = None
        sorotan_produk = [
            {"label": "Produk Paling Profit", "item": paling_profit, "badge": "Paling Untung", "badge_class": "success"},
            {"label": "Produk Paling Rugi", "item": paling_rugi, "badge": "Perlu Dicek", "badge_class": "danger"},
            {"label": "Produk Terlaris", "item": terlaris, "badge": "Terlaris", "badge_class": "primary"},
        ]

        # Breakdown "Profit Kamu Bocor di Sini": porsi tiap komponen biaya terhadap
        # Total Sales, sisanya (sebelum HPP) dihitung sebagai residual supaya batangnya
        # selalu pas 100% biarpun ada pembulatan/selisih kecil antar kategori.
        omzet_basis = r["total_omzet"] or 1
        komponen_biaya = [
            {"label": "Biaya Iklan", "nilai": r["total_biaya_iklan"], "warna": "#ec4899"},
            {"label": "Administrasi", "nilai": r["total_biaya_admin"], "warna": "#3b82f6"},
            {"label": "Biaya Lainnya", "nilai": r["total_biaya_lainnya"], "warna": "#14b8a6"},
            {"label": "Biaya Layanan", "nilai": r["total_biaya_layanan"], "warna": "#8b5cf6"},
        ]
        for k in komponen_biaya:
            k["persen"] = k["nilai"] / omzet_basis * 100
        sisa_sebelum_hpp = r["total_omzet"] - sum(k["nilai"] for k in komponen_biaya)
        persen_sisa = max(sisa_sebelum_hpp, 0) / omzet_basis * 100

        return render_template(
            "marketing/profit_dashboard.html",
            aktif="dashboard",
            bulan_filter=bulan_filter,
            bulan_tersedia=bulan_tersedia,
            ringkasan=data["ringkasan"],
            produk=data["produk"],
            tren=data["tren"],
            bulan_a=bulan_a,
            bulan_b=bulan_b,
            perbandingan=perbandingan,
            hpp_lengkap=hpp_lengkap,
            ad_ratio=ad_ratio,
            health_score=health_score,
            status_profit=status_profit,
            diagnosa=diagnosa,
            sorotan_produk=sorotan_produk,
            komponen_biaya=komponen_biaya,
            sisa_sebelum_hpp=sisa_sebelum_hpp,
            persen_sisa=persen_sisa,
        )

    @app.route("/marketing/profit/dashboard/export")
    @marketing_required
    def profit_dashboard_export():
        bulan_filter = request.args.get("bulan", "")
        data = hitung_profit_agregat(bulan_filter or None)
        data["produk"].sort(key=lambda p: p["profit"])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profit per Produk"
        ws.append(["Nama Produk", "Qty Terjual", "Omzet", "Total Penghasilan", "HPP Total", "Profit", "Margin (%)", "HPP Terisi?"])
        for p in data["produk"]:
            ws.append([
                p["nama_produk"], p["qty"], p["omzet"], p["income"], p["hpp_total"], p["profit"],
                round(p["margin"], 1), "Ya" if p["ada_hpp"] else "Belum",
            ])
        ws.append([])
        r = data["ringkasan"]
        ws.append(["TOTAL", "", r["total_omzet"], r["total_income"], r["total_hpp"], r["total_profit"], round(r["margin"], 1), ""])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nama_file = f"laporan_profit_{bulan_filter or 'semua'}.xlsx"
        return send_file(
            buf, as_attachment=True, download_name=nama_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/marketing/profit/insight")
    @marketing_required
    def profit_insight():
        bulan_filter = request.args.get("bulan", "")
        data = hitung_profit_agregat(bulan_filter or None)

        insight_list = []
        for p in data["produk"]:
            if not p["ada_hpp"] or not p["ada_income"]:
                continue
            if p["profit"] < 0:
                insight_list.append({
                    "produk": p["nama_produk"], "level": "danger", "margin": p["margin"],
                    "pesan": f"RUGI (profit {p['profit']:,} setelah HPP) — segera evaluasi harga jual atau HPP produk ini, atau hentikan sementara promosinya.".replace(",", "."),
                })
            elif p["margin"] < 10:
                insight_list.append({
                    "produk": p["nama_produk"], "level": "warning", "margin": p["margin"],
                    "pesan": "Margin tipis (di bawah 10%) — pertimbangkan naikkan harga jual, cari HPP lebih murah, atau kurangi diskon/voucher.",
                })
            elif p["margin"] > 40 and p["qty"] >= 5:
                insight_list.append({
                    "produk": p["nama_produk"], "level": "success", "margin": p["margin"],
                    "pesan": "Margin sehat & laku banyak — produk ini layak ditambah budget iklan untuk digenjot lebih.",
                })

        insight_list.sort(key=lambda x: {"danger": 0, "warning": 1, "success": 2}[x["level"]])

        bulan_tersedia = sorted({
            r[0].strftime("%Y-%m")
            for r in PesananMarketplace.query.with_entities(PesananMarketplace.tanggal_pesanan).all()
        }, reverse=True)

        return render_template(
            "marketing/profit_insight.html",
            aktif="insight",
            bulan_filter=bulan_filter,
            bulan_tersedia=bulan_tersedia,
            insight_list=insight_list,
            jumlah_produk_ada_hpp=data["ringkasan"]["jumlah_produk_ada_hpp"],
        )

    @app.route("/akun", methods=["GET", "POST"])
    @admin_required
    def akun():
        if request.method == "POST":
            password_baru = request.form.get("password_baru", "")
            password_konfirmasi = request.form.get("password_konfirmasi", "")
            if len(password_baru) < 6:
                flash("Password baru minimal 6 karakter.", "danger")
            elif password_baru != password_konfirmasi:
                flash("Konfirmasi password tidak cocok.", "danger")
            else:
                current_user.set_password(password_baru)
                db.session.commit()
                flash("Password berhasil diubah.", "success")
            return redirect(url_for("akun"))
        return render_template("account.html")

    def abort_404():
        from flask import abort
        abort(404)

    @app.errorhandler(413)
    def file_terlalu_besar(e):
        flash("File yang diupload terlalu besar (maksimal 10MB). Silakan kompres/perkecil filenya lalu coba lagi.", "danger")
        return redirect(request.referrer or url_for("login"))

    @app.context_processor
    def inject_globals():
        pending_izin = 0
        pending_lembur = 0
        if current_user.is_authenticated and getattr(current_user, "role", None) == "admin":
            pending_izin = PengajuanIzin.query.filter_by(status="Menunggu").count()
            pending_lembur = PengajuanLembur.query.filter_by(status="Menunggu").count()
        return {
            "bulan_nama_list": BULAN_NAMA,
            "pending_izin_count": pending_izin,
            "pending_lembur_count": pending_lembur,
            "site_settings": get_settings(),
        }

    return app


app = create_app()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
